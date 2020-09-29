# -*- coding: utf-8 -*-
import pandas as pd
import os
import numpy as np
import datetime
import re
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook

logtxtFilePath = r'C:/PyCharm\pdf-docx\source\Example\池分布\log.txt'


def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))


def writeErrToSheet(sourceFilePath, ErrorMsg):
    excelwb = load_workbook(sourceFilePath)
    if len(ErrorMsg) > 0:
        logSheet = excelwb.create_sheet("格式检查")
        i = 1
        while i <= len(ErrorMsg):
            logSheet["A{0}".format(i)] = ErrorMsg[i - 1]
            i += 1
    else:
        logSheet = excelwb.create_sheet("格式检查")
        logSheet["A1"] = '检验完成，无错误！！！'

    excelwb.save(sourceFilePath)


def ErrLocation(colName):  # 定位错误的具体列
    if colName == 'PaymentPeriodID':
        return 'A'
    elif colName == 'Amount':
        return 'G'
    elif colName == 'BucketSequenceNo':
        return 'E'
    elif colName == 'AmountPercentage':
        return 'H'
    elif colName == 'CountPercentage':
        return 'J'
    elif colName == 'Count':
        return 'I'


def is_Number(value):
    num = '^[-]?\d*[.]?\d*$'
    flag = re.search(num, value)
    if str(flag) == 'None':
        return False
    else:
        return True


def CheckData(sourceFilePath):
    ErrorMsg = []  # 错误信息记录
    file = pd.read_excel(sourceFilePath)
    splitname = sourceFilePath.split('_')[2]

    # PaymentPeriodID列检测
    try:
        if 'PaymentPeriodID' not in file.columns:
            ErrorMsg.append("A1:0:PaymentPeriodID字段缺失，请勿更改模板列标题")
        if 'Amount' not in file.columns:
            ErrorMsg.append("G1:0:Amount字段缺失，请勿更改模板列标题")
        if 'BucketSequenceNo' not in file.columns:
            ErrorMsg.append("E1:0:BucketSequenceNo字段缺失，请勿更改模板列标题")
        if 'AmountPercentage' not in file.columns:
            ErrorMsg.append("H1:0:AmountPercentage字段缺失，请勿更改模板列标题")
        if 'CountPercentage' not in file.columns:
            ErrorMsg.append("J1:0:CountPercentage字段缺失，请勿更改模板列标题")
        if 'Count' not in file.columns:
            ErrorMsg.append("I1:0:Count字段缺失，请勿更改模板列标题")
        if 'DistributionType' not in file.columns:
            ErrorMsg.append("I1:0:DistributionType字段缺失，请勿更改模板列标题")

        if len(ErrorMsg) > 0:  # 如果有字段缺失，先补全字段后继续校验
            writeErrToSheet(sourceFilePath, ErrorMsg)
            return len(ErrorMsg)

        for i in range(file['PaymentPeriodID'].index.values.size):
            value = int(file['PaymentPeriodID'][i]) if file['PaymentPeriodID'][i] == file['PaymentPeriodID'][
                i] else -1  # 用于防止数据是NaN的情况
            if value != 0:
                val = int(file['PaymentPeriodID'].index.values[i])
                ErrorMsg.append("A{0}:0:PaymentPeriodID期数填写错误，应为0期".format(val + 2))

        # 检测BucketSequenceNo列是否有重复

        def checkUnique(file):
            if file['BucketSequenceNo'].is_unique == False:
                cells = ''
                for i in range(file['BucketSequenceNo'].index.values.size):
                    val = int(file['BucketSequenceNo'].index.values[i])
                    cells = cells + ErrLocation('BucketSequenceNo') + str(val + 2) + ';'
                ErrorMsg.append("{0}:0:BucketSequenceNo排序数不唯一，请修改".format(cells.rstrip(';')))

        pd.DataFrame(file.groupby('DistributionType').apply(checkUnique))

        # 检测AmountPercentage/CountPercentage列数值设置是否大于或小于100
        def Verification(file):
            ErrorNum = 0
            for i in range(file['AmountPercentage'].index.values.size):
                val = int(file['AmountPercentage'].index.values[i])
                flag = isinstance(file['AmountPercentage'][val], int) or isinstance(file['AmountPercentage'][val],
                                                                                    float)
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('AmountPercentage') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-AmountPercentage值：{1}不是数字，请修改".format(cell.rstrip(';'),
                                                                                         file['AmountPercentage'][val]))

                val = int(file['CountPercentage'].index.values[i])
                flag = isinstance(file['CountPercentage'][val], int) or isinstance(file['CountPercentage'][val], float)
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('CountPercentage') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-CountPercentage值：{1}不是数字，请修改".format(cell.rstrip(';'),
                                                                                        file['CountPercentage'][val]))
                val = int(file['Count'].index.values[i])
                flag = isinstance(file['Count'][val], int) or isinstance(file['Count'][val], float)
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('Count') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-Count值：{1}不是数字，请修改".format(cell.rstrip(';'), file['Count'][val]))

                val = int(file['Amount'].index.values[i])
                flag = isinstance(file['Amount'][val], int) or isinstance(file['Amount'][val], float)
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('Amount') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-Amount值：{1}不是数字，请修改".format(cell.rstrip(';'), file['Amount'][val]))

            if ErrorNum == 0:  # 类型全是数字，可以进行加总运算
                file['Count'] = file['Count'].astype(float)
                file['Amount'] = file['Amount'].astype(float)

                AmountPercentage_c = file['AmountPercentage'].sum()
                CountPercentage_c = file['CountPercentage'].sum()
                Count = file['Count'].sum()
                Amount = file['Amount'].sum()

                if (
                        AmountPercentage_c > 100.5 or AmountPercentage_c < 99.5) and CountPercentage_c != 0 and AmountPercentage_c != 0:
                    cells = ''
                    for i in range(file['AmountPercentage'].index.values.size):
                        val = int(file['AmountPercentage'].index.values[i])
                        cells = cells + ErrLocation('AmountPercentage') + str(val + 2) + ';'
                    ErrorMsg.append("{0}:0:列数值设置有误-AmountPercentage分布总计值：{1}不等于100(忽略精度)".format(cells.rstrip(';'),
                                                                                                 AmountPercentage_c))

                if (
                        CountPercentage_c >= 100.5 or CountPercentage_c <= 99.5) and CountPercentage_c != 0 and AmountPercentage_c != 0:
                    cells = ''
                    for i in range(file['CountPercentage'].index.values.size):
                        val = int(file['CountPercentage'].index.values[i])
                        cells = cells + ErrLocation('CountPercentage') + str(val + 2) + ';'
                    ErrorMsg.append("{0}:0:列数值设置有误-CountPercentage分布总计值：{1}不等于100(忽略精度)".format(cells.rstrip(';'),
                                                                                                CountPercentage_c))

                # 检测Count\Amount是否录反
                if Count > Amount:
                    cells = ''
                    for i in range(file['Count'].index.values.size):
                        val = int(file['Count'].index.values[i])
                        cells = cells + ErrLocation('Count') + str(val + 2) + ';'
                    for i in range(file['Amount'].index.values.size):
                        val = int(file['Amount'].index.values[i])
                        cells = cells + ErrLocation('Amount') + str(val + 2) + ';'
                    ErrorMsg.append(
                        "{0}:0:Count合计值：{1}不能大于Amount合计值：{2}".format(cells.rstrip(';'), Count, Amount))

        file = file.replace(np.nan, 0)
        pd.DataFrame(file.groupby('DistributionType').apply(Verification))
        writeErrToSheet(sourceFilePath, ErrorMsg)
        return len(ErrorMsg)
    except:
        ErrorMsg.append("文件错误，请确认使用模板，使用规定池分布模板, 或表格中填值不能为空，请填入NA！！！")
        writeErrToSheet(sourceFilePath, ErrorMsg)
        return -1


if __name__ == '__main__':
    sourceFileForder = r"C:/PyCharm\pdf-docx\source\Example\池分布"
    for dirPath, dirNames, fileNames in os.walk(sourceFileForder):
        for fileName in fileNames:
            if not fileName.endswith('池分布.xlsx'):
                msg = "【跳过】文件名称不符合，已跳过文件{0}".format(fileName)
                writeLog(msg)
                continue

            sourceFilePath = os.path.join(dirPath, fileName)

            fileNameAry = fileName.split('_')
            if len(fileNameAry) != 4:
                msg = "【文件名错误】文件名称命名不规范"
                writeLog(msg)
                continue

            excelwb = load_workbook(sourceFilePath)
            if '格式检查' in excelwb.sheetnames:
                excelwb.remove(excelwb['格式检查'])
                excelwb.save(sourceFilePath)

            hasError = CheckData(sourceFilePath)
            if hasError > 0 or hasError < 0:
                writeLog(sourceFilePath)
                writeLog('【有格式错误】详情见文档[格式检查]sheet')
            else:
                writeLog(sourceFilePath)
                writeLog('通过校验，无错误！！！')

