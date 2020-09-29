# -*- coding: utf-8 -*-
"""
Created on Thu Dec 12 15:10:53 2019

@author: HUAWEI
"""

#%%

#TrustId=None
#TrustStartDate=None
#PoolCloseDate=None
#ClosureDate=None
#IssueAmount=0

#Error=0
#Error1=0
def ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,ErrorInformation):
    #ErrorPath=r"C:\Users\HUAWEI\Desktop\Error\TrustError.xlsx"
    wb = openpyxl.load_workbook(errtxtFilePath) # 读取xlsx文件
    ws=wb['Sheet']
    maxrow=ws.max_row
    maxrow+=1
#    print(maxrow)
    ws.cell(maxrow,3,TrustId).value
    ws.cell(maxrow,4,TrustCode).value
    ws.cell(maxrow,5,TrustName).value
    ws.cell(maxrow,6,Excelfilepath).value
    ws.cell(maxrow,7,ErrorInformation).value
    ws.cell(maxrow,9,1).value	
    wb.save(errtxtFilePath)
    
    
   
def Trust(Excelfilepath,TrustCode,TrustName):
#    print(Excelfilepath)
    global TrustId,TrustStartDate,PoolCloseDate,ClosureDate,IssueAmount,V3
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    try:
        selectId="select TrustId from TrustManagement.Trust where TrustCode='{}'".format(TrustCode)
        b1.execute(selectId)
        TrustId=b1.fetchone()[0]
    except:
        A0='文件TrustCode与系统TrustCoden不匹配请检查!'
        
        ErrorMessage('#',TrustCode,TrustName,Excelfilepath,A0)
        print('文件TrustCode与系统TrustCoden不匹配请检查!')
#        sys.exit(1)
        return
    TrustError=0
    V3=0
    DataTrust=pd.read_excel(Excelfilepath,sheet_name="Trust")
    TrustBond=pd.read_excel(Excelfilepath,sheet_name="TrustBond")
    CouponPaymentReferencelist=[]

    for i in TrustBond.index:
        ItemCode1=TrustBond.iloc[i][2]
        ItemValue1=TrustBond.iloc[i][3]
        if ItemCode1=='CouponPaymentReference':
            CouponPaymentReferencelist.append(ItemValue1)

    #获取名称
    try:
        TrustName=DataTrust.ItemValue[0]
    except:
        TrustError=1
        V3=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'模板错误,请检查!')
        print('模板错误，请检查!')
    
    if pd.isnull(TrustName)==True:
        TrustError=1
        TE0='产品名称缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,'产品名称缺失',Excelfilepath,TE0)
        print(TrustId,TrustName,TE0)    
    

    #获取产品简称
    TrustShortName=DataTrust.ItemValue[1]
    if pd.isnull(TrustShortName)==True:
        TrustError=1
        TE0='产品简称缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,'产品简称缺失',Excelfilepath,TE0)
        print(TrustId,TrustName,TE0) 
    
    
    #获取产品主体
    ProductSubject1=DataTrust.ItemValue[5]
    if pd.isnull(ProductSubject1)==True:
        TrustError=1
        TE0='产品主体缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE0)
        print(TrustId,TrustName,TE0)

    #获取发行金额
    IssueAmount=DataTrust.ItemValue[7]
    if type(IssueAmount) not in (int,float):
        TrustError=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'Trust表-发行金额应为数值型!')
        print('发行金额应为数值型!')
    
    if pd.isnull(IssueAmount)==True:
        TrustError=1
        TE1='Trust表-发行金额缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE1)
        print(TrustId,TrustName,TE1)
        
    #封包日资产池余额
    AssetPoolBalance=DataTrust.ItemValue[8]
    if type(AssetPoolBalance) not in (int,float):
        TrustError=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'Trust表-发行金额应为数值型!')
        print('封包日资产池余额应为数值型!')
        
    IsTopUpAvailable_Trust=DataTrust.ItemValue[10]
    if pd.isnull(IsTopUpAvailable_Trust)==True:
        TrustError=1
        print('Trust表-是否为循环购买产品填写为空，请检查!')
    else:
        if IsTopUpAvailable_Trust=='是':
            IsTopUpAvailable_Trust=1
        else:
            IsTopUpAvailable_Trust=0
        
        upIsTopUpAvailable_Trust="update TrustManagement.Trust set IsTopUpAvailable={} where TrustId={}".format(IsTopUpAvailable_Trust,TrustId)
        b1.execute(upIsTopUpAvailable_Trust)
        conn.commit()
        print('是否循环购买已更新!')
    
        
    
    
    if pd.isnull(AssetPoolBalance)==True:
        TE11='封包日资产池余额缺失,请检查!'
        TrustError=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE11)
        print(TrustId,TrustName,TE11)
        
    #获取设立日
    TrustStartDate=DataTrust.ItemValue[12]
    if pd.isnull(TrustStartDate)==True:
        TrustError=1
        TE2='设立日缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE2)
        print(TrustId,TrustName,TE2)
    
    
    #获取封包日
    PoolCloseDate=DataTrust.ItemValue[11]
    if pd.isnull(PoolCloseDate)==True:
        TrustError=1
        TE3='封包日缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE3)
        print(TrustId,TrustName,TE3)
    
    
    #获取法定到期日
    ClosureDate=DataTrust.ItemValue[13]
    if pd.isnull(ClosureDate)==True:
        TrustError=1
        TE4='法定到期日缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE4)
        print(TrustId,TrustName,TE4)
    #----------------------------------------------------------------------------------

    #-----------------------------------
    #牵头主承销
    MainUnderwriter=DataTrust.ItemValue[15]      
    
    if pd.isnull(MainUnderwriter)==True:
        TrustError=1
        TE6='主承销商缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE6)
        print(TrustId,TrustName,TE6)
    
    
    
    elif '/' in MainUnderwriter:
        deleteMainUnderwritersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='MainUnderwriter'".format(TrustId)
        b1.execute(deleteMainUnderwritersql)
        conn.commit()
        
        listMainUnderwriter=MainUnderwriter.split('/')
        for MainUnderwriter_f in listMainUnderwriter:
            AdditionalItemCode='MainUnderwriter'
            ChineseField='主承销商'
            AdditionalContent=MainUnderwriter_f
            insertMainUnderwriter="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
            b1.execute(insertMainUnderwriter)
            conn.commit()
            print('主承销插入完成!')
    elif '\\' in MainUnderwriter:
        deleteMainUnderwritersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='MainUnderwriter'".format(TrustId)
        b1.execute(deleteMainUnderwritersql)
        conn.commit()
        
        listMainUnderwriter=MainUnderwriter.split('\\')
        for MainUnderwriter_f in listMainUnderwriter:
            AdditionalItemCode='MainUnderwriter'
            ChineseField='主承销商'
            AdditionalContent=MainUnderwriter_f
            insertMainUnderwriter="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
            b1.execute(insertMainUnderwriter)
            conn.commit()
            print('主承销插入完成!')
            
            
    else:
        deleteMainUnderwritersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='MainUnderwriter'".format(TrustId)
        b1.execute(deleteMainUnderwritersql)
        conn.commit()
        
        AdditionalItemCode='MainUnderwriter'
        ChineseField='主承销商'
        AdditionalContent=MainUnderwriter
        insertMainUnderwriter="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertMainUnderwriter)
        conn.commit()
        print('主承销插入完成!')
        
    #----------------------------
    
    #资金保管机构
    FundsCustodian=DataTrust.ItemValue[16]
    if pd.isnull(FundsCustodian)==True:
        TrustError=1
        TE7='资金保管机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE7)
        print(TrustId,TrustName,TE7)
    
    else:
        deleteFundsCustodiansql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='FundsCustodian'".format(TrustId)
        b1.execute(deleteFundsCustodiansql)
        conn.commit()
        
        
        AdditionalItemCode='FundsCustodian'
        ChineseField='资金保管机构'
        AdditionalContent=FundsCustodian
        insertFundsCustodian="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertFundsCustodian)
        conn.commit()
        print('资金保管机构插入完成!')
    
    #----------------------------------
    #发行人
    Trustee=DataTrust.ItemValue[17]
    if pd.isnull(Trustee)==True:
        TrustError=1
        TE8='发行人/机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE8)
        print(TrustId,TrustName,TE8)
        
    
        
    else:
        
        deleteTrusteesql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='Trustee'".format(TrustId)
        b1.execute(deleteTrusteesql)
        conn.commit()
        
        AdditionalItemCode='Trustee'
        ChineseField='发行人/机构'
        AdditionalContent=Trustee
        insertTrustee="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertTrustee)
        conn.commit()
        print('发行人/机构插入完成!')
        
    #-----------------------------
     
    #资产服务机构
    ServiceInstitutions=DataTrust.ItemValue[18]
    if pd.isnull(ServiceInstitutions)==True:
        TrustError=1
        TE9='资产服务机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE9)
        print(TrustId,TrustName,TE9)
        
        
    else:
        deleteServiceInstitutionssql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='ServiceInstitutions'".format(TrustId)
        b1.execute(deleteServiceInstitutionssql)
        conn.commit()
        
        AdditionalItemCode='ServiceInstitutions'
        ChineseField='资产服务机构/贷款服务机构'
        AdditionalContent=ServiceInstitutions
        insertServiceInstitutions="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertServiceInstitutions)
        conn.commit()
        print('资产服务机构(贷款服务机构)插入完成!')

    #---------------------------------
    #受托机构
    EntrustOrganization=DataTrust.ItemValue[19]
    if pd.isnull(ServiceInstitutions)==True:

        TrustError=1
        TE10='受托机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE10)
        print(TrustId,TrustName,TE10)
    
        
    
    else:
        
        deleteEntrustOrganizationsql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='EntrustOrganization'".format(TrustId)
        b1.execute(deleteEntrustOrganizationsql)
        conn.commit()
        
        AdditionalItemCode='EntrustOrganization'
        ChineseField='受托机构'
        AdditionalContent=ServiceInstitutions
        insertEntrustOrganization="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertEntrustOrganization)
        conn.commit()
        print('受托机构插入完成!')
        
        #----------------------
    BookkeepingDate=DataTrust.ItemValue[20]
    print(BookkeepingDate)
    BookkeepingDate=str(BookkeepingDate).split(' ')[0]
    if pd.isnull(BookkeepingDate)==True:
        TrustError=1
        TE11='薄记建档日不能为空,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE10)
        print(TrustId,TrustName,TE11)
    else:
        try:
            
            BookkeepingDate=datetime.datetime.strptime(BookkeepingDate,'%Y-%m-%d')
            upBookkeepingDate="update TrustManagement.Trust set BookkeepingDate=N'{}' where TrustId={}".format(BookkeepingDate,TrustId)
            b1.execute(upBookkeepingDate)
            conn.commit()
            print('薄记建档日更新完成!')
            
        except:
            TrustError=1
            Tdata='建档日日期格式不规范!应为日期格式(YYYY-mm-dd)'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,Tdata)
            print(Tdata)
        
    #有浮动利率债券时Trust表基准利率不能为空
    ProductFeatures=DataTrust.ItemValue[21]
    
    if pd.isnull(ServiceInstitutions)==True and '浮动利率' in CouponPaymentReferencelist:
        TrustError=1
        TE22='有浮动利率债券时Trust表基准利率不能为空'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE22)

    else:
        upProductFeatures="update TrustManagement.Trust set ProductFeatures=N'{}' where TrustId={}".format(ProductFeatures,TrustId)
        b1.execute(upProductFeatures)
        conn.commit()
        print('基准利率更新完成!')

    
    
    #---------------------------
    
    #插入产品主体
    selectProductSubject="select ItemCode from TrustManagement.TrustInfoExtension where TrustId={}".format(TrustId)
    b1.execute(selectProductSubject)
    SItemCode=b1.fetchall()
    TrustInfoExtension=[]
    for st in SItemCode:
        TrustInfoExtension.append(st[0])
    if pd.isnull(ProductSubject1)==False:
        if 'ProductSubject' not in TrustInfoExtension:
            ProductSubjectInsert="insert into TrustManagement.TrustInfoExtension values({},GETDATE(),NULL,NULL,'ProductSubject',N'{}')".format(TrustId,ProductSubject1)
#            b1.execute(ProductSubjectInsert)
            conn.commit()
            T0='{},{}产品主体插入完成!'.format(TrustId,TrustName)
#            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T0)
            print(T0)
        else:
            T0='{},{}产品主体已存在!'.format(TrustId,TrustName)
    #        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T0)
            print(T0)
        
    #更新发行金额
    if pd.isnull(IssueAmount)==False and type(IssueAmount)!=str:
        UpdateIssueAmount="update TrustManagement.Trust set IssueAmount={} where TrustId={}".format(IssueAmount,TrustId)
        b1.execute(UpdateIssueAmount)
        conn.commit()
        T1="{},{},'发行金额更新完成!'".format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T1)
        print(T1) 

            
    #更新资产池封包日余额
    if pd.isnull(AssetPoolBalance)==False  and type(AssetPoolBalance)!=str:
#        if SQLIAssetPoolBalance==None:
        UpdateAssetPoolBalance="update TrustManagement.Trust set AssetPoolBalance={} where TrustId={}".format(AssetPoolBalance,TrustId)
        b1.execute(UpdateAssetPoolBalance)
        conn.commit()
        T2="{},{},'封包日资产池余额更新完成!'".format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T1)
        print(T2) 

            
        
def TrustBond(Excelfilepath):
    global Error
    Error=0
    DataTrustBond1=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
    DataPrincipalSchedule=pd.read_excel(Excelfilepath,sheet_name='PrincipalSchedule')
    
    huanben=DataPrincipalSchedule.loc[2][2]
    print(huanben)

    try:
        DataTrustBond=DataTrustBond1[['TrustBondID','ItemID','ItemCode','ItemValue']]
        TrustBondID=DataTrustBond['TrustBondID'].drop_duplicates()

    except:
        T2="TrustBond表字段错误"
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T2)
        print('TrustBond字段错误!')
        return 
    
    DataTrustBond=DataTrustBond.dropna(subset=['ItemCode','ItemID'])
    if DataTrustBond['TrustBondID'].count()>25 and DataTrustBond['TrustBondID'].sum()==0:
        A1="TrustBond表分层标识(TrustBondID)重复'"
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A1)
        print(TrustId,TrustName,'分层标识重复(TrustBondID)')
        Error=1
    AmountC=0
    
    ZClass=0
    #各层级产品简称
    ShortNamelist=[]
    #各层级发行规模
    OfferAmountlist=[]
    #获取各层级起息日
    IssueDatelist=[]
    for i in DataTrustBond.index:       
        TrustBondId=DataTrustBond.loc[i][0]
        ItemId=int(DataTrustBond.loc[i][1])
        ItemCode=DataTrustBond.loc[i][2]
        ItemValue=DataTrustBond.loc[i][3]
        
        
        if ItemCode=='SecurityExchangeCode' and bool(re.search('[a-zA-Z]+', str(ItemValue)))==True:
            A2="{},{},'不能包含英文字符'".format(TrustBondId,ItemCode)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A2)
            print(TrustId,TrustName,TrustBondId,'债券代码不能包含英文字符')
            Error=1
            
        if ItemCode=='SecurityExchangeCode' and pd.isnull(ItemValue)==True:
            Error=1
            AJJ1='证券代码不能为空'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ1)
            print(TrustId,TrustName,AJJ1)
            
		#必填项不能为空
        if ItemCode in ('ShortName','SecurityExchangeCode','CurrencyOfIssuance','Denomination','LegalMaturityDate','InterestPaymentType','PaymentConvention','IssueDate','InterestRateCalculation','InterestDays','CouponBasis','CouponPaymentReference','RatingAgent','OriginalCreditRating','ClassName') and pd.isnull(ItemValue)==True:
            Error=1
            AJJ2='{}不能为空!'.format(ItemCode)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ2)
            print(TrustId,TrustName,AJJ2)
			
			
        
        if TrustBondId==0 and ItemCode=='ClassType' and ItemValue!='FirstClass':
            A3='优先级层级标识填写错误，请检查!'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A3)
            print(TrustId,TrustName,'优先级层级标识填写错误，请检查!')
            Error=1
            
        if DataTrustBond.TrustBondID.drop_duplicates().count()==1 and ItemCode=='ClassType' and ItemValue !='FirstClass':
            A4='{},债券层级填写错误'.format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A4)
            print(TrustId,TrustName,'{},债券层级填写错误'.format(TrustBondId))
            Error=1
            
        if DataTrustBond.TrustBondID.drop_duplicates().count()==2 and ItemCode=='ClassType' and ItemValue not in ('FirstClass','EquityClass'):
            A5='{},债券层级填写错误'.format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A5)
            print(TrustId,TrustName,'{},债券层级填写错误'.format(TrustBondId))
            Error=1
            
        if ItemCode=='CouponBasis' and type(ItemValue) not in (int,float):
            A6="{},票面利率(CouponBasis)数据类型应该为int、float".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A6)
            print(TrustId,TrustName,A6)
            Error=1
			
            
        if ItemCode=='ClassType' and ItemValue in ('FirstClass','SubClass'):
            ZClass=1

        if ItemCode=='CouponBasis' and ItemValue==0 and ZClass==1:
            print('ZCass:',ZClass)
            Error=1
            AJ='优先级或次优先级票面利率不能为0,请检查!'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJ)
            print(TrustId,TrustName,'优先级或次优先级票面利率不能为0')
            
        if ItemCode=='CouponBasis' and ItemValue!=0 and ZClass==1:
            ZClass=0
        
        if ItemCode=='OfferAmount' and type(ItemValue) not in (int,float):
            Error=1
            A7="{},发行规模(OfferAmount)数据类型应该为int、float".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A7)
            print(TrustId,TrustName,A7)
			
        if ItemCode=='OfferAmount' and pd.isna(ItemValue)==True:
            Error=1
            A17="{},发行规模(OfferAmount)不能为空".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A17)
            print(TrustId,TrustName,A17)
			
		
		#判断预计到期日是否为日期格式
        try:
            
            if ItemCode=='LegalMaturityDate':
                #print(ItemValue)
                ItemValue=str(ItemValue)
                ItemValue=datetime.datetime.strptime(ItemValue,"%Y-%m-%d %H:%M:%S")
        except:
            Error=1
            A18="{},预计到期日(LegalMaturityDate)应为日期格式(YYYY-mm-dd),且不能为空!".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A18)
            print(TrustId,TrustName,A18)	
		
		
            
            
        if ItemCode=='ClassType' and ItemValue not in ('FirstClass','SubClass','EquityClass'):
            A8="{},债券类别(ClassType)填写值为[{}]错误!只能填写'FirstClass','SubClass','EquityClass'".format(TrustBondId,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A8)
            print(TrustId,TrustName,"{},债券类别(ClassType)填写值为[{}]错误!只能填写'FirstClass','SubClass','EquityClass'".format(TrustBondId,ItemValue))
            Error=1

        
        if ItemCode=='InterestPaymentType' and ItemValue not in ('BaseOnPaymentDay'):
            A9="{},付息频率(InterestPaymentType)填写值[{}]错误!只能填写BaseOnPaymentDay".format(TrustBondId,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A9)
            print(TrustId,TrustName,"{},付息频率(InterestPaymentType)填写值['{}']错误!只能填写1,3,6,12".format(TrustBondId,ItemValue))
            Error=1


        if ItemCode=='InterestRateCalculation' and ItemValue not in ('天','月','半年','年','季'):
            A10="{},计息方式(InterestRateCalculation)填写值错误!只能是'天','月','半年','年','季'".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A10)
            print(A10)
            Error=1

                
        if ItemCode=='PaymentConvention' and ItemValue not in ('固定摊还','一次性还本付息','到期一次性还本','过手摊还','按期等额本金','按期等额本息'):
            A11="{},还本付息方式(PaymentConvention)填写值错误!只能为'固定摊还','一次性还本付息','到期一次性还本','过手摊还','按期等额本金','按期等额本息'".format(TrustBondId)
            
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A11)
            print(TrustId,TrustName,A11)
            Error=1
            
        if ItemCode=='PaymentConvention' and ItemValue=='固定摊还' and pd.isnull(huanben)==True:
            AJJ='{},固定摊还应有还本计划'.format(TrustBondId)
            print(TrustId,TrustName,AJJ)
            Error=1
        

        if ItemCode=='CouponPaymentReference' and ItemValue not in ('浮动利率','固定利率'):
            A12="利率形式(CouponPaymentReference)填写值[{}]错误!只能为'浮动利率','固定利率'".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A12)
            print(TrustId,TrustName,"利率形式(CouponPaymentReference)填写值[{}]错误!只能为'浮动利率','固定利率'".format(ItemValue))
            Error=1
            
            
        if ItemCode in ('OriginalCreditRating','ClassName') and ItemValue not in ('A','AA','AAA',np.nan,'NR','A+','AA+'):
            A13='{},债券评级填写值【{}】错误!'.format(TrustBondId,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A13)
            print(TrustId,TrustName,'{},债券评级填写值【{}】错误!'.format(TrustBondId,ItemValue))
            Error=1
            
            
        if ItemCode=='OfferAmount':
            Amount=ItemValue
            AmountC=AmountC+Amount
            
        #获取产品简称的中文字段
        if ItemCode=='ShortName':
            ItemValue=re.sub("[A-Za-z0-9\!\%\[\]\,\。\_]", "", ItemValue)
            ShortNamelist.append(ItemValue)
        #获取各层级的发行规模
        if ItemCode=='OfferAmount':
            OfferAmountlist.append(ItemValue)
        #获取起息日
        if ItemCode=='IssueDate':
            IssueDatelist.append(ItemValue)
        
    ShortNamelist=list((set(ShortNamelist)))
    if len(ShortNamelist)>1:
        Error=1
        A16="TrustBond表-债券简称的中文字段不一样请检查!"
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A16)
        print(A16)
        
    OfferAmountlist=list(set(OfferAmountlist))
    print(OfferAmountlist)
    print('层级数:',TrustBondID)
    if len(OfferAmountlist)!=len(set(TrustBondID)):
        Error=1
        A17='TrustBond表-债券不同层级发行规模有相等金额,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A17)
        print(A17)
		
    #判断各层级起息日是否相同
    IssueDatelist=list(set(IssueDatelist))
    if len(IssueDatelist)>1:
        Error=1
        A18='TrustBond表-各层级起息日不相同,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A18)
    #判断各层级计息日与Trust表设立日是否相同
    for m in IssueDatelist:
        m = parse(str(m))
        if m!=TrustStartDate:
            Error=1
            A19='TrustBond表的层级起息日[{}]与Trust表的设立日{}不一致，请检查!'.format(m,TrustStartDate)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A19)
            print(A19)

        
    if  IssueAmount-AmountC>1000 or IssueAmount-AmountC<-1000:  #7
        A14='发行金额{}与层级金额{}相加不符'.format(IssueAmount,AmountC)
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A14)
        
        print(TrustId,TrustName,'发行金额{}与层级金额{}相加不符'.format(IssueAmount,AmountC))
        Error=1
    print('Error:',Error)
    Errorlist.append(Error)
    
    if Error==0:
#        A15='TrustBond表校验通过!'
        #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A15)
        print(TrustId,TrustName,'TrustBond表校验通过!')
        TrustBondImport(Excelfilepath,TrustId,TrustName)
        


def TrustExtension(Excelfilepath):
    global Error1
    Error1=0
    DataTrustExtension=pd.read_excel(Excelfilepath,sheet_name='TrustExtension')
    try:
        DataTrustExtension=DataTrustExtension[['ItemCode','ItemValue']]
    except:
        Error1=1
        B000='TrustExtension表字段错误!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B000)
        print(TrustId,TrustName,B000)
        return
			
    for Extensionindex in DataTrustExtension.index:
        ItemCode=DataTrustExtension.loc[Extensionindex][0]
        ItemValue=DataTrustExtension.loc[Extensionindex][1]
        
        if ItemCode=='PoolCloseDate' and ItemValue!=PoolCloseDate:
            B0='Trust表封包日[{}]与TrustExtension表封包日[{}]不符请检查!'.format(PoolCloseDate,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B0)
            print(TrustId,TrustName,'Trust表封包日[{}]与TrustExtension表封包日[{}]不符请检查!'.format(PoolCloseDate,ItemValue))
            Error1=1     
            
        if ItemCode=='TrustStartDate' and ItemValue!=TrustStartDate:
            B1='Trust表设立日日[{}]与TrustExtension表设立日[{}]不符请检查!'.format(TrustStartDate,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B1)
            print(TrustId,TrustName,'Trust表设立日日[{}]与TrustExtension表设立日[{}]不符请检查!'.format(TrustStartDate,ItemValue))
            Error1=1
            
            
        if ItemCode=='ClosureDate' and ItemValue!=ClosureDate:
            B2='Trust表法定到期日[{}]与TrustExtension表法定到期日[{}]不符请检查!'.format(ClosureDate,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B2)
            print(TrustId,TrustName,'Trust表法定到期日[{}]与TrustExtension表法定到期日[{}]不符请检查!'.format(ClosureDate,ItemValue))
            Error1=1
            
        if ItemCode=='PaymentDate' and ItemValue not in ('InterestCollectionDate','PaymentDate'):
            Error1=1
            B="用何种日期进行计息(PaymentDate)填写值[{}]错误!只能填写'InterestCollectionDate','PaymentDate'".foramt(ItemValue)
            ErrorMessage(TrustId,TrustCode,Excelfilepath,B)
            print(TrstId,TrustName,B)
            
            
        if ItemCode=='IsTopUpAvailable' and ItemValue not in (False,True):
            B3="是否循环购买(IsTopUpAvailable)填写值错误!只能填写False,True"
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B3)
            print(TrustId,TrustName,B3)
            Error1=1
        
        

        if ItemCode in ('B_PaymentDate','R_CollectionDate','R_InterestCollectionDate','R_PaymentDate') and pd.isnull(ItemValue)==False:
            B4="日期(B_InterestCollectionDate','B_CollectionDate','B_PaymentDate'等)填写值为[{}]错误 --只能为空值".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B4)
            print(TrustId,TrustName,"日期(B_InterestCollectionDate','B_CollectionDate','B_PaymentDate'等)填写值为[{}]错误 --只能为空值".format(ItemValue))
            Error1=1
            
            
        #第一个计息日与产品成立日和法定到期日的对比
        if ItemCode =='B_InterestCollectionDate_FirstDate' and DataTrustExtension.ItemValue[3]=='InterestCollectionDate':
            ItemValue=str(ItemValue)
            ItemValue=ItemValue.split(' ')[0]
            try:
                ItemValue=datetime.datetime.strptime(ItemValue,"%Y-%m-%d")
                if ItemValue<TrustStartDate or ItemValue>ClosureDate:
                    B5="TrustExtensionb表-第一个计息日(B_InterestCollectionDate_FirstDate)应大于产品成立日(TrustStartDate)小于法定到期日(ClosureDate)"
                    ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B5)
                    print(TrustId,TrustName,B5)
                    Error1=1
            except:
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个计息日应该为日期格式')
                Error1=1
                print('第一个计息日,应为日期格式')
                
         #第一个计算日与产品成立日和法定到期日的对比       
        if ItemCode=='B_CollectionDate_FirstDate':
            ItemValue=str(ItemValue)
            ItemValue=ItemValue.split(' ')[0]
            try:
                ItemValue=datetime.datetime.strptime(ItemValue,"%Y-%m-%d")
                if ItemValue<TrustStartDate or ItemValue>ClosureDate:
                    B55="TrustExtensionb表-第一个计算日(B_CollectionDate_FirstDate)应大于产品成立日(TrustStartDate)小于法定到期日(ClosureDate)"
                    ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B5)
                    print(TrustId,TrustName,B55)
                    Error1=1
            except:
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个计算日应该为日期格式')
                Error1=1
                print('第一个计算日,应为日期格式')
        
        #、第一个兑付日与产品成立日和法定到期日的对比 
        if ItemCode =='B_PaymentDate_FirstDate':
            ItemValue=str(ItemValue)
            ItemValue=ItemValue.split(' ')[0]
            try:
                ItemValue=datetime.datetime.strptime(ItemValue,"%Y-%m-%d")
                if ItemValue<TrustStartDate or ItemValue>ClosureDate:
                    B555="TrustExtensionb表-第一个兑付日(B_PaymentDate_FirstDate)应大于产品成立日(TrustStartDate)小于法定到期日(ClosureDate)"
                    ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B5)
                    print(TrustId,TrustName,B555)
                    Error1=1
            except:
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个兑付日应该为日期格式')
                Error1=1
                print('第一个兑付日,应为日期格式')   
            
            
        if ItemCode in ('B_CollectionDate_Frequency','B_PaymentDate_Frequency') and ItemValue not in (1,3,6,12):
            B6="频率('B_CollectionDate_Frequency','B_PaymentDate_Frequency')填写值[{}]错误!只能填写1、3、6、12".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B6)
            print(TrustId,TrustName,"频率('B_InterestCollectionDate_Frequency','B_CollectionDate_Frequency','B_PaymentDate_Frequency')填写值[{}]错误!只能填写1、3、6、12".format(ItemValue))
            Error1=1
            
            
        if ItemCode in ('B_CollectionDate_Condition','B_PaymentDate_Condition') and ItemValue not in ('True','False'):
            B7="条件('B_CollectionDate_Condition','B_PaymentDate_Condition')填写值[{}]错误!只能填写True、False".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B7)
            print(TrustId,TrustName,B7)
            Error1=1
            
            
        if ItemCode in ('B_CollectionDate_ConditionCalendar','B_PaymentDate_ConditionCalendar') and ItemValue not in ('WorkingDay','NaturalDay','TradingDay'):
            B8="条件日历('B_CollectionDate_ConditionCalendar','B_PaymentDate_ConditionCalendar')填写值[{}]错误!只能是'WorkingDay','NaturalDay','TradingDay'".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B8)
            print(TrustId,TrustName,B8)
            Error1=1
            
            
        if ItemCode in ('B_CollectionDate_ConditionDay','B_PaymentDate_ConditionDay') and ItemValue not in (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31):
            B9="条件日期('B_CollectionDate_ConditionDay','B_PaymentDate_ConditionDay')填写值[{}]错误!填写规则详见excel备注信息.".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B9)
            print(TrustId,TrustName,B9)
            Error1=1
            
            
        if ItemCode in ('B_CollectionDate_ConditionTarget','B_PaymentDate_ConditionTarget') and ItemValue not in ('BeginingOfMonth','EndOfMonth'):
            B10="条件对象('B_CollectionDate_ConditionTarget','B_PaymentDate_ConditionTarget')填写值[{}]错误!只能填写'BeginingOfMonth','EndOfMonth'".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B10)       
            print(TrustId,TrustName,B10)
            Error1=1
            
            
        if ItemCode in ('B_CollectionDate_WorkingDateAdjustment','B_PaymentDate_WorkingDateAdjustment') and ItemValue not in (0,1,-1):
            B11="工作日调整('B_CollectionDate_WorkingDateAdjustment')填写值[{}]错误!只能填写0(不调整)、1(调整)"
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B11)
            print(TrustId,TrustName,B11)
            Error1=1
            
            
        if ItemCode in ('B_CollectionDate_Canlendar','B_PaymentDate_Canlendar') and ItemValue not in ('WorkingDay','NaturalDay','TradingDay'):
            B12="日历('B_CollectionDate_Canlendar','B_PaymentDate_Canlendar')填写值为[{}]错误!只能填写'WorkingDay','NaturalDay','TradingDay'".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B12)
            print(TrustId,TrustName,B12)
            Error1=1
			
			
		#当以何种方式计息为兑付日
        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and pd.isnull(DataTrustExtension.ItemValue[7])==True:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-第一个计息日不能为空')
            print(TrustId,TrustName,'TrustExtension表-第一个计息日不能为空')
            Error1=1
		
        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and DataTrustExtension.ItemValue[8] not in (1,3,6,12):
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-产品计息日[频率]填写错误,请检查!')
            print(TrustId,TrustName,'TrustExtension表-产品计息日[频率]填写错误,请检查!')
            Error1=1
			
        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and DataTrustExtension.ItemValue[9] !='True':
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-产品计息日[条件]填写错误,请检查!')
            print(TrustId,TrustName,'TrustExtension表-产品计息日[条件]填写错误,请检查!')
            Error1=1			
			
        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and DataTrustExtension.ItemValue[10] !='NaturalDay':
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-产品计息日[条件日历]填写错误,请检查!')
            print(TrustId,TrustName,'TrustExtension表-产品计息日[条件日历]填写错误,请检查!')
            Error1=1

        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and DataTrustExtension.ItemValue[11] not in (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31):
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-产品计息日[条件日期]填写错误,请检查!')
            print(TrustId,TrustName,'TrustExtension表-产品计息日[条件日期]填写错误,请检查!')
            Error1=1
			
        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and DataTrustExtension.ItemValue[12] not in ('BeginingOfMonth','EndOfMonth'):
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-产品计息日[条件对象]填写错误,请检查!')
            print(TrustId,TrustName,'TrustExtension表-产品计息日[条件对象]填写错误,请检查!')
            Error1=1

        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and DataTrustExtension.ItemValue[13] not in (0,1,-1):
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-产品计息日[工作日调整]填写错误,请检查!')
            print(TrustId,TrustName,'TrustExtension表-产品计息日[工作日调整]填写错误,请检查!')
            Error1=1
			
        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and DataTrustExtension.ItemValue[14] not in ('NaturalDay','WorkingDay','TradingDay'):
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-产品计息日[日历]填写错误,请检查!')
            print(TrustId,TrustName,'TrustExtension表-产品计息日[日历]填写错误,请检查!')
            Error1=1
			
            
        try:
                #取计息日
            date=1
            if ItemCode=='B_InterestCollectionDate_FirstDate' and DataTrustExtension.ItemValue[3]=='InterestCollectionDate':
                            
                B_InterestCollectionDate_FirstDate=str(ItemValue)
                B_InterestCollectionDate_FirstDate=B_InterestCollectionDate_FirstDate.split(' ')[0]
                B_InterestCollectionDate_FirstDate=datetime.datetime.strptime(B_InterestCollectionDate_FirstDate,"%Y-%m-%d")
#                print(B_PaymentDate_FirstDate)
                date=0
        except:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个计息日应为日期格式（YYYY-mm-dd）')
            Error1=1
            date=1
            print('计息日应为日期格式')
            print(date)
        
        try:
            
                #取计算日
            if ItemCode=='B_CollectionDate_FirstDate':
                B_CollectionDate_FirstDate=str(ItemValue)
                B_CollectionDate_FirstDate=B_CollectionDate_FirstDate.split(' ')[0]
                B_CollectionDate_FirstDate=datetime.datetime.strptime(B_CollectionDate_FirstDate,"%Y-%m-%d")
                print(B_CollectionDate_FirstDate)
                date1=0
        except:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个计算日应为日期格式（YYYY-mm-dd）')
            Error1=1
            date1=1
            print('计算日应为日期格式')
            print(date)      
            
        try:
                #取兑付日
            if ItemCode=='B_PaymentDate_FirstDate':
                B_PaymentDate_FirstDate=str(ItemValue)
                B_PaymentDate_FirstDate=B_PaymentDate_FirstDate.split(' ')[0]
                B_PaymentDate_FirstDate=datetime.datetime.strptime(B_PaymentDate_FirstDate,"%Y-%m-%d")
#                print(B_InterestCollectionDate_FirstDate)
                date2=0
        except:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个兑付日应为日期格式（YYYY-mm-dd）')
            Error1=1			
            date2=1
            print('兑付日应为日期格式')
            print(date)
        

         #检查计算日与兑付日
    if date1==0 and date2==0 and B_CollectionDate_FirstDate>=B_PaymentDate_FirstDate:
        Error1=1
        AJJ3='TrustExtensionb表-计算日(B_CollectionDate_FirstDate)【{}】应该在兑付日(B_PaymentDate_FirstDate)【{}】之前!'.format(B_CollectionDate_FirstDate,B_PaymentDate_FirstDate)
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ3)
        print(TrustId,TrustName,AJJ3)
    
    #计息与兑付
    if date==0 and date2==0 and DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and B_PaymentDate_FirstDate<B_InterestCollectionDate_FirstDate:
        Error1=1
        AJJ4="TrustExtensionb表-第一个兑付日应大于等于第一个计息日"
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ4)
        print(TrustId,TrustName,AJJ4)
        
        #B_PaymentDate_FirstDate=datetime.datetime.strptime(B_PaymentDate_FirstDate, '%Y-%m-%d')
        #B_InterestCollectionDate_FirstDate=datetime.datetime.strptime(B_PaymentDate_FirstDate, '%Y-%m-%d')
        
        jxryear=B_PaymentDate_FirstDate.year
        dfryear=B_InterestCollectionDate_FirstDate.year
        jxrmonth=B_PaymentDate_FirstDate.month
        dfrmonth=B_InterestCollectionDate_FirstDate.month
        
        if jxryear!=dfryear or jxrmonth!=dfrmonth:
            Error1=1
            AJJ4_1='TrustExtensionb表-第一个兑付日与第一个计息日首次日期应在同一年同一月'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ4_1)
            print(TrustId,TrustName,AJJ4_1) 
     
    #判断循环购买产品循环期长是否为数值
    if DataTrustExtension.ItemValue[4]==True:
        if pd.isnull(DataTrustExtension.ItemValue[5])==True:
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-循环购买产品应有循环期长!')
            print('TrustExtension表-循环购买产品应有循环期长!')
            
        else:
            if isinstance(DataTrustExtension.ItemValue[5],int)==True:
                pass
            else:
                Error1=1
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-循环期长应为数值型!')
                print('TrustExtension表-循环期长应为数值型!')
        
        #---------
        #判断循环购买产品是否含有循环期数据
    if DataTrustExtension.ItemValue[4]==True:
        if DataTrustExtension.ItemValue[3]=='InterestCollectionDate':
            if pd.isnull(DataTrustExtension.ItemValue[34])==True or pd.isnull(DataTrustExtension.ItemValue[38])==True or pd.isnull(DataTrustExtension.ItemValue[43])==True or pd.isnull(DataTrustExtension.ItemValue[47])==True or pd.isnull(DataTrustExtension.ItemValue[52])==True or pd.isnull(DataTrustExtension.ItemValue[56])==True:
                AJJ5_1='此产品为循环期购买产品,TrustExtension表-缺失循环期数据，请检查!'
                Error1=1
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ5_1)
                print(AJJ5_1)
        else:
            if pd.isnull(DataTrustExtension.ItemValue[43])==True or pd.isnull(DataTrustExtension.ItemValue[47])==True or pd.isnull(DataTrustExtension.ItemValue[52])==True or pd.isnull(DataTrustExtension.ItemValue[56])==True:
                AJJ5_12='此产品为循环期购买产品,TrustExtension表-缺失循环期数据，请检查!'
                Error1=1
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ5_12)
                print(AJJ5_12)
				
    else:
        if pd.isnull(DataTrustExtension.ItemValue[5])==False:
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'TrustExtension表-非循环购买产品，循环期长应为空!')
            print('TrustExtension表-非循环购买产品，循环期长应为空!')
    
    
    for name in ['和信','融腾','和享','和智','永动','捷赢','幸福','兴晴']:
        if name in TrustName and DataTrustExtension.ItemValue[4]==False:
            Error1=1
            AJJ6_1="产品名称包含,'和信','融腾','和享','和智','永动','捷赢','幸福','兴晴'的为循环购买产品,与TrustExtension-表中-是否循环购买（IsTopUpAvailable）冲突,请检查!"
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ6_1)
            print(AJJ6_1)
        
    #循环期产品，循环期第一个日期—+循环期长=摊还期第一个日期
    if DataTrustExtension.ItemValue[4]==True and pd.isnull(DataTrustExtension.ItemValue[5])==False and isinstance(DataTrustExtension.ItemValue[5],int)==True:
	
        shifdate=0
        RevolvingPeriod=DataTrustExtension.ItemValue[5]
        #摊还期计息日
        B_FirstDate_JX=DataTrustExtension.ItemValue[7]
        try:
            if DataTrustExtension.ItemValue[3]=='InterestCollectionDate':
                B_FirstDate_JX=parse(str(B_FirstDate_JX))
        except:
            shifdate=1
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'摊还期第一个计息日非日期格式请检查!')
			
        #摊还期计算日
        B_FirstDate_JS=DataTrustExtension.ItemValue[16]
        try:
            B_FirstDate_JS=parse(str(B_FirstDate_JS))
        except:
            shifdate=1
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'摊还期第一个计算日非日期格式请检查!')
			
			
        #摊还期兑付日
        B_FirstDate_DF=DataTrustExtension.ItemValue[25]
		
        try:
            B_FirstDate_DF=parse(str(B_FirstDate_DF))
        except:
            shifdate=1
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'摊还期第一个兑付日非日期格式请检查!')
			
        #循环期计息日
        R_FirstDate_JX=DataTrustExtension.ItemValue[34]
        try:
            if DataTrustExtension.ItemValue[3]=='InterestCollectionDate':
                R_FirstDate_JX=parse(str(R_FirstDate_JX))
        except:
            shifdate=1
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'循环期第一个计息日非日期格式请检查!')
			
        #循环期计算日
        R_FirstDate_JS=DataTrustExtension.ItemValue[43]
        try:
            R_FirstDate_JS=parse(str(R_FirstDate_JS))
        except:
            shifdate=1
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'循环期第一个计算日非日期格式请检查!')
			
        #循环期兑付日
        R_FirstDate_DF=DataTrustExtension.ItemValue[52]
        try:
            R_FirstDate_DF=parse(str(R_FirstDate_DF))
        except:
            shifdate=1
            Error1=1
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'循环期第一个兑付日非日期格式请检查!')
        
        if shifdate==0:
            if DataTrustExtension.ItemValue[3]=='InterestCollectionDate':
                R_FirstDate_JX_RevolvingPeriod=R_FirstDate_JX+relativedelta(months=+RevolvingPeriod)
			
            R_FirstDate_JS_RevolvingPeriod=R_FirstDate_JS+relativedelta(months=+RevolvingPeriod)
			
            R_FirstDate_DF_RevolvingPeriod=R_FirstDate_DF+relativedelta(months=+RevolvingPeriod)
			
			
            if DataTrustExtension.ItemValue[3]=='InterestCollectionDate' and R_FirstDate_JX_RevolvingPeriod!=B_FirstDate_JX:
                Error1=1
                AJJ7_1='循环购买产品,循环期第一个计息日加循环期长应等于摊还期第一个计息日,请检查!'
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ7_1)
                print(AJJ7_1)
			
            if R_FirstDate_JS_RevolvingPeriod!=B_FirstDate_JS:
                Error1=1
                AJJ7_2='循环购买产品,循环期第一个计算日加循环期长应等于摊还期第一个计算日,请检查!'
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ7_2)
                print(AJJ7_2)
			
            if R_FirstDate_DF_RevolvingPeriod!=B_FirstDate_DF:
                Error1=1
                AJJ7_3='循环购买产品,循环期第一个兑付日加循环期长应等于摊还期第一个兑付日，请检查!'
                ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ7_3)
                print(AJJ7_3)
        
            
            
                
    print('Error1:',Error1)
    Errorlist1.append(Error1)
    if Error1==0:
        B13='TrustExtension表校验通过!'
        #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B13)
        print(TrustId,TrustName,B13)
        TrustExtensionImport(Excelfilepath)


            

def TrustBondImport(Excelfilepath,TrustId,TrustName):
    
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    deletepjsql="delete DV.RatingAgency where TrustId={} and RatingAgencyCode in ('RatingAgent','RatingAgent-1','OriginalCreditRating','OriginalCreditRating-1','ClassName','ClassName-1')".format(TrustId)
    b1.execute(deletepjsql)
    conn.commit()
	
	 #将评级插入RatingAgency
    imRatingAgencydata=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
    imRatingAgencydata=imRatingAgencydata[['TrustBondID','ItemID','ItemCode','ItemValue']]
    for Agency in imRatingAgencydata.index:
        TrustBondId=imRatingAgencydata.loc[Agency][0]
#        ItemId=int(imRatingAgencydata.loc[Agency][1])
        ItemCode=imRatingAgencydata.loc[Agency][2]
        ItemValue=imRatingAgencydata.loc[Agency][3]
        
        if ItemCode in ('RatingAgent','RatingAgent-1'):
            agencyname='评级机构'
            insertRatingAgency="insert into DV.RatingAgency values({},{},'{}',N'{}',N'{}')".format(TrustId,TrustBondId,ItemCode,agencyname,ItemValue)
            b1.execute(insertRatingAgency)
            conn.commit()
            print('证券机构插入完成!')
        if ItemCode in ('OriginalCreditRating','OriginalCreditRating-1'):
            agencyname='原始证券评级'
            insertRatingAgency="insert into DV.RatingAgency values({},{},'{}',N'{}',N'{}')".format(TrustId,TrustBondId,ItemCode,agencyname,ItemValue)
            b1.execute(insertRatingAgency)
            conn.commit()
            print('原始评级插入完成')
        if ItemCode in ('ClassName','ClassName-1'):
            agencyname='当前证券评级'
            insertRatingAgency="insert into DV.RatingAgency values({},{},'{}',N'{}',N'{}')".format(TrustId,TrustBondId,ItemCode,agencyname,ItemValue)
            b1.execute(insertRatingAgency)
            conn.commit()
            print('当前评级插入完成!')
			
        #删除上次插入的TrustBond数据
    deleteTrustBond="delete TrustManagement.TrustBond2 where TrustId={}".format(TrustId)
    b1.execute(deleteTrustBond)
    conn.commit()        
            

    DataTrustBond=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
    DataTrustBond=DataTrustBond[['TrustBondID','ItemID','ItemCode','ItemValue']]
    DataTrustBond.dropna(subset=['ItemCode'],inplace=True)
    DataTrustBond.dropna(subset=['ItemID'],inplace=True)
#        del DataTrustBond['Unnamed: 5']
#        del DataTrustBond['StartDate']
#        del DataTrustBond['EndDate']
    for i in DataTrustBond.index:
        TrustBondId=DataTrustBond.loc[i][0]
        ItemId=int(DataTrustBond.loc[i][1])
        ItemCode=DataTrustBond.loc[i][2]
        ItemValue=DataTrustBond.loc[i][3]
#            print(TrustBondId,ItemId,ItemCode,ItemValue)  
        if ItemCode=='CouponBasis' and ItemValue<1:
            ItemValue=ItemValue*100
        InsertBond="insert into TrustManagement.TrustBond2(TrustBondId,TrustId,StartDate,EndDate,ItemId,ItemCode,ItemValue) values({},{},GETDATE(),NULL,{},'{}',N'{}')".format(TrustBondId,TrustId,ItemId,ItemCode,ItemValue)
    #    print(sql)
        b1.execute(InsertBond)
    conn.commit()
    BI='TrustBond导入完成!'
    #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,BI)
    print(TrustId,TrustName,'TrustBond导入完成!')
		
    #调用导入还本计划函数
    PrincipalSchedule(ExceLFilePath)
		#临时表推正式表
#        BondUp1="update TrustManagement.TrustBond2 set ItemValue=replace(ItemValue,N'nan','') where TrustId={}".format(TrustId)
#        BondUp2="update TrustManagement.TrustBond2 set ItemValue=replace(ItemValue,N' 00:00:00','') where TrustId={}".format(TrustId)
#        Bondinsert="insert into TrustManagement.TrustBond select * from TrustManagement.TrustBond2 where TrustId={}".format(TrustId)
#        b1.execute(BondUp1)
#        conn.commit()
#        b1.execute(BondUp2)
#        conn.commit()
#        b1.execute(Bondinsert)
#        conn.commit()
#        print('TrustBond已插入正式表')
     
        
def TrustExtensionImport(Excelfilepath):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()

    deleteTrustExtension="delete TrustManagement.TrustExtension1 where TrustId={}".format(TrustId)
    b1.execute(deleteTrustExtension)
    conn.commit()
    
    TrustExtension=pd.read_excel(Excelfilepath,sheet_name='TrustExtension')
    TrustExtension=TrustExtension[['ItemCode','ItemValue']]
    #TrustExtension=TrustExtension[1:]
    
    if str(TrustExtension.ItemValue[4])=='False':
        index=32
    else:
        index=60
#    print(TrustExtension.ItemValue[4])
#    print(index)
#    print(TrustExtension.ItemCode[index])

    for tindex in range(0,index):
        TrustExtension_ItemCode=TrustExtension.iloc[tindex][0]
        TrustExtension_ItemValue=TrustExtension.iloc[tindex][1] 
			 
        if TrustExtension_ItemCode=='B_InterestCollectionDate_Canlendar':
            TrustExtension_ItemCode='B_InterestCollectionDate_Calendar'
        if TrustExtension_ItemCode=='B_CollectionDate_Canlendar':
            TrustExtension_ItemCode='B_CollectionDate_Calendar'
        if TrustExtension_ItemCode=='B_PaymentDate_Canlendar':
            TrustExtension_ItemCode='B_PaymentDate_Calendar'
            #转化Bool
        if TrustExtension_ItemCode=='B_InterestCollectionDate_Frequency' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_CollectionDate_Frequency':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_PaymentDate_Frequency':
				
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_InterestCollectionDate_Frequency' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_CollectionDate_Frequency':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_PaymentDate_Frequency':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
            
            #----
        if TrustExtension_ItemCode=='B_InterestCollectionDate_WorkingDateAdjustment' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_CollectionDate_WorkingDateAdjustment':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_PaymentDate_WorkingDateAdjustment':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_InterestCollectionDate_WorkingDateAdjustment' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_CollectionDate_WorkingDateAdjustment':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_PaymentDate_WorkingDateAdjustment':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        
        if TrustExtension_ItemCode=='B_InterestCollectionDate_ConditionDay' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_CollectionDate_ConditionDay':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_PaymentDate_ConditionDay':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_InterestCollectionDate_ConditionDay' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_CollectionDate_ConditionDay':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_PaymentDate_ConditionDay':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
				
        if TrustExtension_ItemCode=='RevolvingPeriod' and pd.isnull(TrustExtension_ItemValue)==False:
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)            
			
			
        Itemid_sql="select ItemId from TrustManagement.Item where ItemCode='{}'".format(TrustExtension_ItemCode)
        b1.execute(Itemid_sql)
        ItemId=b1.fetchone()[0]

        TrustExtensionInsert="insert into TrustManagement.TrustExtension1(TrustId,StartDate,EndDate,ItemId,ItemCode,ItemValue) values({},GETDATE(),NULL,{},'{}','{}')".format(TrustId,ItemId,TrustExtension_ItemCode,TrustExtension_ItemValue)
            
        b1.execute(TrustExtensionInsert)
    conn.commit()
    EI='TrustExtension导入完成!'
    #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,EI)
    print(TrustId,TrustName,EI)
		
		
		#临时表推正式表
#    EUp1="update TrustManagement.TrustExtension1 set ItemValue=replace(ItemValue,N'nan','') where TrustId={}".format(TrustId)
#    EUp2="update TrustManagement.TrustExtension1 set ItemValue=replace(ItemValue,N' 00:00:00','') where TrustId={}".format(TrustId)
		
		
#    Einsert="insert into TrustManagement.TrustExtension select * from TrustManagement.TrustExtension1 where TrustId={} and TrustId not in (select TrustId from TrustManagement.TrustExtension where TrustId={})".format(TrustId,TrustId)
#    b1.execute(EUp1)
#    conn.commit()
#    b1.execute(EUp2)
#    conn.commit()
#    b1.execute(Einsert)
#    conn.commit()
#    print('TrustExtension表数据已插入正式表!')
#		
#		#调用存储过程生成日期期间
#    try:
#        execsql="exec [TrustManagement].[usp_GenerateTrustPeriod] '{}'".format(TrustCode)
#        b1.execute(execsql)
#    except:
#        print('日期期间生成失败!')
		
        
def PrincipalSchedule(Excelfilepath):
    global huanben
    
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    deletePrincipalSchedule="delete dbo.PrincipalSchedule1 where TrustId={}".format(TrustId)
    b1.execute(deletePrincipalSchedule)
    conn.commit()
    
    DataPrincipalSchedule=pd.read_excel(Excelfilepath,sheet_name='PrincipalSchedule')
    huanben=DataPrincipalSchedule.loc[2][2]
    huanben1=DataPrincipalSchedule.loc[3][2]
    
    if pd.isnull(huanben)==True and pd.isnull(huanben1)==True:
#        print(huanben)
        print('无还本计划!')
    else:
#        DataPrincipalSchedule=pd.read_excel(Excelfilepath,sheet_name='PrincipalSchedule',header=None)
        DataPrincipalSchedule=DataPrincipalSchedule.iloc[:,:3]
#        print(DataPrincipalSchedule)
        for Prindex in DataPrincipalSchedule.index:
            TrustBondID=DataPrincipalSchedule.loc[Prindex][0]
            Pdate=DataPrincipalSchedule.loc[Prindex][1]
            
            DebtPlan=DataPrincipalSchedule.loc[Prindex][2]  
            InsertPrincipalSchedule="insert into dbo.PrincipalSchedule1 values({},'{}','{}','{}')".format(TrustId,TrustBondID,Pdate,DebtPlan)
            b1.execute(InsertPrincipalSchedule)
        conn.commit()
        print('还本计划导入完成!')
        
        
        
        
def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        writeLog(str(ex))
        writeErr("\n【{0}】".format(sourceFilePath))
        writeErr(str(ex))
        print(str(ex))
        # raise ex
    finally:
        cnxn.close()

def InsertVerificationLog(userId, CheckType, IsSucess, filePath, result):
    sql = "exec TaskCollection.dbo.usp_InsertVerificationLog N'{0}',N'{1}',N'{2}',N'{3}',N'{4}'".format(userId,CheckType,IsSucess, filePath, result)
    execSQLCmd(sql) 

def InsertTrusteeCheckByTrustId(userId, TrustId,ckResultLen):
    sql = "exec TaskCollection.dbo.usp_InsertTrusteeCheckByTrustId N'{0}',N'{1}',N'{2}',N'{3}',N'{4}'".format(userId,9, TrustId, 0, ckResultLen)
    execSQLCmd(sql)

           
def execSQLCmdFetchAll(sql):
    print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)

    try:
        cursor = cnxn.cursor()
        rows = cursor.execute(sql).fetchall()
        cnxn.commit()
        return rows
    except Exception as ex:
        raise ex
    finally:
        cnxn.close()		
		
def runDBDataValidation(TrustId):
    sql = "exec TaskCollection.dbo.usp_VerifyTrustServiceBasicByTrust N'{0}'".format(TrustId)
    dbCheckResult = execSQLCmdFetchAll(sql)
    if len(dbCheckResult) > 0:
        wb = Workbook()  # 新建工作簿
        ws1 = wb.active
        wb.save(errtxtFilePath)
        writeLog("校验完成，错误详情请查看:{0}".format(errtxtFilePath))
        excelwb = load_workbook(errtxtFilePath)
        logSheet = excelwb['Sheet']
        logSheet["A{0}".format(1)] = dbCheckResult[0].cursor_description[0][0]
        logSheet["B{0}".format(1)] = dbCheckResult[0].cursor_description[1][0]
        logSheet["C{0}".format(1)] = dbCheckResult[0].cursor_description[2][0]
        logSheet["D{0}".format(1)] = dbCheckResult[0].cursor_description[3][0]
        logSheet["E{0}".format(1)] = dbCheckResult[0].cursor_description[4][0]
        logSheet["F{0}".format(1)] = dbCheckResult[0].cursor_description[5][0]
        # logSheet["G{0}".format(1)] = dbCheckResult[0].cursor_description[6][0]
        i = 2
        for r in dbCheckResult:
            logSheet["A{0}".format(i)] = r[0]
            logSheet["B{0}".format(i)] = r[1]
            logSheet["C{0}".format(i)] = r[2]
            logSheet["D{0}".format(i)] = r[3]
            logSheet["E{0}".format(i)] = r[4]
            logSheet["F{0}".format(i)] = r[5]
            # logSheet["G{0}".format(i)] = r[6]
            i += 1
        excelwb.save(errtxtFilePath)
        IsSucess = 0
    else:
        IsSucess = 1

	

   
	
if __name__=="__main__":
    import xlrd,xlwt
    import pandas as pd,numpy as np
    import numpy as np
    import os
    import time
    import sys
    import datetime
    import pymssql
    import pyodbc
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from dateutil.relativedelta import *
    from dateutil.parser import parse
    import re
    
    #path=r'\\172.16.6.143\StudentsProducts\1206新产品和受托报告\说明书\个人住房贷款\20组 工元宜居2019年第七期个人住房抵押贷款资产支持证券说明书'
    # filepath = str(sys.argv[1])
    # dateId = str(sys.argv[2])
    # userId = str(sys.argv[3])
    filepath = ''
    dateId = ''
    userId = ''
    TrustId = 10673
    dbConnectionStr="DRIVER={SQL Server};SERVER=172.16.6.143\MSSQL;DATABASE=TaskCollection;UID=sa;PWD=PasswordGS2017"

	
	
	
    errtxtFilePath = os.path.join('C:\\PyCharm\\pdf-docx\\source\\Example\\1','Error_基础表校验结果_{0}.xlsx'.format(dateId))

    if not os.path.exists(errtxtFilePath):
        wb = Workbook()  # 新建工作簿
        ws1 = wb.active
        wb.save(errtxtFilePath)
       # ErrorMessage('TrustId','TrustCode','TrustName','Path','ErrorInformation')
        wb = openpyxl.load_workbook(errtxtFilePath) # 读取xlsx文件
        ws=wb['Sheet']
        maxrow=ws.max_row
        ws.cell(maxrow,1,'负责人').value
        ws.cell(maxrow,2,'操作人').value
        ws.cell(maxrow,3,'TrustId').value
        ws.cell(maxrow,4,'TrustCode').value
        ws.cell(maxrow,5,'TrustName').value
        ws.cell(maxrow,6,'Path').value
        ws.cell(maxrow,7,'ErrorInformation').value
        ws.cell(maxrow,8,'备注').value
        ws.cell(maxrow,9,'Type').value
        wb.save(errtxtFilePath)
		
	

			
    Errorlist=[]
    Errorlist1=[]
    for root,dirs,files in os.walk(filepath):
        for name in files:
            if not name.endswith('基础表.xlsx'):
#                print(name,'文件名称不匹配!')
                continue
            ExceLFilePath = os.path.join(root,name)
            if ';' in name:
                SplitName=name.split(';')
                TrustCode=SplitName[0]
                TrustName=SplitName[1]
                print(TrustCode,TrustName)
                Trust(ExceLFilePath,TrustCode,TrustName)
                if V3==0:
                    TrustBond(ExceLFilePath)
                    #Errorlist.append(Error)
                    TrustExtension(ExceLFilePath)
                    #Errorlist1.append(Error1)
                    PrincipalSchedule(ExceLFilePath)
                    data=pd.read_excel(errtxtFilePath)
                    ckResultLen=len(data)
                    InsertTrusteeCheckByTrustId(userId, TrustId,ckResultLen)

                else:
                    Error=1
                    Errorlist.append(Error)
            else:
                Error=1
                Errorlist.append(Error)
                FileError='文件名称格式错误!应为[TrustCode;TrustName;基础表.xlsx]'
                ErrorMessage('*','*','*',ExceLFilePath,FileError)
                print(name,"--文件名称格式错误!应为[TrustCode;TrustName;基础表.xlsx]")

    dbCheckResult = runDBDataValidation(TrustId)
	



		
    print(sum(Errorlist),sum(Errorlist1))
    if sum(Errorlist)==0 and sum(Errorlist1)==0:
        IsSucess=1
        result='计划说明书基础表校验通过!'
        InsertVerificationLog(userId,9,IsSucess,errtxtFilePath,result)
    else:
        IsSucess=0
        result='计划说明书基础表校验失败,详细错误请下载查看!'
        InsertVerificationLog(userId,9,IsSucess,errtxtFilePath,result)
    
    
    
    
    