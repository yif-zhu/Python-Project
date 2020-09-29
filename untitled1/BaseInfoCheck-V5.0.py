# -*- coding: utf-8 -*-
"""
Created on Thu Dec 12 15:10:53 2019

@author: HUAWEI
"""

#%%

#TrustId=None
#TrustStartDate=None
#PoolCloseDate=None
#ClosureDate=None
#IssueAmount=0

#Error=0
#Error1=0
def ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,ErrorInformation):
    #ErrorPath=r"C:\Users\HUAWEI\Desktop\Error\TrustError.xlsx"
    wb = openpyxl.load_workbook(errtxtFilePath) # 读取xlsx文件
    sheets = wb.sheetnames
    ws=wb[sheets[0]]
    maxrow=ws.max_row
    maxrow+=1
#    print(maxrow)
    ws.cell(maxrow,3,TrustId).value
    ws.cell(maxrow,4,TrustCode).value
    ws.cell(maxrow,5,TrustName).value
    ws.cell(maxrow,6,Excelfilepath).value
    ws.cell(maxrow,7,ErrorInformation).value
    ws.cell(maxrow,9,1).value   
    wb.save(errtxtFilePath)
    
    
   
def Trust(Excelfilepath,TrustCode,TrustName):
#    print(Excelfilepath)
    global TrustId,TrustStartDate,PoolCloseDate,ClosureDate,IssueAmount,V3
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    try:
        selectId="select TrustId from TrustManagement.Trust where TrustCode='{}'".format(TrustCode)
        b1.execute(selectId)
        TrustId=b1.fetchone()[0]
    except:
        A0='文件TrustCode与系统TrustCoden不匹配请检查!'
        
        ErrorMessage('#',TrustCode,TrustName,Excelfilepath,A0)
        print('文件TrustCode与系统TrustCoden不匹配请检查!')
#        sys.exit(1)
        return
    TrustError=0
    V3=0
    DataTrust=pd.read_excel(Excelfilepath,sheet_name="Trust")
    TrustBond=pd.read_excel(Excelfilepath,sheet_name="TrustBond")
    CouponPaymentReferencelist=[]

    for i in TrustBond.index:
        ItemCode1=TrustBond.iloc[i][2]
        ItemValue1=TrustBond.iloc[i][3]
        if ItemCode1=='CouponPaymentReference':
            CouponPaymentReferencelist.append(ItemValue1)

    #获取名称
    try:
        TrustName=DataTrust.ItemValue[0]
    except:
        TrustError=1
        V3=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'模板错误,请检查!')
        print('模板错误，请检查!')
    
    if pd.isnull(TrustName)==True:
        TrustError=1
        TE0='产品名称缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,'产品名称缺失',Excelfilepath,TE0)
        print(TrustId,TrustName,TE0)    
    

    #获取产品简称
    TrustShortName=DataTrust.ItemValue[1]
    if pd.isnull(TrustShortName)==True:
        TrustError=1
        TE0='产品简称缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,'产品简称缺失',Excelfilepath,TE0)
        print(TrustId,TrustName,TE0) 
    
    
    #获取产品主体
    ProductSubject1=DataTrust.ItemValue[5]
    if pd.isnull(ProductSubject1)==True:
        TrustError=1
        TE0='产品主体缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE0)
        print(TrustId,TrustName,TE0)

    #获取发行金额
    IssueAmount=DataTrust.ItemValue[7]
    if type(IssueAmount) not in (int,float):
        TrustError=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'Trust表-发行金额应为数值型!')
        print('发行金额应为数值型!')
    
    if pd.isnull(IssueAmount)==True:
        TrustError=1
        TE1='Trust表-发行金额缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE1)
        print(TrustId,TrustName,TE1)
        
    #封包日资产池余额
    AssetPoolBalance=DataTrust.ItemValue[8]
    if type(AssetPoolBalance) not in (int,float):
        TrustError=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'Trust表-发行金额应为数值型!')
        print('封包日资产池余额应为数值型!')
        
    IsTopUpAvailable_Trust=DataTrust.ItemValue[10]
    if pd.isnull(IsTopUpAvailable_Trust)==True:
        TrustError=1
        print('Trust表-是否为循环购买产品填写为空，请检查!')
    else:
        if IsTopUpAvailable_Trust=='是':
            IsTopUpAvailable_Trust=1
        else:
            IsTopUpAvailable_Trust=0
        
        upIsTopUpAvailable_Trust="update TrustManagement.Trust set IsTopUpAvailable={} where TrustId={}".format(IsTopUpAvailable_Trust,TrustId)
        b1.execute(upIsTopUpAvailable_Trust)
        conn.commit()
        print('是否循环购买已更新!')
    
        
    
    
    if pd.isnull(AssetPoolBalance)==True:
        TE11='封包日资产池余额缺失,请检查!'
        TrustError=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE11)
        print(TrustId,TrustName,TE11)
        
    #获取设立日
    TrustStartDate=DataTrust.ItemValue[12]
    if pd.isnull(TrustStartDate)==True:
        TrustError=1
        TE2='设立日缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE2)
        print(TrustId,TrustName,TE2)
    
    
    #获取封包日
    PoolCloseDate=DataTrust.ItemValue[11]
    if pd.isnull(PoolCloseDate)==True:
        TrustError=1
        TE3='封包日缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE3)
        print(TrustId,TrustName,TE3)
    
    
    #获取法定到期日
    ClosureDate=DataTrust.ItemValue[13]
    if pd.isnull(ClosureDate)==True:
        TrustError=1
        TE4='法定到期日缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE4)
        print(TrustId,TrustName,TE4)
    #----------------------------------------------------------------------------------

    #-----------------------------------
    #牵头主承销
    MainUnderwriter=DataTrust.ItemValue[15]      
    
    if pd.isnull(MainUnderwriter)==True:
        TrustError=1
        TE6='主承销商缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE6)
        print(TrustId,TrustName,TE6)
    
    
    
    elif '/' in MainUnderwriter:
        deleteMainUnderwritersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='MainUnderwriter'".format(TrustId)
        b1.execute(deleteMainUnderwritersql)
        conn.commit()
        
        listMainUnderwriter=MainUnderwriter.split('/')
        for MainUnderwriter_f in listMainUnderwriter:
            AdditionalItemCode='MainUnderwriter'
            ChineseField='主承销商'
            AdditionalContent=MainUnderwriter_f
            insertMainUnderwriter="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
            b1.execute(insertMainUnderwriter)
            conn.commit()
            print('主承销插入完成!')
    elif '\\' in MainUnderwriter:
        deleteMainUnderwritersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='MainUnderwriter'".format(TrustId)
        b1.execute(deleteMainUnderwritersql)
        conn.commit()
        
        listMainUnderwriter=MainUnderwriter.split('\\')
        for MainUnderwriter_f in listMainUnderwriter:
            AdditionalItemCode='MainUnderwriter'
            ChineseField='主承销商'
            AdditionalContent=MainUnderwriter_f
            insertMainUnderwriter="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
            b1.execute(insertMainUnderwriter)
            conn.commit()
            print('主承销插入完成!')
            
            
    else:
        deleteMainUnderwritersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='MainUnderwriter'".format(TrustId)
        b1.execute(deleteMainUnderwritersql)
        conn.commit()
        
        AdditionalItemCode='MainUnderwriter'
        ChineseField='主承销商'
        AdditionalContent=MainUnderwriter
        insertMainUnderwriter="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertMainUnderwriter)
        conn.commit()
        print('主承销插入完成!')
        
    #----------------------------
    
    #资金保管机构
    FundsCustodian=DataTrust.ItemValue[16]
    if pd.isnull(FundsCustodian)==True:
        TrustError=1
        TE7='资金保管机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE7)
        print(TrustId,TrustName,TE7)
    
    else:
        deleteFundsCustodiansql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='FundsCustodian'".format(TrustId)
        b1.execute(deleteFundsCustodiansql)
        conn.commit()
        
        
        AdditionalItemCode='FundsCustodian'
        ChineseField='资金保管机构'
        AdditionalContent=FundsCustodian
        insertFundsCustodian="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertFundsCustodian)
        conn.commit()
        print('资金保管机构插入完成!')
    
    #----------------------------------
    #发行人
    Trustee=DataTrust.ItemValue[17]
    if pd.isnull(Trustee)==True:
        TrustError=1
        TE8='发行人/机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE8)
        print(TrustId,TrustName,TE8)
        
    
        
    else:
        
        deleteTrusteesql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='Trustee'".format(TrustId)
        b1.execute(deleteTrusteesql)
        conn.commit()
        
        AdditionalItemCode='Trustee'
        ChineseField='发行人/机构'
        AdditionalContent=Trustee
        insertTrustee="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertTrustee)
        conn.commit()
        print('发行人/机构插入完成!')
        
    #-----------------------------
     
    #资产服务机构
    ServiceInstitutions=DataTrust.ItemValue[18]
    if pd.isnull(ServiceInstitutions)==True:
        TrustError=1
        TE9='资产服务机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE9)
        print(TrustId,TrustName,TE9)
        
        
    else:
        deleteServiceInstitutionssql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='ServiceInstitutions'".format(TrustId)
        b1.execute(deleteServiceInstitutionssql)
        conn.commit()
        
        AdditionalItemCode='ServiceInstitutions'
        ChineseField='资产服务机构/贷款服务机构'
        AdditionalContent=ServiceInstitutions
        insertServiceInstitutions="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertServiceInstitutions)
        conn.commit()
        print('资产服务机构(贷款服务机构)插入完成!')

    #---------------------------------
    #受托机构
    EntrustOrganization=DataTrust.ItemValue[19]
    if pd.isnull(ServiceInstitutions)==True:

        TrustError=1
        TE10='受托机构缺失,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE10)
        print(TrustId,TrustName,TE10)
    
        
    
    else:
        
        deleteEntrustOrganizationsql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='EntrustOrganization'".format(TrustId)
        b1.execute(deleteEntrustOrganizationsql)
        conn.commit()
        
        AdditionalItemCode='EntrustOrganization'
        ChineseField='受托机构'
        AdditionalContent=ServiceInstitutions
        insertEntrustOrganization="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertEntrustOrganization)
        conn.commit()
        print('受托机构插入完成!')
		
		
		
   #---------------------
    #实际融资人/归集主体
    TheActualFinancier=DataTrust.ItemValue[22]
    if pd.isnull(TheActualFinancier)==True:
        TrustError=1
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'实际融资人/归集主体,不能为空!')
        print(TrustId,TrustName,'实际融资人/归集主体,不能为空!')
		
    elif '/' in TheActualFinancier:
        deleteTheActualFinancier="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='TheActualFinancier'".format(TrustId)
        b1.execute(deleteTheActualFinancier)
        conn.commit()
        
        listTheActualFinancier=TheActualFinancier.split('/')
        for TheActualFinancier_f in listTheActualFinancier:
            AdditionalItemCode='TheActualFinancier'
            ChineseField='实际融资人/归集主体'
            AdditionalContent=MainUnderwriter_f
            insertTheActualFinancier="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
            b1.execute(insertTheActualFinancier)
            conn.commit()
            print('实际融资人/归集主体插入完成!')
    elif '\\' in TheActualFinancier:
        deleteTheActualFinanciersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='TheActualFinancier'".format(TrustId)
        b1.execute(deleteTheActualFinanciersql)
        conn.commit()
        
        listTheActualFinancier=TheActualFinancier.split('/')
        for TheActualFinancier_f in listTheActualFinancier:
            AdditionalItemCode='TheActualFinancier'
            ChineseField='实际融资人/归集主体'
            AdditionalContent=MainUnderwriter_f
            insertTheActualFinancier="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
            b1.execute(insertTheActualFinancier)
            conn.commit()
            print('实际融资人/归集主体插入完成!')
            
            
    else:
        deleteMainUnderwritersql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='MainUnderwriter'".format(TrustId)
        b1.execute(deleteMainUnderwritersql)
        conn.commit()
        
        AdditionalItemCode='TheActualFinancier'
        ChineseField='实际融资人/归集主体'
        AdditionalContent=MainUnderwriter
        insertTheActualFinancier="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertTheActualFinancier)
        conn.commit()
        print('实际融资人/归集主体插入完成!') 
		
		
		
		
		
		
		
        
        #----------------------
    BookkeepingDate=DataTrust.ItemValue[20]
    print(BookkeepingDate)
    BookkeepingDate=str(BookkeepingDate).split(' ')[0]
    if pd.isnull(BookkeepingDate)==True:
        TrustError=1
        TE11='薄记建档日不能为空,请检查!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE10)
        print(TrustId,TrustName,TE11)
    else:
        try:
            
            BookkeepingDate=datetime.datetime.strptime(BookkeepingDate,'%Y-%m-%d')
            upBookkeepingDate="update TrustManagement.Trust set BookkeepingDate=N'{}' where TrustId={}".format(BookkeepingDate,TrustId)
            b1.execute(upBookkeepingDate)
            conn.commit()
            print('薄记建档日更新完成!')
            
        except:
            TrustError=1
            Tdata='建档日日期格式不规范!应为日期格式(YYYY-mm-dd)'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,Tdata)
            print(Tdata)
        
    #有浮动利率债券时Trust表基准利率不能为空
    ProductFeatures=DataTrust.ItemValue[21]
    
    if pd.isnull(ServiceInstitutions)==True and '浮动利率' in CouponPaymentReferencelist:
        TrustError=1
        TE22='有浮动利率债券时Trust表基准利率不能为空'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,TE22)

    else:
        upProductFeatures="update TrustManagement.Trust set ProductFeatures=N'{}' where TrustId={}".format(ProductFeatures,TrustId)
        b1.execute(upProductFeatures)
        conn.commit()
        print('基准利率更新完成!')

    
    
    #---------------------------
	#插入产品主体
    if pd.isnull(ProductSubject1)==False:

        deleteServiceInstitutionssql="delete DV.AdditionalInformation where TrustId={} and AdditionalItemCode='ProductSubject'".format(TrustId)
        b1.execute(deleteServiceInstitutionssql)
        conn.commit()

        AdditionalItemCode='ProductSubject'
        ChineseField='原始权益人\产品主体'
        AdditionalContent=ProductSubject1
        insertServiceInstitutions="insert into DV.AdditionalInformation Values({},'{}',N'{}',N'{}')".format(TrustId,AdditionalItemCode,ChineseField,AdditionalContent)
        b1.execute(insertServiceInstitutions)
        conn.commit()
        print('原始权益人\产品主体插入完成!')
	

   
    #更新发行金额
    if pd.isnull(IssueAmount)==False and type(IssueAmount)!=str:
        UpdateIssueAmount="update TrustManagement.Trust set IssueAmount={} where TrustId={}".format(IssueAmount,TrustId)
        b1.execute(UpdateIssueAmount)
        conn.commit()
        T1="{},{},'发行金额更新完成!'".format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T1)
        print(T1) 

            
    #更新资产池封包日余额
    if pd.isnull(AssetPoolBalance)==False  and type(AssetPoolBalance)!=str:
#        if SQLIAssetPoolBalance==None:
        UpdateAssetPoolBalance="update TrustManagement.Trust set AssetPoolBalance={} where TrustId={}".format(AssetPoolBalance,TrustId)
        b1.execute(UpdateAssetPoolBalance)
        conn.commit()
        T2="{},{},'封包日资产池余额更新完成!'".format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T1)
        print(T2) 

    MarketCategoryType=DataTrust.ItemValue[3]
    try:
        selectId="SELECT ISNULL(ItemValue,'MarketCategoryIsNull') AS MarketCategory  FROM TrustManagement.TrustInfoExtension WHERE TrustId ={} AND ItemCode='MarketCategory'".format(TrustId)
        b1.execute(selectId)
        MarketCategory=b1.fetchone()[0]
        selectId="SELECT  ISNULL(ItemCode,'AssetTypeIsNull')  AS AssetType  FROM DV.Item WHERE CategoryCode='AssetType' AND ItemTitle=N'{}'".format(MarketCategoryType)
        b1.execute(selectId)
        AssetType=b1.fetchone()[0]
        if MarketCategory=='ABN' and AssetType=='AssetTypeIsNull':
            TrustError=1
            Tdata='资产类型在数据库不存在，请联系管理！'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,Tdata)
            print(Tdata)
        if MarketCategory=='ABN' and AssetType!='AssetTypeIsNull':
            print(MarketCategory,MarketCategoryType,AssetType)
            sql=" MERGE TrustManagement.TrustInfoExtension AS a USING(SELECT {} AS TrustID,'{}' AS ItemCode) AS source ".format(TrustId,'AssetType')
            sql+=" ON source.TrustId=a.TrustID  AND source.ItemCode=a.ItemCode  WHEN MATCHED THEN UPDATE SET ItemValue ='{}' ".format(AssetType)
            sql+=" WHEN NOT MATCHED THEN INSERT ( TrustId, StartDate,EndDate,ItemId,ItemCode,ItemValue) VALUES ( {}, CONVERT(DATE,GETDATE()) ,null,13171, 'AssetType',N'{}');".format(TrustId,AssetType)
            print(sql)
            b1.execute(sql)
            conn.commit()
            T2="{},{},'资产支持票据市场分类更新完成!'".format(TrustId,TrustName)
    except: 
        pass 
        return

            
        
def TrustBond(Excelfilepath,errtxtFilePath):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    global Error
    Error=0
    DataTrustBond1=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
    DataPrincipalSchedule=pd.read_excel(Excelfilepath,sheet_name='PrincipalSchedule')
    
    huanben=DataPrincipalSchedule.loc[2][2]
    print(huanben)

    try:
        DataTrustBond=DataTrustBond1[['TrustBondID','ItemID','ItemCode','ItemValue']]
        TrustBondID=DataTrustBond['TrustBondID'].drop_duplicates()

    except:
        T2="TrustBond表字段错误"
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T2)
        print('TrustBond字段错误!')
        return 
    
    DataTrustBond=DataTrustBond.dropna(subset=['ItemCode','ItemID'])
    if DataTrustBond['TrustBondID'].count()>25 and DataTrustBond['TrustBondID'].sum()==0:
        A1="TrustBond表分层标识(TrustBondID)重复'"
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A1)
        print(TrustId,TrustName,'分层标识重复(TrustBondID)')
        Error=1
#    AmountC=0
#    
#    ZClass=0
#    #各层级产品简称
#    ShortNamelist=[]
#    #各层级发行规模
#    OfferAmountlist=[]
#    #获取各层级起息日
#    IssueDatelist=[]
    for i in DataTrustBond.index:       
        TrustBondId=DataTrustBond.loc[i][0]
        ItemId=int(DataTrustBond.loc[i][1])
        ItemCode=DataTrustBond.loc[i][2]
        ItemValue=DataTrustBond.loc[i][3]
        
        

        #必填项不能为空
        if ItemCode in ('ShortName','SecurityExchangeCode','CurrencyOfIssuance','Denomination','LegalMaturityDate','InterestPaymentType','PaymentConvention','IssueDate','InterestRateCalculation','InterestDays','CouponBasis','CouponPaymentReference','RatingAgent','OriginalCreditRating','ClassName') and pd.isnull(ItemValue)==True:
            Error=1
            AJJ2='{}不能为空!'.format(ItemCode)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ2)
            print(TrustId,TrustName,AJJ2)
            
            
     
        if ItemCode=='OfferAmount' and type(ItemValue) not in (int,float):
            Error=1
            A7="{},发行规模(OfferAmount)数据类型应该为int、float".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A7)
            print(TrustId,TrustName,A7)
            
        if ItemCode=='OfferAmount' and pd.isna(ItemValue)==True:
            Error=1
            A17="{},发行规模(OfferAmount)不能为空".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A17)
            print(TrustId,TrustName,A17)
            
        
        #判断预计到期日是否为日期格式
        try:
            
            if ItemCode=='LegalMaturityDate':
                #print(ItemValue)
                ItemValue=str(ItemValue)
                ItemValue=datetime.datetime.strptime(ItemValue,"%Y-%m-%d %H:%M:%S")
        except:
            Error=1
            A18="{},预计到期日(LegalMaturityDate)应为日期格式(YYYY-mm-dd),且不能为空!".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A18)
            print(TrustId,TrustName,A18)    

        
            
    
    
    
    
    #查询是否有备注信息
    ErroeMsglist=[]
    selectNone="select ErroeMsg from dbo.specification_note where TrustId={} and PaymentPeriod=0 and PointType=1".format(TrustId)
    b1.execute(selectNone)
    ErroeMsg=b1.fetchall()
    
    for E in ErroeMsg:
        ErroeMsglist.append(E[0])
    print(ErroeMsglist)
    ErrorDate=pd.read_excel(errtxtFilePath)
    
    for error in ErrorDate.index:
        
        if ErrorDate.iloc[error][6] in ErroeMsglist:
            print(ErrorDate.iloc[error][6])
            ErrorDate=ErrorDate.drop(index=[error])
            
    if len(ErrorDate['ErrorInformation'])==0:
        Error=0
        print(Error)
    else:
        Error=1
    ErrorDate.to_excel(errtxtFilePath,index=False)
    
    print('Error:',Error)
    if Error==0:
#        A15='TrustBond表校验通过!'
        #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A15)
        print(TrustId,TrustName,'TrustBond表校验通过!')
        TrustBondImport(Excelfilepath,TrustId,TrustName)

        
        


def TrustExtension(Excelfilepath,errtxtFilePath):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    global Error1
    Error1=0
    DataTrustExtension=pd.read_excel(Excelfilepath,sheet_name='TrustExtension')
    try:
        DataTrustExtension=DataTrustExtension[['ItemCode','ItemValue']]
    except:
        Error1=1
        B000='TrustExtension表字段错误!'
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B000)
        print(TrustId,TrustName,B000)
        return
            
    for Extensionindex in DataTrustExtension.index:
        ItemCode=DataTrustExtension.loc[Extensionindex][0]
        ItemValue=DataTrustExtension.loc[Extensionindex][1]
         
            
        try:
                #取计息日
            date=1
            if ItemCode=='B_InterestCollectionDate_FirstDate' and DataTrustExtension.ItemValue[3]=='InterestCollectionDate':
                            
                B_InterestCollectionDate_FirstDate=str(ItemValue)
                B_InterestCollectionDate_FirstDate=B_InterestCollectionDate_FirstDate.split(' ')[0]
                B_InterestCollectionDate_FirstDate=datetime.datetime.strptime(B_InterestCollectionDate_FirstDate,"%Y-%m-%d")
#                print(B_PaymentDate_FirstDate)
                date=0
        except:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个计息日应为日期格式（YYYY-mm-dd）')
            Error1=1
            date=1
            print('计息日应为日期格式')
            print(date)
        
        try:
            
                #取计算日
            if ItemCode=='B_CollectionDate_FirstDate':
                B_CollectionDate_FirstDate=str(ItemValue)
                B_CollectionDate_FirstDate=B_CollectionDate_FirstDate.split(' ')[0]
                B_CollectionDate_FirstDate=datetime.datetime.strptime(B_CollectionDate_FirstDate,"%Y-%m-%d")
                print(B_CollectionDate_FirstDate)
                date1=0
        except:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个计算日应为日期格式（YYYY-mm-dd）')
            Error1=1
            date1=1
            print('计算日应为日期格式')
            print(date)      
            
        try:
                #取兑付日
            if ItemCode=='B_PaymentDate_FirstDate':
                B_PaymentDate_FirstDate=str(ItemValue)
                B_PaymentDate_FirstDate=B_PaymentDate_FirstDate.split(' ')[0]
                B_PaymentDate_FirstDate=datetime.datetime.strptime(B_PaymentDate_FirstDate,"%Y-%m-%d")
#                print(B_InterestCollectionDate_FirstDate)
                date2=0
        except:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'第一个兑付日应为日期格式（YYYY-mm-dd）')
            Error1=1            
            date2=1
            print('兑付日应为日期格式')
            print(date)
        
 
            
        #查询是否有备注信息
    ErroeMsglist1=[]
    selectNone1="select ErroeMsg from dbo.specification_note where TrustId={} and PaymentPeriod=0 and PointType=1".format(TrustId)
    b1.execute(selectNone1)
    ErroeMsg1=b1.fetchall()
    
    for Er in ErroeMsg1:
        ErroeMsglist1.append(Er[0])
        
    ErrorDate1=pd.read_excel(errtxtFilePath)
    
    for error1 in ErrorDate1.index:
        if ErrorDate1.iloc[error1][6] in ErroeMsglist1:
            ErrorDate1.drop(index=[error1])
            
    if len(ErrorDate1['ErrorInformation'])==0:
        Error1=0
        print(Error1)
    ErrorDate1.to_excel(errtxtFilePath,index=False)
    
                
    print('Error1:',Error1)
    Errorlist1.append(Error1)
    if Error1==0:
        B13='TrustExtension表校验通过!'
        #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,B13)
        print(TrustId,TrustName,B13)
        TrustExtensionImport(Excelfilepath)


            

def TrustBondImport(Excelfilepath,TrustId,TrustName):
    
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    deletepjsql="delete DV.RatingAgency where TrustId={} and RatingAgencyCode in ('RatingAgent','RatingAgent-1','OriginalCreditRating','OriginalCreditRating-1','ClassName','ClassName-1')".format(TrustId)
    b1.execute(deletepjsql)
    conn.commit()
    
     #将评级插入RatingAgency
    imRatingAgencydata=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
    imRatingAgencydata=imRatingAgencydata[['TrustBondID','ItemID','ItemCode','ItemValue']]
    for Agency in imRatingAgencydata.index:
        TrustBondId=imRatingAgencydata.loc[Agency][0]
#        ItemId=int(imRatingAgencydata.loc[Agency][1])
        ItemCode=imRatingAgencydata.loc[Agency][2]
        ItemValue=imRatingAgencydata.loc[Agency][3]
        
        if ItemCode in ('RatingAgent','RatingAgent-1'):
            agencyname='评级机构'
            insertRatingAgency="insert into DV.RatingAgency values({},{},'{}',N'{}',N'{}')".format(TrustId,TrustBondId,ItemCode,agencyname,ItemValue)
            b1.execute(insertRatingAgency)
            conn.commit()
            print('证券机构插入完成!')
        if ItemCode in ('OriginalCreditRating','OriginalCreditRating-1'):
            agencyname='原始证券评级'
            insertRatingAgency="insert into DV.RatingAgency values({},{},'{}',N'{}',N'{}')".format(TrustId,TrustBondId,ItemCode,agencyname,ItemValue)
            b1.execute(insertRatingAgency)
            conn.commit()
            print('原始评级插入完成')
        if ItemCode in ('ClassName','ClassName-1'):
            agencyname='当前证券评级'
            insertRatingAgency="insert into DV.RatingAgency values({},{},'{}',N'{}',N'{}')".format(TrustId,TrustBondId,ItemCode,agencyname,ItemValue)
            b1.execute(insertRatingAgency)
            conn.commit()
            print('当前评级插入完成!')
            
        #删除上次插入的TrustBond数据
    deleteTrustBond="delete TrustManagement.TrustBond2 where TrustId={}".format(TrustId)
    b1.execute(deleteTrustBond)
    conn.commit() 
    print('历史数据已清除')
            

    DataTrustBond=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
    DataTrustBond=DataTrustBond[['TrustBondID','ItemID','ItemCode','ItemValue']]
    DataTrustBond.dropna(subset=['ItemCode'],inplace=True)
    DataTrustBond.dropna(subset=['ItemID'],inplace=True)
#        del DataTrustBond['Unnamed: 5']
#        del DataTrustBond['StartDate']
#        del DataTrustBond['EndDate']
    for i in DataTrustBond.index:
        TrustBondId=DataTrustBond.loc[i][0]
        ItemId=int(DataTrustBond.loc[i][1])
        ItemCode=DataTrustBond.loc[i][2]
        ItemValue=DataTrustBond.loc[i][3]
#            print(TrustBondId,ItemId,ItemCode,ItemValue)  
        if ItemCode=='CouponBasis' and ItemValue<1:
            ItemValue=ItemValue*100
        InsertBond="insert into TrustManagement.TrustBond2(TrustBondId,TrustId,StartDate,EndDate,ItemId,ItemCode,ItemValue) values({},{},GETDATE(),NULL,{},'{}',N'{}')".format(TrustBondId,TrustId,ItemId,ItemCode,ItemValue)
    #    print(sql)
        b1.execute(InsertBond)
    conn.commit()
    BI='TrustBond导入完成!'
    #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,BI)
    print(TrustId,TrustName,'TrustBond导入完成!')
        
    #调用导入还本计划函数
    PrincipalSchedule(ExceLFilePath)
        #临时表推正式表

     
        
def TrustExtensionImport(Excelfilepath):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()

    deleteTrustExtension="delete TrustManagement.TrustExtension1 where TrustId={}".format(TrustId)
    b1.execute(deleteTrustExtension)
    conn.commit()
    
    TrustExtension=pd.read_excel(Excelfilepath,sheet_name='TrustExtension')
    TrustExtension=TrustExtension[['ItemCode','ItemValue','对三种日期的文字描述']]
    #TrustExtension=TrustExtension[1:]
    
    if str(TrustExtension.ItemValue[4])=='False':
        index=32
    else:
        index=60
#    print(TrustExtension.ItemValue[4])
#    print(index)
#    print(TrustExtension.ItemCode[index])

    for tindex in range(0,index):
        TrustExtension_ItemCode=TrustExtension.iloc[tindex][0]
        TrustExtension_ItemValue=TrustExtension.iloc[tindex][1]
        if TrustExtension_ItemCode=='B_InterestCollectionDate_Canlendar':
            TrustExtension_ItemCode='B_InterestCollectionDate_Calendar'
        if TrustExtension_ItemCode=='B_CollectionDate_Canlendar':
            TrustExtension_ItemCode='B_CollectionDate_Calendar'
        if TrustExtension_ItemCode=='B_PaymentDate_Canlendar':
            TrustExtension_ItemCode='B_PaymentDate_Calendar'
            #转化Bool
        if TrustExtension_ItemCode=='B_InterestCollectionDate_Frequency' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
            #if Acceleratedliquidation=='null':
            #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'文字描述部分不能有空，请检查！')
        if TrustExtension_ItemCode=='B_CollectionDate_Frequency':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_PaymentDate_Frequency':
                
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_InterestCollectionDate_Frequency' and TrustExtension.ItemValue[3]=='InterestCollectionDate' and index>32:
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_CollectionDate_Frequency':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='R_PaymentDate_Frequency':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
            
            #----
        if TrustExtension_ItemCode=='B_InterestCollectionDate_WorkingDateAdjustment' and TrustExtension.ItemValue[3]=='InterestCollectionDate':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_CollectionDate_WorkingDateAdjustment':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)
        if TrustExtension_ItemCode=='B_PaymentDate_WorkingDateAdjustment':
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)


                
        if TrustExtension_ItemCode=='RevolvingPeriod' and pd.isnull(TrustExtension_ItemValue)==False:
            TrustExtension_ItemValue=int(TrustExtension_ItemValue)            
            
            
        Itemid_sql="select ItemId from TrustManagement.Item where ItemCode='{}'".format(TrustExtension_ItemCode)
        b1.execute(Itemid_sql)
        ItemId=b1.fetchone()[0]

        TrustExtensionInsert="insert into TrustManagement.TrustExtension1(TrustId,StartDate,EndDate,ItemId,ItemCode,ItemValue) values({},GETDATE(),NULL,{},N'{}',N'{}')".format(TrustId,ItemId,TrustExtension_ItemCode,TrustExtension_ItemValue)
        b1.execute(TrustExtensionInsert)
        Description_ItemValue=None if pd.isnull(TrustExtension.iloc[tindex][2])==True else TrustExtension.iloc[tindex][2]
        if Description_ItemValue ==None and tindex== int(tindex/10+1)*9-3:
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'计算日、计息日、支付日的具体文字描述不能为空，请检查！')
            print('计算日、计息日、支付日的具体文字描述不能为空，请检查！')
            Error1=1
        if Description_ItemValue !=None:
            TrustExtensionInsert="insert into TrustManagement.TrustExtension1(TrustId,StartDate,EndDate,ItemId,ItemCode,ItemValue) values({},GETDATE(),NULL,{},N'{}',N'{}')".format(TrustId,ItemId,TrustExtension_ItemCode+"Description",Description_ItemValue)
            b1.execute(TrustExtensionInsert)
    conn.commit()
    EI='TrustExtension导入完成!'
    #ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,EI)
    print(TrustId,TrustName,EI)
        
        
  
        
def PrincipalSchedule(Excelfilepath):
    global huanben
    
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    deletePrincipalSchedule="delete dbo.PrincipalSchedule1 where TrustId={}".format(TrustId)
    b1.execute(deletePrincipalSchedule)
    conn.commit()
    
    DataPrincipalSchedule=pd.read_excel(Excelfilepath,sheet_name='PrincipalSchedule')
    huanben=DataPrincipalSchedule.loc[2][2]
    huanben1=DataPrincipalSchedule.loc[3][2]
    
    if pd.isnull(huanben)==True and pd.isnull(huanben1)==True:
#        print(huanben)
        print('无还本计划!')
    else:
#        DataPrincipalSchedule=pd.read_excel(Excelfilepath,sheet_name='PrincipalSchedule',header=None)
        DataPrincipalSchedule=DataPrincipalSchedule.iloc[:,:3]
#        print(DataPrincipalSchedule)
        for Prindex in DataPrincipalSchedule.index:
            TrustBondID=DataPrincipalSchedule.loc[Prindex][0]
            Pdate=DataPrincipalSchedule.loc[Prindex][1]
            
            DebtPlan=DataPrincipalSchedule.loc[Prindex][2]
            if TrustBondID=='0.0':
                TrustBondID=0
            InsertPrincipalSchedule="insert into dbo.PrincipalSchedule1 values({},{},'{}','{}')".format(TrustId,TrustBondID,Pdate,DebtPlan)
            print(InsertPrincipalSchedule)
            b1.execute(InsertPrincipalSchedule)
        conn.commit()
        print('还本计划导入完成!')
       
       
def TermSensitivityAnalysis(Excelfilepath):

    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    print("期限敏感性分析入库")
    deleteTermSensitivityAnalysis="delete DV.TermSensitivityAnalysis where TrustId={}".format(TrustId)
    b1.execute(deleteTermSensitivityAnalysis)
    conn.commit()
    SqlTrustInfoExtension="  IF NOT EXISTS(SELECT TOP 1  ISNULL(StartDate,GETDATE()) AS StartDate  FROM TrustManagement.TrustInfoExtension WHERE  TrustId={})SELECT GETDATE()  AS StartDate ELSE SELECT TOP 1  ISNULL(StartDate,GETDATE()) AS StartDate  FROM TrustManagement.TrustInfoExtension WHERE  TrustId={}".format(TrustId,TrustId)
    b1.execute(SqlTrustInfoExtension)
    StartDate=b1.fetchone()[0]
    # 获取sheet的汇总数据
    excel= xlrd.open_workbook(Excelfilepath)
    sheet = excel.sheet_by_name("期限敏感性分析")
    nrows= sheet.nrows
    ncols=sheet.ncols
    TermSensitivityAnalysisData=pd.read_excel(Excelfilepath,sheet_name='期限敏感性分析')
    global Rows
    Rows=0
    global Error2
    Error2=0
    for i in range(7,16): #7-16
        for j in range(0,ncols): #列
            TrustBondId='' if pd.isnull(TermSensitivityAnalysisData.iloc[6][j])==True else TermSensitivityAnalysisData.iloc[6][j]
            ItemCode="违约率假设"
            ItemValue=None if pd.isnull(TermSensitivityAnalysisData.iloc[i][j])==True else TermSensitivityAnalysisData.iloc[i][j]
            if ItemCode!= None and ItemValue != None:
                if is_number(TrustBondId)!=True:
                    TrustBondId=-1
                TrustBondId=int(TrustBondId)
                ItemCode=str(ItemCode)
                InsertTermSensitivityAnalysis="insert into DV.TermSensitivityAnalysis values({},{},'{}',{},{},N'{}',N'{}')".format(TrustBondId,TrustId,StartDate,i,j,ItemCode,ItemValue)
                b1.execute(InsertTermSensitivityAnalysis)


    for i in range(21,TermSensitivityAnalysisData.shape[0]): #21-28
        for j in range(0,ncols): #列
            TrustBondId='' if pd.isnull(TermSensitivityAnalysisData.iloc[20][j])==True else TermSensitivityAnalysisData.iloc[20][j]
            ItemCode="早偿率假设"
            print(i,j)
            ItemValue=None if pd.isnull(TermSensitivityAnalysisData.iloc[i][j])==True else TermSensitivityAnalysisData.iloc[i][j]
            if ItemCode!= None and ItemValue != None:
                if is_number(TrustBondId)!=True:
                    TrustBondId=-1
                TrustBondId=int(TrustBondId)
                ItemCode=str(ItemCode)
                InsertTermSensitivityAnalysis="insert into DV.TermSensitivityAnalysis values({},{},'{}',{},{},N'{}',N'{}')".format(TrustBondId,TrustId,StartDate,i,j,ItemCode,ItemValue)
                b1.execute(InsertTermSensitivityAnalysis)

    conn.commit()
    print('期限敏感性分析导入完成!')

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
 
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
 
    return False        
        
def ProductEventDetails(Excelfilepath):

    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    print("文字描述开始入库")
    # 获取sheet的汇总数据
    excel= xlrd.open_workbook(Excelfilepath)
    sheet = excel.sheet_by_name("文字描述")
    nrows= sheet.nrows
    ncols=sheet.ncols
    ProductEventDetails=pd.read_excel(Excelfilepath,sheet_name='文字描述')
    global Rows
    Rows=0
    global Error3
    Error3=0
    Acceleratedliquidation='null' if pd.isnull(ProductEventDetails.iloc[2][0])==True else ProductEventDetails.iloc[2][0]
    ContractBreach='null' if pd.isnull(ProductEventDetails.iloc[14][0])==True else ProductEventDetails.iloc[14][0]
    ClearanceRepo='null' if pd.isnull(ProductEventDetails.iloc[26][0])==True else ProductEventDetails.iloc[26][0]
    AnalysisOfcreditRatingAgencies='null' if pd.isnull(ProductEventDetails.iloc[38][0])==True else ProductEventDetails.iloc[38][0]
    ContractBreachbefore='null' if pd.isnull(ProductEventDetails.iloc[51][0])==True else ProductEventDetails.iloc[51][0]
    ContractBreachRear='null' if pd.isnull(ProductEventDetails.iloc[64][0])==True else ProductEventDetails.iloc[64][0]
    if Acceleratedliquidation=='null' or  ContractBreach=='null' or ClearanceRepo=='null'or  AnalysisOfcreditRatingAgencies=='null' or ContractBreachbefore=='null'or    ContractBreachRear=='null':
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,'文字描述部分不能有空，请检查！')
        Error1=1
    sql="MERGE DV.ProductEventDetails AS a USING(SELECT {} AS TrustId, N'{}' AS TrustName ) AS source".format(TrustId,TrustName)
    sql+=" ON source.TrustId=a.TrustID AND source.TrustName=a.TrustName  WHEN MATCHED THEN UPDATE SET TrustName =N'{}'".format(TrustName)
    sql+=" WHEN NOT MATCHED THEN INSERT (TrustID,  TrustName ) VALUES ( {}, N'{}' );".format(TrustId,  TrustName)
    b1.execute(sql)
    conn.commit() 
    sql="MERGE DV.ProductEventDetails AS a USING(SELECT {} AS TrustId, N'{}' AS TrustName ) AS source".format(TrustId,TrustName)
    sql+=" ON source.TrustId=a.TrustID AND source.TrustName=a.TrustName  WHEN MATCHED THEN UPDATE SET Acceleratedliquidation =N'{}'".format(Acceleratedliquidation)
    sql+=" WHEN NOT MATCHED THEN INSERT (TrustID, TrustName,Acceleratedliquidation) VALUES ( {}, N'{}',N'{}' );".format(TrustId,TrustName,Acceleratedliquidation)
    b1.execute(sql)
    conn.commit() 
    sql="MERGE DV.ProductEventDetails AS a USING(SELECT {} AS TrustId, N'{}' AS TrustName ) AS source".format(TrustId,TrustName)
    sql+=" ON source.TrustId=a.TrustID AND source.TrustName=a.TrustName  WHEN MATCHED THEN UPDATE SET ContractBreach =N'{}'".format(ContractBreach)
    sql+=" WHEN NOT MATCHED THEN INSERT (TrustID, TrustName,ContractBreach) VALUES ( {}, N'{}',N'{}' );".format(TrustId,TrustName,ContractBreach)
    b1.execute(sql)
    conn.commit() 
    sql="MERGE DV.ProductEventDetails AS a USING(SELECT {} AS TrustId,N'{}' AS TrustName ) AS source".format(TrustId,TrustName)
    sql+=" ON source.TrustId=a.TrustID AND source.TrustName=a.TrustName  WHEN MATCHED THEN UPDATE SET ClearanceRepo =N'{}'".format(ClearanceRepo)
    sql+=" WHEN NOT MATCHED THEN INSERT (TrustID, TrustName,ClearanceRepo) VALUES ( {}, N'{}',N'{}' );".format(TrustId,TrustName,ClearanceRepo)
    b1.execute(sql)
    conn.commit() 
    sql="MERGE DV.ProductEventDetails AS a USING(SELECT {} AS TrustId, N'{}' AS TrustName ) AS source".format(TrustId,TrustName)
    sql+=" ON source.TrustId=a.TrustID AND source.TrustName=a.TrustName  WHEN MATCHED THEN UPDATE SET AnalysisOfcreditRatingAgencies =N'{}'".format(AnalysisOfcreditRatingAgencies)
    sql+=" WHEN NOT MATCHED THEN INSERT (TrustID, TrustName,AnalysisOfcreditRatingAgencies) VALUES ( {}, N'{}',N'{}' );".format(TrustId,TrustName,AnalysisOfcreditRatingAgencies)
    b1.execute(sql)
    conn.commit() 
    sql="MERGE DV.ProductEventDetails AS a USING(SELECT {} AS TrustId, N'{}' AS TrustName ) AS source".format(TrustId,TrustName)
    sql+=" ON source.TrustId=a.TrustID AND source.TrustName=a.TrustName  WHEN MATCHED THEN UPDATE SET ContractBreachbefore =N'{}'".format(ContractBreachbefore)
    sql+=" WHEN NOT MATCHED THEN INSERT (TrustID, TrustName,ContractBreachbefore) VALUES ( {}, N'{}',N'{}' );".format(TrustId,TrustName,ContractBreachbefore)
    b1.execute(sql)
    conn.commit() 
    sql="MERGE DV.ProductEventDetails AS a USING(SELECT {} AS TrustId, N'{}' AS TrustName ) AS source".format(TrustId,TrustName)
    sql+=" ON source.TrustId=a.TrustID AND source.TrustName=a.TrustName  WHEN MATCHED THEN UPDATE SET ContractBreachRear =N'{}'".format(ContractBreachRear)
    sql+=" WHEN NOT MATCHED THEN INSERT (TrustID, TrustName,ContractBreachRear) VALUES ( {}, N'{}',N'{}' );".format(TrustId,TrustName,ContractBreachRear)
    b1.execute(sql)
    conn.commit() 
    print('文字描述导入完成!')
       
        
def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:

        print(str(ex))
        # raise ex
    finally:
        cnxn.close()
        
    
        
def InsertVerificationLog(userId, CheckType, IsSucess, filePath, result,filepath):
    sql = "exec TaskCollection.dbo.usp_InsertVerificationLog N'{0}',N'{1}',N'{2}',N'{3}',N'{4}',N'{5}'".format(userId,CheckType,IsSucess, filePath, result,filepath)
    execSQLCmd(sql) 
    
def InsertTrusteeCheckByTrustId(userId, TrustId,ckResultLen):
    sql = "exec TaskCollection.dbo.usp_InsertTrusteeCheckByTrustId N'{0}',N'{1}',N'{2}',N'{3}',N'{4}'".format(userId,9, TrustId, 0, ckResultLen)
    execSQLCmd(sql)

#更新任务状态
def UpdateProductStatus(TrustId):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='TaskCollection', charset='utf8')
    b1=conn.cursor()
    
    SelectTrustDocumentID="select TrustDocumentID from TaskCollection.[dbo].[view_TrusteeReportForDocumentId] where TrustId={} and Period=0 and FileType='BasicInformation'".format(TrustId)
    b1.execute(SelectTrustDocumentID)
    TrustDocumentID=b1.fetchone()[0]
#    print(TrustDocumentID)
    execsql="[TaskCollection].[dbo].[UpdateProductStatus] {},{},'{}',1".format(TrustId,TrustDocumentID,'BasicInformation')
#    print(execsql)
    b1.execute(execsql)
    conn.commit()

           
def execSQLCmdFetchAll(sql):
    print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        print(str(ex))
    finally:
        cnxn.close()        
        
def runDBDataValidation(TrustId):
    sql = "exec TaskCollection.dbo.usp_VerifyTrustServiceBasicByTrust N'{0}'".format(TrustId)
    dbCheckResult = execSQLCmdFetchAll(sql)
    selectId = "SELECT top 1  TrustId FROM PortfolioManagement.[DV].[VerifyTrustServiceBasicRemarksLog] WHERE TrustId= {}".format(TrustId)
    cnxn = pyodbc.connect(dbConnectionStr)
    b1 = cnxn.cursor()
    b1.execute(selectId)
    Result = b1.fetchone()[0]
    return Result 


   
    
if __name__=="__main__":
    import xlrd,xlwt
    import pandas as pd,numpy as np
    import numpy as np
    import os
    import time
    import sys
    import datetime
    import pymssql
    import pyodbc
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from dateutil.relativedelta import *
    from dateutil.parser import parse
    import re
    
    filepath=r'\\172.16.6.143\Products\计划说明书\202008\ABS\基础表校验-V5\基础表\卢迪\周帅\杭盈2020年第二期个人住房抵押贷款资产支持证券'
    # filepath = str(sys.argv[1])
    # dateId = str(sys.argv[2])
    # userId = str(sys.argv[3])
    IsResultSucess = 0
    dbConnectionStr="DRIVER={SQL Server};SERVER=172.16.6.143\MSSQL;DATABASE=TaskCollection;UID=sa;PWD=PasswordGS2017"

    
    
    
    errtxtFilePath = os.path.join(r'C:\Users\DELL\Desktop\华能信托·惠橙6号集合资金信托计划','Error_基础表校验结果_{0}.xlsx'.format('100000'))
    #
    if not os.path.exists(errtxtFilePath):
        wb = Workbook()  # 新建工作簿
        ws1 = wb.active
        wb.save(errtxtFilePath)
       # ErrorMessage('TrustId','TrustCode','TrustName','Path','ErrorInformation')
        wb = openpyxl.load_workbook(errtxtFilePath) # 读取xlsx文件
        ws=wb['Sheet']
        maxrow=ws.max_row
        ws.cell(maxrow,1,'负责人').value
        ws.cell(maxrow,2,'操作人').value
        ws.cell(maxrow,3,'TrustId').value
        ws.cell(maxrow,4,'TrustCode').value
        ws.cell(maxrow,5,'TrustName').value
        ws.cell(maxrow,6,'Path').value
        ws.cell(maxrow,7,'ErrorInformation').value
        ws.cell(maxrow,8,'备注').value
        ws.cell(maxrow,9,'Type').value
        wb.save(errtxtFilePath)
        
    

            
    Errorlist=[]
    Errorlist1=[]
    for root,dirs,files in os.walk(filepath):
        for name in files:
            if not name.endswith('基础表.xlsx'):
                continue
            ExceLFilePath = os.path.join(root,name)
            if ';' in name:
                SplitName=name.split(';')
                TrustCode=SplitName[0]
                TrustName=SplitName[1]
                print(TrustCode,TrustName)
                Trust(ExceLFilePath,TrustCode,TrustName)
                V3=0
                if V3==0:
                    # TrustBond(ExceLFilePath,errtxtFilePath)
                    # Errorlist.append(Error)
                    # TrustExtension(ExceLFilePath,errtxtFilePath)
                    # Errorlist1.append(Error1)
                    # PrincipalSchedule(ExceLFilePath)
                    # data=pd.read_excel(errtxtFilePath)
                    TermSensitivityAnalysis(ExceLFilePath)
                    Errorlist1.append(Error2)
                    ProductEventDetails(ExceLFilePath)
                    Errorlist1.append(Error3)
                    ckResultLen=len(data)
                    InsertTrusteeCheckByTrustId(userId, TrustId,ckResultLen)
                    UpdateProductStatus(TrustId)
                else:
                    Error=1
                    Errorlist.append(Error)
            else:
                Error=1
                Errorlist.append(Error)
                FileError='文件名称格式错误!应为[TrustCode;TrustName;基础表.xlsx]'
                ErrorMessage('*','*','*',ExceLFilePath,FileError)
                print(name,"--文件名称格式错误!应为[TrustCode;TrustName;基础表.xlsx]")





        
    print(sum(Errorlist),sum(Errorlist1))
    if sum(Errorlist)==0 and sum(Errorlist1)==0:
        IsSucess=1
        result='计划说明书基础表校验通过!'
        InsertVerificationLog(userId,9,IsSucess,errtxtFilePath,result,filepath)
    else:
        IsSucess=0
        result='计划说明书基础表校验失败,详细错误请下载查看!'
        InsertVerificationLog(userId,9,IsSucess,errtxtFilePath,result,filepath)
    
    
    
    