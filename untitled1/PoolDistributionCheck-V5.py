

# -*- coding: utf-8 -*-
"""
Created on Wed Jan 15 14:38:26 2020

@author: HUAWEI
"""


#%%dbConnectionStr = None
errtxtFilePath = None
Error=0
Excelfilepath=None
def ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,ErrorInformation):
    wb = openpyxl.load_workbook(errtxtFilePath) #读取xlsx文件
    ws=wb['Sheet']
    maxrow=ws.max_row
    maxrow+=1
   # print(maxrow)
    ws.cell(maxrow,3,TrustId).value
    ws.cell(maxrow,4,TrustCode).value
    ws.cell(maxrow,5,TrustName).value
    ws.cell(maxrow,6,Excelfilepath).value
    ws.cell(maxrow,7,ErrorInformation).value
    ws.cell(maxrow,9,2).value   
    wb.save(errtxtFilePath)

    
    
def PoolDistributions(Excelfilepath,FileTrustCode,TrustName):
    
    global TrustId,Error,errtxtFilePath
    Error=0
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    try:
        selectId="select TrustId from TrustManagement.Trust where TrustCode='{}'".format(FileTrustCode)
        b1.execute(selectId)
        TrustId=b1.fetchone()[0]
    except:
        Error=1
        C0='文件TrustCode【{}】与系统TrustCoden不匹配请检查!'.format(FileTrustCode)
        ErrorMessage('#',FileTrustCode,TrustName,Excelfilepath,C0)
        print(TrustName,C0)
        sys.exit(1)
    
    data=pd.read_excel(Excelfilepath,skiprows=1,errors='ignore')
    data=data.iloc[:,:12]
    PoolDistributions_columns=['PaymentPeriodID','资产池分布类型','DistributionType','DatabaseItem','BucketSequenceNo','Bucket','Amount','AmountPercentage','Count','CountPercentage','CustomerCount','CustomerCountPercentage']
    
    for columnsP in PoolDistributions_columns:
        if columnsP not in data.columns:
#            print(columnsP)
            print(TrustId,TrustName,'表格字段错误不能含有中文或字段缺失!')
            C9='表格字段错误不能含有中文或字段缺失!'
            
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C9)
            return 
            
#            sys.exit(1)
    
    data.dropna(subset=['Amount','Bucket','AmountPercentage'],inplace=True)
    data=data[['PaymentPeriodID','DistributionType','DatabaseItem','BucketSequenceNo','Bucket','Amount','AmountPercentage','Count','CountPercentage','CustomerCount','CustomerCountPercentage']]
    
    for typeNo in data.index:
#        try:
        DatabaseItem= 'null' if pd.isnull(data.loc[typeNo][2])==True else data.loc[typeNo][2]
        PaymentPeriodID='null' if pd.isnull(data.loc[typeNo][0])==True else data.loc[typeNo][0]
        BucketSequenceNo='null' if pd.isnull(data.loc[typeNo][3])==True else data.loc[typeNo][3]
#            BucketSequenceNolist.append(BucketSequenceNo)
        AmountPercentage='null' if pd.isnull(data.loc[typeNo][6])==True else data.loc[typeNo][6]
        if type(AmountPercentage) in (float,int,np.float32,np.float64,np.int32,np.int64) or AmountPercentage=='null' or  AmountPercentage=='NA' :
            pass
        else:
            CJ1='AmountPercentage，第{}行,数据类型错误!应为数值型。'.format(typeNo)
            Error=1
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,CJ1)
            print(CJ1,AmountPercentage,'AmountPercentage')

        
        CountPercentage='null' if pd.isnull(data.loc[typeNo][8])==True else data.loc[typeNo][8]
        if type(CountPercentage) in (float,int,np.float32,np.float64,np.int32,np.int64) or CountPercentage=='null' or CountPercentage=='NA':
            pass
        else:
            CJ11='CountPercentage，第{}行,数据类型错误!应为数值型。'.format(typeNo)
            Error=1
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,CJ11)
            print(CJ11,CountPercentage,'CountPercentage')

        
        Amount='null' if pd.isnull(data.loc[typeNo][5])==True else data.loc[typeNo][5]
        if type(Amount) in (float,int,np.float32,np.float64,np.int32,np.int64) or Amount=='null' or  Amount=='NA':
            pass
        else:
            CJ111='Amount，第{}行,数据类型错误!应为数值型。'.format(typeNo)
            Error=1
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,CJ111)
            print(CJ111,Amount,'Amount')
        
        
        Count='null' if pd.isnull(data.loc[typeNo][7])==True else data.loc[typeNo][7]
        if type(Count) in (float,int,np.float32,np.float64,np.int32,np.int64) or Count=='null' or Count=='NA':
            pass
        else:
            CJ1111='Count，第{}行,数据类型错误!应为数值型。'.format(typeNo)
            Error=1
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,CJ1111)
            print(CJ1111,Count,'Count')

        CustomerCount ='null' if pd.isnull(data.loc[typeNo][9])==True else data.loc[typeNo][9]
        if type(CustomerCount) in (float, int, np.float32, np.float64, np.int32, np.int64) or CustomerCount=='null' or CustomerCount=='NA':
            pass
        else:
            MgCustomerCount = '{},数据类型错误!应为数值型。'.format(typeNo)
            Error = 1
            ErrorMessage(TrustId, FileTrustCode, TrustName, Excelfilepath, MgCustomerCount)
            print(MgCustomerCount,CustomerCount,'CustomerCount')

        CustomerCountPercentage ='null' if pd.isnull(data.loc[typeNo][10])==True else data.loc[typeNo][10]
        if type(CustomerCountPercentage) in (float, int, np.float32, np.float64, np.int32, np.int64) or CustomerCountPercentage=='null'or CustomerCountPercentage=='NA':
            pass
        else:
            MGCustomerCountPercentage = '{},数据类型错误!应为数值型。'.format(typeNo)
            Error = 1
            ErrorMessage(TrustId, FileTrustCode, TrustName, Excelfilepath, MGCustomerCountPercentage)
            print(MGCustomerCountPercentage,CustomerCountPercentage,'CustomerCountPercentage')

        
#            print(DatabaseItem,AmountPercentage,CountPercentage,Amount,Count)
        DistributionType=data.loc[typeNo][1]
        Bucket=data.loc[typeNo][4]

                
        if PaymentPeriodID!=0:
            Error=1
            C8='PaymentPeriodID填写值【{}】错误!只能填【0】'.format(PaymentPeriodID)
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C8)
            print(C8)
       
    print(Error)
    if Error==0:
        print(TrustId,TrustName,'池分布数据校验通过!')
        PoolDistributionsImport(Excelfilepath,FileTrustCode)

            

def PoolDistributionsImport(Excelfilepath,FileTrustCode):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    #删除上次插入数据
    deletePoolDistributions1="delete dbo.PoolDistributions1 where TrustId={} and PaymentPeriodID=0".format(TrustId)
    b1.execute(deletePoolDistributions1)
    conn.commit()
    print('历史数据已清除')

    DataPoolDistributions=pd.read_excel(Excelfilepath,header=1,errors='ignore')
    DataPoolDistributionsImport=DataPoolDistributions[['PaymentPeriodID','DistributionType','DatabaseItem','BucketSequenceNo','Bucket','Amount','AmountPercentage','Count','CountPercentage','CustomerCount','CustomerCountPercentage']]
    DataPoolDistributionsImport=DataPoolDistributionsImport.dropna(subset=['Amount','Count',])
    #print(DataPoolDistributionsImport)
    
    for Cindex in DataPoolDistributionsImport.index:
        PaymentPeriodID='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][0])==True else DataPoolDistributionsImport.loc[Cindex][0]
        #PaymentPeriodID=int(PaymentPeriodID)
        
        DatabaseItem='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][2])==True else DataPoolDistributionsImport.loc[Cindex][2]
        
        DistributionTypeCode='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][1])==True else DataPoolDistributionsImport.loc[Cindex][1]
        BucketSequenceNo= 'null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][3])==True else DataPoolDistributionsImport.loc[Cindex][3]
        #BucketSequenceNo=int(BucketSequenceNo)
        Bucket='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][4])==True else DataPoolDistributionsImport.loc[Cindex][4]
        Amount= 'null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][5])==True else DataPoolDistributionsImport.loc[Cindex][5]
        if Amount=='NA':
            Amount='null'

        AmountPercentage='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][6])==True else DataPoolDistributionsImport.loc[Cindex][6]
        if AmountPercentage=='NA':
            AmountPercentage='null'
        if AmountPercentage!='null':
            AmountPercentage=round(AmountPercentage,4)
            
        Count='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][7])==True else  DataPoolDistributionsImport.loc[Cindex][7]
        if Count=='NA':
            Count='null'

        CountPercentage='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][8])==True else  DataPoolDistributionsImport.loc[Cindex][8]
        if CountPercentage=='NA':
            CountPercentage='null'
        if CountPercentage!='null':
            CountPercentage=round(CountPercentage,4)

        CustomerCount = 'null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][9])==True else  DataPoolDistributionsImport.loc[Cindex][9]
        if CustomerCount=='NA':
            CustomerCount='null'
        if CustomerCount!='null':
            CustomerCount=round(CustomerCount,0)

        CustomerCountPercentage ='null' if pd.isnull(DataPoolDistributionsImport.loc[Cindex][10])==True else  DataPoolDistributionsImport.loc[Cindex][10]
        if CustomerCountPercentage=='NA':
            CustomerCountPercentage='null'
        if CustomerCountPercentage!='null':
            CustomerCountPercentage=round(CustomerCountPercentage,4)
            
        #print(AmountPercentage)
        if  AmountPercentage!='null':
            if AmountPercentage<1:
                AmountPercentage=AmountPercentage*100
        if CountPercentage!='null':
            if CountPercentage<1:
                CountPercentage=CountPercentage*100
        if  CustomerCountPercentage!='null':       
            if CustomerCountPercentage<1:
                CustomerCountPercentage=CustomerCountPercentage*100
		
        if Bucket!='null':
            PoolDistributions1Insert="insert into dbo.PoolDistributions1 values({},{},'{}',{},N'{}',{},{},{},{},N'{}',{},{})".format(TrustId,PaymentPeriodID,DistributionTypeCode,BucketSequenceNo,Bucket,Count,CountPercentage,Amount,AmountPercentage,DatabaseItem,CustomerCount,CustomerCountPercentage)
            #print(PoolDistributions1Insert)
            b1.execute(PoolDistributions1Insert)
            conn.commit()
            
        if Bucket=='null':
            PoolDistributions1Insert="insert into dbo.PoolDistributions1 values({},{},'{}',{},{},{},{},{},{},N'{}',{},{})".format(TrustId,PaymentPeriodID,DistributionTypeCode,BucketSequenceNo,Bucket,Count,CountPercentage,Amount,AmountPercentage,DatabaseItem,CustomerCount,CustomerCountPercentage)
            #print(PoolDistributions1Insert)
            b1.execute(PoolDistributions1Insert)
            conn.commit()
    b1.close()
    conn.close()
    # C6='池分布导入完成!'
    # ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C6)
    print("{},{},{},池分布导入完成!".format(TrustId,FileTrustCode,TrustName)) 

        
def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        writeLog(str(ex))
        writeErr("\n【{0}】".format(sourceFilePath))
        writeErr(str(ex))
        print(str(ex))
        # raise ex
    finally:
        cnxn.close()
        
        
        
def InsertVerificationLog(userId, CheckType, IsSucess, filePath, result,filepath):
    sql = "exec TaskCollection.dbo.usp_InsertVerificationLog N'{0}',N'{1}',N'{2}',N'{3}',N'{4}',N'{5}'".format(userId,CheckType,IsSucess, filePath, result,filepath)
    execSQLCmd(sql)
    
    
def InsertTrusteeCheckByTrustId(userId, TrustId,ckResultLen):
    sql = "exec TaskCollection.dbo.usp_InsertTrusteeCheckByTrustId N'{0}',N'{1}',N'{2}',N'{3}',N'{4}'".format(userId,10, TrustId, 0, ckResultLen)
    execSQLCmd(sql)
	
#更新任务状态
def UpdateProductStatus(TrustId):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='TaskCollection', charset='utf8')
    b1=conn.cursor()
    
    SelectTrustDocumentID="select TrustDocumentID from TaskCollection.[dbo].[view_TrusteeReportForDocumentId] where TrustId={} and Period=0 and FileType='BasicInformation'".format(TrustId)
    b1.execute(SelectTrustDocumentID)
    TrustDocumentID=b1.fetchone()[0]
#    print(TrustDocumentID)
    execsql="[TaskCollection].[dbo].[UpdateProductStatus] {},{},'{}',1".format(TrustId,TrustDocumentID,'PoolDistribution')
#    print(execsql)
    b1.execute(execsql)
    conn.commit()	

	
def runDBDataValidation(TrustId):
    sql = "exec TaskCollection.dbo.[usp_VerifyTrustPoolDistributionsByTrust] N'{0}'".format(TrustId)
    dbCheckResult = execSQLCmd(sql)
    selectId = "SELECT top 1  TrustId FROM PortfolioManagement.[DV].[VerifyTrustPoolDistributionsByTrustRemarksLog] WHERE TrustId= {}".format(TrustId)
    cnxn = pyodbc.connect(dbConnectionStr)
    b1 = cnxn.cursor()
    b1.execute(selectId)
    Result = b1.fetchone()[0]
    return Result  
        
    
if __name__=="__main__":
    import xlrd,xlwt
    import pandas as pd,numpy as np
    import numpy as np
    import os
    import time
    import sys
    import datetime
    import pymssql
    import pyodbc
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl import Workbook
    
    filepath=r'\\172.16.6.143\Products\计划说明书\202008\ABN\池分布校验-V5\池分布\王刚依\王刚依\深圳前海联易融商业保理有限公司2020年度第二期资产支持票据'
    # filepath = str(sys.argv[1])
    # dateId = str(sys.argv[2])
    # userId = str(sys.argv[3])
    IsResultSucess=0
    dbConnectionStr="DRIVER={SQL Server};SERVER=172.16.6.143\MSSQL;DATABASE=TaskCollection;UID=sa;PWD=PasswordGS2017"
    
    errtxtFilePath = os.path.join(r'C:\Users\DELL\Desktop\华能信托·惠橙6号集合资金信托计划','Error_池分布校验结果_{0}.xlsx'.format(100000))
    if not os.path.exists(errtxtFilePath):
        wb = Workbook()  # 新建工作簿
        ws1 = wb.active
        wb.save(errtxtFilePath)
        wb = openpyxl.load_workbook(errtxtFilePath) # 读取xlsx文件
        ws=wb['Sheet']
        maxrow=ws.max_row
        ws.cell(maxrow,1,'负责人').value
        ws.cell(maxrow,2,'操作人').value
        ws.cell(maxrow,3,'TrustId').value
        ws.cell(maxrow,4,'TrustCode').value
        ws.cell(maxrow,5,'TrustName').value
        ws.cell(maxrow,6,'Path').value
        ws.cell(maxrow,7,'ErrorInformation').value
        ws.cell(maxrow,8,'备注').value
        ws.cell(maxrow,9,'Type').value
        wb.save(errtxtFilePath)
        
            
    #循环遍查找指定文件夹
    ErrorList=[]
    for root,dirs,files in os.walk(filepath):
        for name in files:
            if not name.endswith('池分布.xlsx'):
#                print(name,'文件名称不匹配!')
                continue
            Excelfilepath = os.path.join(root,name)
            if ';' in name:
                SplitName=name.split(';')
                TrustCode=SplitName[0]
                TrustName=SplitName[1]
                print(TrustCode,TrustName)
                PoolDistributions(Excelfilepath,TrustCode,TrustName)
                data=pd.read_excel(errtxtFilePath)
                ckResultLen=len(data)
                InsertTrusteeCheckByTrustId(userId, TrustId,0)
                UpdateProductStatus(TrustId)
                ErrorList.append(Error)
            else:
                Error=1
                ErrorList.append(Error)
                FileError='文件名称格式错误!应为[TrustCode;TrustName;池分布.xlsx]'
                ErrorMessage('*','*','*',Excelfilepath,FileError)
                print(name,"--文件名称格式错误!应为[TrustCode;TrustName;池分布.xlsx]")       
    

    if sum(ErrorList)==0:
        IsSucess=1
        result='计划说明书池分布校验通过!'
        InsertVerificationLog(userId, 10, IsSucess, errtxtFilePath, result,filepath)
    else:
        IsSucess=0
        result='计划说明书池分布校验错误，详细错误请下载查看!'
        InsertVerificationLog(userId, 10, IsSucess, errtxtFilePath, result,filepath)