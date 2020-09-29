# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import re
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import openpyxl.styles as sty

logtxtFilePath = None
class cObj(object):
    def __init__(self, cell, desc, val):
        self.cell = cell
        self.desc = desc
        self.val = val

#mapping文件添加map节点
def addNodeForMapXml(MapPath, sourcecell, destCell, comment, sourcecolor, destcolor):
    tree = XETree.parse(MapPath)
    root = tree.getroot()
    node = root.find("Mapping")
    MapNode = XETree.Element('map')  # 创建节点,单个文件的mapping文件
    MapNode.set("destcolor", destcolor)
    MapNode.set("sourcecolor", sourcecolor)
    MapNode.set("comment", comment)
    MapNode.set("destCell", destCell)
    MapNode.set("sourcecell", sourcecell)
    node.append(MapNode)
    indent(node)
    tree.write(MapPath, encoding='utf-8', xml_declaration=True)
    return
def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def is_number(s):
    try:
        float(s.replace(',', ''))
        return True
    except ValueError:
        return False

def getNumberVal(val):
    if isinstance(val, int) or isinstance(val, float):
        return val
    return None if val == 'NA' or val =='-' or val == '' else float(val.replace('\t', '').replace('\n', '').replace(',', '').replace('%', '').replace(' ', ''))

def markCellError(sheet, ctag, errtype):
    color = 'E93936' #datatype
    if errtype == 1:#empty
        color = 'AA2927'
    sheet[ctag].fill=sty.PatternFill(fill_type='solid', fgColor=color)

def checkCellValue(sheet, ctag, cvalue, cdtype):
    if ((cdtype == 'mumber' or cdtype == 'rate') and isinstance(cvalue, float)
        ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return 1, 0, 0

    cvalue = str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '')

    if cvalue == 'NA' or cvalue == '-':
        return 1, 1, 0

    if cvalue == '':
        markCellError(sheet, ctag, 1)
        return 0, 0, 1

    dtcheck = 0
    if cdtype == 'int':
        dtcheck = 1 if cvalue.isdigit() else 0
    elif cdtype == 'number':
        dtcheck = 1 if is_number(cvalue) else 0
    elif cdtype == 'date':
        reg = r"(\d{4}[-/]\d{1,2}([-/]\d{1,2})?)|((\d{1,2}[-/])?\d{1,2}[-/]\d{4})|(\d{4}年\d{1,2}月(\d{1,2}日)?)|(\d{6})|(\d{8})"
        dtcheck = 1 if re.search(reg, cvalue) is not None else 0
    elif cdtype == 'rate':
        dtcheck = 1 if is_number(cvalue.rstrip('%')) else 0
    else:  # string
        dtcheck = 1

    if dtcheck == 0:
        markCellError(sheet, ctag, 2)
        return 0, 0, 2

    return 1, 0, 0

def specificCellsCheck(cfgItem, sheet):
    checkmsg = []
    r = []
    for cell in cfgItem:
        ctag = cell.tag
        cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

        isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)

        if isValid != 1:
            etype = "不能为空" if errorType == 1 else "格式错误"
            checkmsg.append("{0}:{3}:[{1}] - {2}".format(ctag, cdesc, etype, errorType))

        r.append(cObj(ctag, cdesc, cvalue))

    return checkmsg, r

def specificColsCheck(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 1
    breakRowsLoop = 0
    rs = []
    while rStart <= rEnd:  # rows loop
        r = []
        for cell in cNode: # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break #break cells loop

            isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)
            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            if isValid != 1:
                etype = "不能为空" if errorType == 1 else "格式错误"
                checkmsg.append("{0}:{3}:[{1}] - {2}".format(ctag, cdesc, etype, errorType))

            r.append(cObj(ctag, cdesc, cvalue))

        if (isFirstRow == 1 and firstRowAllNA == 1) or breakRowsLoop == 1:
            rStart = rEnd + 1 #break rows loop
        else:
            isFirstRow = 0
            rStart += 1
            rs.append(r)

    return checkmsg, rs

def writeSheetLog(wb, res, mapPath):
    logSheet = wb.create_sheet("格式检查")
    i = 1
    while i <= len(res):
        if len(res[i-1].split(':')) == 3:
            sourceCell = res[i-1].split(':')[0]
            destCell = "A{0}".format(i)
            errorType = res[i-1].split(':')[1]
            comment = res[i-1].split(':')[2]
            sourceColor = "255,230,153"
            destColor = ''
            if errorType == 1:
                destColor = "170,41,39"
            elif errorType == 2:
                destColor = "233,57,54"
            else:
                destColor = "255,230,153"
            addNodeForMapXml(mapPath, sourceCell, destCell, comment, sourceColor, destColor)
            logSheet[destCell] = str(sourceCell + '：' + comment)
            i += 1
        else:
            logSheet["A{0}".format(i)] = res[i - 1]
            i += 1

def getValidationResult(ds):
    ckResult = []

    bondsAmt = 0
    cells = ''
    for d in ds[0]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestPaied = getNumberVal(d[3].val)
        pendAmt = getNumberVal(d[4].val)

        if pInterestPaied < 100000:
            ckResult.append("{1}:{2}:【收益分配记录】债券利息还款 [{0}] 小于10万,数据范围错误".format(pInterestPaied, d[3].cell, 0))
        if pendAmt < 100000:
            ckResult.append("{1}:{2}【收益分配记录】债券期末金额 [{0}] 小于10万,数据范围错误".format(pendAmt, d[4].cell, 0))

        if pendAmt is not None and pstartAmt is not None and pendAmt > pstartAmt:
            ckResult.append("{2};{3}:{4}:【收益分配记录】债券期末余额 [{0}] 大于期初余额{1},数据对应关系错误".format(pendAmt, pstartAmt, d[4].cell, d[1].cell, 0))
        if pPrincipalAmt is not None and pstartAmt is not None and pPrincipalAmt > pstartAmt:
            ckResult.append("{2};{3}:{4}:【收益分配记录】债券本金金额 [{0}] 大于期初余额{1}，数据对应关系错误".format(pPrincipalAmt, pstartAmt, d[2].cell, d[1].cell, 0))

        if pstartAmt is not None and pPrincipalAmt is not None and pendAmt is not None and pstartAmt - pPrincipalAmt != pendAmt:
            ckResult.append("{3};{4};{5}:{6}:【收益分配记录】期初余额 [{0}] 减去本金金额 [{1}] 不等于期末余额 [{2}] 错误".format(pstartAmt,
                                                                                                   pPrincipalAmt, pendAmt, d[1].cell, d[2].cell, d[4].cell, 0))
        if pendAmt is not None:
            bondsAmt += pendAmt
            cells += str(d[4].cell) + ';'

    if bondsAmt is not None and bondsAmt != 0:
        assetAmt = getNumberVal(ds[5][2].val)
        if assetAmt is not None and assetAmt != 0 and assetAmt < 100000:
            ckResult.append("{1}:{2}:【笔数与金额特征】资产池总金额 [{0}] 小于10万,数据范围错误".format(assetAmt, ds[5][2].cell, 0))
        if assetAmt is not None and assetAmt != 0 and abs(assetAmt - bondsAmt) / assetAmt > 0.05:
            ckResult.append("{2}{3}:{4}:【收益分配记录 与 笔数与金额特征】各债券期初余额之和 [{0}] 同资产池统计的总金额 [{1}] 相差在5%之上错误".format(bondsAmt,
                                                                                                         assetAmt, cells, ds[5][2].cell, 0))

    t2 = 0
    t4 = 0
    cells2 = ''
    cells4 = ''
    for d in ds[1]:
        loanCount = getNumberVal(d[2].val)
        amtCount = getNumberVal(d[4].val)
        if loanCount is not None:
            t2 += loanCount
            cells2 += str(d[2].cell) + ';'
        if amtCount is not None:
            t4 += amtCount
            cells4 += str(d[4].cell) + ';'
    if t2 != 0 and (t2 != 100 or t2 != 1):
        ckResult.append("{0}:{1}:【资产池整体表现情况】笔数占比之和不等于100%".format(cells2[:-1], 0))
    if t4 != 0 and (t4 != 100 or t4 != 1):
        ckResult.append("{0}:{1}:【资产池整体表现情况】金额占比之和不等于100%".format(cells4[:-1], 0))

    pstartLast = None
    pPrincipalLast = None
    pInterestlast = None
    for d in ds[2]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestAmt = getNumberVal(d[3].val)

        if pPrincipalAmt < 100000:
            ckResult.append("{1}:{2}:【现金流归集表】应收本金金额 [{0}] 小于10万,数据范围错误".format(pPrincipalAmt, d[2].cell, 0))
        if pInterestAmt < 100000:
            ckResult.append("{1}:{2}:【现金流归集表】应收利息金额 [{0}] 小于10万,数据范围错误".format(pInterestAmt. d[3].cell, 0))

        if pstartAmt is not None and pstartAmt is not None and pstartAmt < pPrincipalAmt:
            ckResult.append("{2};{3}:{4}:【现金流归集表】当期期初本金金额 [{0}] 小于应收本金金额 [{1}] 数据对应关系错误".format(pstartAmt,
                                                                                            pPrincipalAmt, d[1].cell, d[2].cell, 0))

        '''
        if pstartLast is not None and pstartAmt is not None and pstartLast < pstartAmt:
            ckResult.append("【现金流归集表】当期期初本金金额 [{0}] 大于上期期初本金金额 [{1}] 错误".format(pstartAmt, pstartLast))        
        if pPrincipalLast is not None and pPrincipalAmt is not None and pPrincipalLast < pPrincipalAmt:
            ckResult.append("【现金流归集表】当期应收本金金额 [{0}] 大于上期应收本金金额 [{1}] 错误".format(pPrincipalAmt, pPrincipalLast))
        if pInterestlast is not None and pInterestAmt is not None and pInterestlast < pInterestAmt:
            ckResult.append("【现金流归集表】当期应收利息金额 [{0}] 大于上期应收利息金额 [{1}] 错误".format(pInterestAmt, pInterestlast))
        '''
        pstartLast = pstartAmt
        pPrincipalLast = pPrincipalAmt
        pInterestlast = pInterestAmt

    assetPoolCF = ds[3]
    if len(assetPoolCF) > 0:
        val = getNumberVal(assetPoolCF[0].val)
        if val is not None and val != 0 and val < 100000:#利息-正常回收
            ckResult.append("{2}:{3}:【资产池现金流详情】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolCF[0].desc, val, assetPoolCF[0].cell, 0))
        val = getNumberVal(assetPoolCF[8].val)
        if val is not None and val != 0 and val < 100000:#本金-正常回收
            ckResult.append("{2}:{3}:【资产池现金流详情】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolCF[8].desc, val, assetPoolCF[8].cell, 0))

    assetPoolST = ds[4]
    if len(assetPoolST) > 0:
        val = getNumberVal(assetPoolST[0].val)
        if val is not None and val != 0 and val < 100000:  # 收入合计
            ckResult.append("{2}:{3}:【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[0].desc, val, assetPoolST[0].cell, 0))
        val = getNumberVal(assetPoolST[1].val)
        if val is not None and val != 0 and val < 100000:  # 本金合计
            ckResult.append("{2}:{3}:【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[1].desc, val, assetPoolST[1].cell, 0))
        val = getNumberVal(assetPoolST[2].val)
        if val is not None and val != 0 and val < 10000:  # 税费支出-税收
            ckResult.append("{2}:{3}:【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[2].desc, val, assetPoolST[2].cell, 0))
        val = getNumberVal(assetPoolST[3].val)
        if val is not None and val != 0 and val < 10000:  # 税费支出-资产服务报酬
            ckResult.append("{2}:{3}:【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[3].desc, val, assetPoolST[3].cell, 0))

    maxPAmt = getNumberVal(ds[5][3].val)
    avgPAmt = getNumberVal(ds[5][4].val)
    if maxPAmt is not None and avgPAmt is not None and maxPAmt < avgPAmt:
        ckResult.append("{2};{3}:{4}:【笔数与金额特征】最高本金余额 [{0}] 小于平均本金余额 [{1}] 数据对应关系错误".format(maxPAmt, avgPAmt, ds[5][3].cell, ds[5][4].cell, 0))

    maxRT = getNumberVal(ds[5][8].val)
    minRT = getNumberVal(ds[5][9].val)
    avgRT = getNumberVal(ds[5][7].val)
    if maxRT is not None and minRT is not None and maxRT < minRT:
        ckResult.append("{2};{3}:{4}:【笔数与金额特征】最长剩余期限 [{0}] 小于最短剩余期限 [{1}] 数据对应关系错误".format(maxRT, minRT, ds[5][8].cell, ds[5][9].cell, 0))
    if maxRT is not None and minRT is not None and avgRT is not None and (maxRT < avgRT or minRT > avgRT):
        ckResult.append("{3};{4};{5}:{6}:【笔数与金额特征】平均剩余期限 [{2}] 未介于最长剩余期限 [{0}] 小于最短剩余期限 [{1}] 之间,数据对应关系错误".format(maxRT,
                                                                                                              minRT, avgRT, ds[5][8].cell, ds[5][9].cell, ds[5][7].cell, 0))

    maxRT = getNumberVal(ds[5][11].val)
    minRT = getNumberVal(ds[5][12].val)
    avgRT = getNumberVal(ds[5][10].val)
    if maxRT is not None and minRT is not None and maxRT < minRT:
        ckResult.append("{2};{3}:{4}:【笔数于金额特征】最高利率 [{0}] 小于最低利率 [{1}] 数据对应关系错误".format(maxRT, minRT, ds[5][11].cell, ds[5][12].cell, 0))
    if maxRT is not None and minRT is not None and avgRT is not None and (maxRT < avgRT or minRT > avgRT):
        ckResult.append("{3};{4};{5}:{6}:【笔数于金额特征】平均利率 [{2}] 未介于最高利率 [{0}] 和最低利率 [{1}] 之间,数据对应关系错误".format(maxRT,
                                                                                                       minRT, avgRT, ds[5][11].cell, ds[5][12].cell, ds[5][10].cell, 0))

    #累计违约率有可能会降低，所以暂时不验证是否逐期递增，只验证是否每期的数字都在0.1以内
    for d in ds[7]:
        cdr = getNumberVal(d[1].val)
        if cdr is not None and cdr != 0:
            if (str(d[1].val).endswith('%') and cdr > 10) or (not str(d[1].val).endswith('%') and cdr > 0.1):
                ckResult.append("{1}:{2}:【累计违约率】违约率数值 [{0}] 大于10%，取值范围过大错误".format(cdr, d[1].cell, 0))

    return ckResult

def checkFileFormat(wb, cfgItems, MapPath):
    sheet = wb['Sheet1']
    ckResult = []
    tp = [[], [], [], [], [], [], [], []]
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']
        itemtype = cfgItem.attrib['type'] if 'type' in cfgItem.attrib else ''

        itemCkResult = []
        r = []
        if itemtype == 'SpecificCells':
            itemCkResult, r = specificCellsCheck(cfgItem, sheet)
        else:
            itemCkResult, r = specificColsCheck(cfgItem, sheet)

        tp[i] = r
        if len(itemCkResult) > 0:
            ckResult.append("【{0}】".format(itemdesc))
            ckResult.extend(itemCkResult)

    if len(ckResult) == 0:
        ckResult = getValidationResult(tp)
    if len(ckResult) > 0:
        writeSheetLog(wb, ckResult, MapPath)

    return len(ckResult)
#给xml增加换行符
def indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i
#新建mamppingxml， multiply 用于判断是单个文件还是文件夹
def createXml(xmlPath, inputfile, outputfile, multiply):
    if multiply.lower() == 'false':
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "false")
            root.set("inputFile", inputfile)
            root.set("outputFile", outputfile)
            root.set("inputSheetIndex", "0")
            root.set("outputSheetIndex", "1")
            root.set("inputSheetName", "Sheet1")
            root.set("outputSheetName", r"格式检查")
            tree = XETree.ElementTree(root)  # 创建文档
            Mapping1 = XETree.Element('Mapping') #创建子节点
            Mapping1.set("description", r'源文件，目标文件对应情况')
            root.append(Mapping1)
            indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
    else:
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "true")
            tree = XETree.ElementTree(root)  # 创建文档
            # indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
#开始函数，解析xml获取源文件，目标文件
def main(configFilePath, dateId):
    global logtxtFilePath

    scriptFolder = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolder, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolder, 'Logs',
                                  'Log_FormatCheck_{0}.txt'.format(datetime.datetime.now().strftime('%m-%d %H%M%S')))
    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    sourceFolderPath = cfgRoot.attrib['sourcefolder']
    dir_path = os.path.dirname(os.path.abspath(__file__)) + '\\MappingXml\\'  # mapping文件存放路径
    mappingPath = dir_path + dateId + '.xml'
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)
    for dirPath, dirNames, fileNames in os.walk(sourceFolderPath):
        config = 1
        createXml(mappingPath, '', '', 'true')
        for fileName in fileNames:
            if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告数据提取;'):
                print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
                continue

            sourceFilePath = os.path.join(dirPath, fileName)

            msg = "\n{0}".format(sourceFilePath)
            writeLog(msg)

            fileNameAry = fileName.split(';')
            if len(fileNameAry) != 4:
                msg = "【文件名错误】文件名称命名不规范"
                writeLog(msg)
                continue

            paymentPeriodID = 0

            paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
            if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
                msg = "【错误】文件名中的TrustCode或报告期数设置有误"
                writeLog(msg)
                continue

            mulPath = dir_path + dateId + '_' + str(config) + '.xml'
            createXml(mulPath, sourceFilePath, sourceFilePath, 'false')
            config += 1
            tree = XETree.parse(mappingPath)
            root = tree.getroot()
            MapPath = XETree.Element('filename')  # 创建节点,单个文件的mapping文件
            MapPath.set("path", mulPath)
            root.append(MapPath)
            indent(root)
            tree.write(mappingPath, encoding='utf-8', xml_declaration=True)

            excelwb = load_workbook(sourceFilePath)
            if '格式检查' in excelwb.sheetnames:
                excelwb.remove(excelwb['格式检查'])
                excelwb.save(sourceFilePath)

            hasError = checkFileFormat(excelwb, cfgRoot, mulPath)
            if hasError > 0:
                excelwb.save(sourceFilePath)
                writeLog('【有格式错误】详情见文档[格式检查]sheet')


