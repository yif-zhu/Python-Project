import os
import re
import os.path
import pyodbc
import datetime
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
from openpyxl import Workbook

sourceExcel = None
destExcel = None
dbConnectionStr = 'DRIVER={SQL Server};SERVER=172.16.6.143\MSSQL;DATABASE=TaskCollection;UID=sa;PWD=PasswordGS2017'
dbConnectionStr1 = 'DRIVER={SQL Server};SERVER=10.0.0.173\MSSQL;DATABASE=FixedIncomeSuite;UID=sa;PWD=PasswordGS2017'
logtxtFilePath = None
errtxtFilePath = None

def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def writeErr(msg):
    if not os.path.exists(errtxtFilePath):
        f = open(errtxtFilePath, "w")
    # print(msg)
    with open(errtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        writeLog(str(ex))
        print(str(ex))
        # raise ex
    finally:
        cnxn.close()

def execSQLCmdFetchOne(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr1)
    try:
        cursor = cnxn.cursor()
        row = cursor.execute(sql).fetchall()
        return row
    except Exception as ex:
        writeLog(str(ex))
        raise ex
    finally:
        cnxn.close()

def execSQLCmdFetchAll(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        rows = cursor.execute(sql).fetchall()
        return rows
    except Exception as ex:
        writeLog(str(ex))
    finally:
        cnxn.close()

def getServiceChargeByTrustCode(TrustCode):  #获取系统的超额服务费和model类型
    sql = "exec FixedIncomeSuite.Analysis.usp_GetServiceChargeByTrustCode  N'{0}'".format(TrustCode)
    try:
        fee = execSQLCmdFetchOne(sql)
        return fee
    except Exception as ex:
        writeErr(str(ex))
        return 0

def getValuesFromVerificationLog(TrustCode):  #获取匹配备注信息之后的报错信息
    sql = "exec TaskCollection.dbo.usp_getTrustCalculationError  N'{0}'".format(TrustCode)
    msg = execSQLCmdFetchAll(sql)
    return msg

def DeleteValuesToVerificationLog(TrustCode):  #清除原有报错信息
    sql = "exec TaskCollection.dbo.usp_DeleteOldVeriicationLog  N'{0}'".format(TrustCode)
    execSQLCmd(sql)

def getCalculationInfo(TrustCode):  #获取信托测试数据库数据，用于对比
    sql = "exec TaskCollection.dbo.usp_getCalculationInfoForTrustId N'{0}'".format(TrustCode)
    msg = execSQLCmdFetchAll(sql)
    return msg

def getTrustCodeFromXMl(sourceFolder):  #从xml 中获取TrustCode
    mappingTree = XETree.parse(sourceFolder)
    cfgRoot = mappingTree.getroot()
    return cfgRoot[0][6][0][2][0].text

'''
TableName:表示在哪张表里取值
columName:表示判断表的对应位置的表列名
ItemCode: 表示对于列名中对应的值
ValueColumn: 表示对应的值所在列名
IsTrustBond: 用来判断当是分层信息是，取优先级的数据
'''
def getValueFromXML(sourcePath, TableName, columnName, ItemCode, ValueColumn):
    mappingTree = XETree.parse(sourcePath)
    cfgRoot = mappingTree.getroot()
    for table in cfgRoot:
        if table[0].text == TableName:  #确定找到对应的表名
            for tagName in table: #找到对应的表的Data节点
                if tagName.tag == 'data':
                    ValueRowNumber = -1  # 获取对应列的ItemCode对应的位置
                    isNone = 1
                    for column in tagName: #找到对应列的ItemCode和对应的值的位置
                        if column[0].text == columnName:
                            columnLen = len(column[2])
                            for RowName in column[2]:
                                ValueRowNumber += 1
                                if RowName.text == ItemCode: #找到对应的ItemCode所在位置
                                    isNone = 0
                                    break
                                if columnLen == (ValueRowNumber + 1): #当不存在该ItemCode字段时
                                    isNone = 1
                        if column[0].text == ValueColumn and isNone == 0:
                            return column[2][ValueRowNumber].text
    return -1

def getValueFromExcel(sourcePath, Cell):    #Cell 为获取的EXCEL具体的单元格
    excelwb = load_workbook(sourcePath)
    sheet = excelwb._sheets[0]
    return sheet[Cell].value

def getColumnNumForValue(sheet, CellValue):  #根据单元格值查找单元格列数
    maxRow = sheet.max_row
    maxColumn = sheet.max_column
    i = 1
    while i < maxRow:
        j = 1
        while j < maxColumn:
            if sheet.cell(i, j).value == CellValue:
                return j
            j += 1
        i += 1
    return 0

def getRowNumForValue(sheet, CellValue):  #根据单元格值查找单元格行数
    maxRow = sheet.max_row
    maxColumn = sheet.max_column
    i = 1
    while i < maxRow:
        j = 1
        while j < maxColumn:
            if sheet.cell(i, j).value == CellValue:
                return i
            j += 1
        i += 1
    return 0

def getValueFromExcelCellName(sourcePath, CellName, Model):    #Cell 为获取的EXCEL对应列名的数据,model用来确认取的是模板一的还是模板2的
    excelwb = load_workbook(sourcePath, data_only=True)
    sheet = excelwb._sheets[0]
    rowModelNum = 0
    if Model == 0:
        rowModelNum = getRowNumForValue(sheet, '模式1')
    elif Model == 1:
        rowModelNum = getRowNumForValue(sheet, '模式2')
    columnNum = getColumnNumForValue(sheet, '超额服务费')
    if rowModelNum == 0 or columnNum == 0:
        return 0
    return sheet.cell(rowModelNum+2, columnNum).value  #对应的行数为查到的表格位置的下两行

def getRowCountFromExcel(sourcePath):
    excelwb = load_workbook(sourcePath)
    sheet = excelwb._sheets[0]
    return sheet.max_row - 1  #第一行是列名，排除

def main(sourceFolder, TrustCode):
    global DATANOTFOUND
    global cdfp
    global sourceExcel
    global destExcel

    dataBaseData = getCalculationInfo(TrustCode)
    serviceCharge = getServiceChargeByTrustCode(TrustCode)
    ErrorMsg = []
    if len(dataBaseData) == 0:
        ErrorMsg.append("XML文件中TrustCode字段：'{0}'与系统中TrustCode不一致，请确认！".format(TrustCode))
    else:
        for parent, dirnames, filenames in os.walk(sourceFolder, followlinks=True):
            for filename in filenames:
                if len(filename.split('_')) > 1:
                    Type = filename.split('_')[1]
                    sourcePath = os.path.join(parent, filename)
                    if Type == 'TrustInfoImportAndExportModel' and '$' not in filename:
                        OfferAmount = getValueFromXML(sourcePath, 'TrustManagement.TrustBond', 'ItemCode', 'OfferAmount', 'ItemValue') #项目规模
                        CouponBasis = getValueFromXML(sourcePath, 'TrustManagement.TrustBond', 'ItemCode', 'CouponBasis', 'ItemValue') #优先A级利率
                        PoolCloseDate = getValueFromXML(sourcePath, 'TrustManagement.TrustExtension', 'ItemCode', 'PoolCloseDate', 'ItemValue')#资产池封包日
                        TrustStartDate = getValueFromXML(sourcePath, 'TrustManagement.TrustExtension', 'ItemCode', 'TrustStartDate', 'ItemValue') #产品设立日
                        ClosureDate = getValueFromXML(sourcePath, 'TrustManagement.TrustExtension', 'ItemCode', 'ClosureDate', 'ItemValue') #法定到期日
                        TrusteeRadio1 = getValueFromXML(sourcePath, 'TrustManagement.TrustFeeEntity', 'ItemCode', 'TrusteeRemuneration_Fee_Ratio_1', 'ItemValue')#信托通道费
                        CustoRadio1 = getValueFromXML(sourcePath, 'TrustManagement.TrustFeeEntity', 'ItemCode', 'Custodian_Fee_Ratio_1', 'ItemValue')   #托管费
                        CustoRadio2 = getValueFromXML(sourcePath, 'TrustManagement.TrustFeeEntity', 'ItemCode', 'Custodian_Fee_Ratio_2', 'ItemValue') #代销费

                        if OfferAmount == -1:
                            ErrorMsg.append("优先级A发行规模数据在XML文件中不存在，请确认！")
                        elif round(float(OfferAmount), 2) != round(float(dataBaseData[0].OfferAmount), 2):
                            ErrorMsg.append("优先级A发行规模：{0}与项目规模：{1}不相等，请确认！".format(OfferAmount, dataBaseData[0].OfferAmount))
                        if CouponBasis == -1:
                            ErrorMsg.append("优先A票面利率设置数据在XML文件中不存在，请确认！")
                        elif round(float(CouponBasis), 2) != round(float(dataBaseData[0].CouponBasis), 2):
                            ErrorMsg.append("优先A票面利率设置：{0}与优先A级利率：{1}不相等，请确认！".format(CouponBasis, dataBaseData[0].CouponBasis))
                        if TrusteeRadio1 == -1:
                            ErrorMsg.append("信托报酬费率设置数据在XML文件中不存在，请确认！")
                        elif round(float(TrusteeRadio1), 2) != round(float(dataBaseData[0].TrusteeRadio1), 2):
                            ErrorMsg.append("信托报酬费率设置：{0}与信托通道费：{1}不相等，请确认！".format(TrusteeRadio1, dataBaseData[0].TrusteeRadio1))
                        if CustoRadio1 == -1:
                            ErrorMsg.append("托管费费率设置数据在XML文件中不存在，请确认！")
                        elif round(float(CustoRadio1), 2) != round(float(dataBaseData[0].CustoRadio1), 2):
                            ErrorMsg.append("托管费费率设置：{0}与托管费：{1}不相等，请确认！".format(CustoRadio1, dataBaseData[0].CustoRadio1))
                        if CustoRadio2 == -1:
                            ErrorMsg.append("代销费费率设置数据在XML文件中不存在，请确认！")
                        elif round(float(CustoRadio2), 2) != round(float(dataBaseData[0].CustoRadio2), 2):
                            ErrorMsg.append("代销费费率设置：{0}与财顾、代销费：{1}不相等，请确认！".format(CustoRadio2, dataBaseData[0].CustoRadio2))
                        if PoolCloseDate == -1:
                            ErrorMsg.append("资产池封包日(基准日)数据在XML文件中不存在，请确认！")
                        elif PoolCloseDate != dataBaseData[0].TrustStartDate:
                            ErrorMsg.append("资产池封包日(基准日)：{0}与成立日/预计成立日：{1}不相等，请确认！".format(PoolCloseDate, dataBaseData[0].TrustStartDate))
                        if TrustStartDate == -1:
                            ErrorMsg.append("产品设立日数据在XML文件中不存在，请确认！")
                        elif TrustStartDate != dataBaseData[0].TrustStartDate:
                            ErrorMsg.append("产品设立日：{0}与成立日/预计成立日：{1}不相等，请确认！".format(TrustStartDate, dataBaseData[0].TrustStartDate))
                        # if ClosureDate != dataBaseData[0].ClosureDate:
                        #     ErrorMsg.append("法定到期日：{0}与到期日：{1}不相等，请确认！".format(ClosureDate, dataBaseData[0].ClosureDate))

                    elif Type == '存续期归集现金流' and '$' not in filename:
                        TotalBanlance = getValueFromExcel(sourcePath, 'J3')
                        if round(float(TotalBanlance), 2) != round(float(dataBaseData[0].TotalBanlance), 2):
                            ErrorMsg.append("存续期归集现金流 第一行数据的 期初规模列：{0}与最终放款金额（单位：元）：{1}不相等，请确认！".format(TotalBanlance, dataBaseData[0].TotalBanlance))

                    elif Type == '账户收支历史' and '$' not in filename:
                        rowCount = getRowCountFromExcel(sourcePath)
                        if rowCount < 2:
                            ErrorMsg.append("账户收支历史 数据条数为：{0} 小余2条，请补充！".format(rowCount))
                    elif Type == '测算模板' and '$' not in filename:
                        dataCharge = getValueFromExcelCellName(sourcePath, '超额服务费', serviceCharge[0].Model)
                        if round(float(dataCharge), 2) != round(float(serviceCharge[0].ServiceCharge), 2):
                            ErrorMsg.append("超额服务费 测算模板数据为：{0} 与系统数据：{1}不相等，请确认！".format(round(float(dataCharge), 2), round(float(serviceCharge[0].ServiceCharge), 2)))

    # 当存在错误时，输出错误文件
    ErrorLen = len(ErrorMsg)
    DeleteValuesToVerificationLog(TrustCode)
    if ErrorLen > 0:
        TrustId = dataBaseData[0].TrustId
        TrustName = dataBaseData[0].TrustName
        sqlStr = "insert into TaskCollection.dbo.TrustCalculationVerificationLog(TrustId ,TrustCode ,TrustName ,UserId ,ErrorMsg ,TaskType ) values "
        ValueStr = ""
        i = 1
        while i <= ErrorLen:
            ValueStr +="({0},'{1}',N'{2}','{3}',N'{4}','{5}'),".format(TrustId, TrustCode, TrustName, '12B3461F-F3F6-E911-80EE-00155D060656', ErrorMsg[i - 1], 'TrustCalculation')
            i += 1

        exesql = "{0}{1}".format(sqlStr, ValueStr.rstrip(','))   #将报错信息入库，匹配完备注信息在重新输出
        execSQLCmd(exesql)

        dbCheckResult = getValuesFromVerificationLog(TrustCode)

        if len(dbCheckResult) > 0:
            wb = Workbook()  # 新建工作簿
            ws1 = wb.active
            wb.save(errtxtFilePath)
            excelwb = load_workbook(errtxtFilePath)
            logSheet = excelwb['Sheet']
            logSheet["A{0}".format(1)] = '序号'
            logSheet["B{0}".format(1)] = '产品Id'
            logSheet["C{0}".format(1)] = '产品Code'
            logSheet["D{0}".format(1)] = '产品名称'
            logSheet["E{0}".format(1)] = '错误详情'
            logSheet["F{0}".format(1)] = '备注'
            i = 1
            for r in dbCheckResult:
                logSheet["A{0}".format(i + 1)] = i
                logSheet["B{0}".format(i + 1)] = r.TrustId
                logSheet["C{0}".format(i + 1)] = r.TrustCode
                logSheet["D{0}".format(i + 1)] = r.TrustName
                logSheet["E{0}".format(i + 1)] = r.ErrorMsg
                logSheet["F{0}".format(i + 1)] = ''
                i += 1

            excelwb.save(errtxtFilePath)

    else:
        print("校验成功无错误！")




sourceFolder = r'\\172.16.6.143\Products\信托测算\任务标接–平安普惠(4)\陈亚峰\徐思怡\华能信托·丰橙20号集合资金信托计划'
errtxtFilePath = os.path.join(sourceFolder, 'Log_ExtractInsert_{0}.xlsx'.format(datetime.datetime.now().strftime('%m-%d %H%M%S')))
logtxtFilePath = os.path.join(sourceFolder, 'Logs', 'Log_ExtractInsert_{0}.txt'.format(datetime.datetime.now().strftime('%m-%d %H%M%S')))
for parent, dirnames, filenames in os.walk(sourceFolder, followlinks=True):
    for filename in filenames:
        if len(filename.split('_')) > 1:
            Type = filename.split('_')[1]
            sourcePath = os.path.join(parent, filename)
            if Type == 'TrustInfoImportAndExportModel':
                TrustCode = getTrustCodeFromXMl(sourcePath)

                main(sourceFolder, TrustCode)

