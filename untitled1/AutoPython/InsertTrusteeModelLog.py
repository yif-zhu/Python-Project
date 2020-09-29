# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import pyodbc

##### Helper Methods #####
def writeLog(msg):
    print(msg)

def writeErr(msg):
    print(msg)


def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        writeLog(str(ex))
        print(str(ex))
        # raise ex
    finally:
        cnxn.close()


def execSQLCmdFetchAll(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        rows = cursor.execute(sql).fetchall()
        return rows
    except Exception as ex:
        writeLog(str(ex))
    finally:
        cnxn.close()



def cleanOldData():
    sql = "exec DVImport.CleanTrusteeReportModelLog "
    execSQLCmd(sql)


def runDBDataValidation():
    sql = "exec DV.UpdateTrusteeReportModelLog "
    execSQLCmd(sql)


##### Extract Implemention Methods #####
#return: cellValue, isNA
def getCellValue(cvalue, cdtype):
    if ((cdtype == 'number' or cdtype == 'rate') and isinstance(cvalue, float)
       ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return str(cvalue), 0

    #cvalue = str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '')
    if cvalue == 'NA' or cvalue == '-':
        return cvalue, 1

    return cvalue, 0

def specificColsExtract(cfgItem, sheet):
    statement = cfgItem.attrib['stat']
    itemdesc = cfgItem.attrib['desc']
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    tmpl = ""
    isFirstRow = 1
    firstRowAllNA = 1
    # breakRowsLoop = 0
    while rStart <= rEnd:  # rows loop
        rvalues = ''

        for cell in cNode:  # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            crv, isNA = getCellValue(cvalue, cdtype)

            if isFirstRow != 1 and crv == '' and cemptybreak == '1':
                rvalues = ''
                break  # break cells loop

            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            if crv == 'NULL':
                rvalues += "{0},".format(crv)
            else:
                rvalues += "N'{0}',".format(crv)


        if (isFirstRow == 1 and firstRowAllNA == 1):
            rStart = rEnd + 1  # break rows loop
        else:
            isFirstRow = 0
            if rvalues != "":
                tmpl += "({0}),".format(rvalues.rstrip(','))
            rStart += 1

    if tmpl == "":
        writeLog("【无数据提示】 [{0}] 数据全部NA".format(itemdesc))
        return ''

    return "{0}{1}".format(statement, tmpl.rstrip(','))

def extractExcel(wb, cfgItems):
    sheet = wb._sheets[0]

    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']

        if not 'stat' in cfgItem.attrib:
            msg = '【程序配置错误】config.xml中[{0}]节点中未配置statement，节点无法操作!'.format(itemdesc)
            writeLog(msg)
            continue

        exesql = ""

        exesql = specificColsExtract(cfgItem, sheet)

        if exesql != '':
            writeLog("【{0}-已提取】开始提交 [{1}] 至数据库".format(i, itemdesc))
            execSQLCmd(exesql)
        else:
            writeLog("【{0}-未提取】未获取不提交 [{1}] ".format(i, itemdesc))

################### Config Settings and Source Folder ###################
#scriptFolder = sys.argv[0].replace(sys.argv[0].split('\\')[-1], '')
scriptFolder = os.getcwd()
configFilePath = os.path.join(scriptFolder, 'InsertTrusteeModelLog_setting.xml')
mappingTree = XETree.parse(configFilePath)

cfgRoot = mappingTree.getroot()
sourceFilePath =cfgRoot.attrib['sourceFilePath']
dbConnectionStr = cfgRoot.attrib['dbconnstr']

excelwb = load_workbook(sourceFilePath)
cleanOldData()
extractExcel(excelwb, cfgRoot)

runDBDataValidation()


