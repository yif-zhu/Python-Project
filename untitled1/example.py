import os
import re
import os.path
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
import  shutil
sourceExcel = None
destExcel = None
xmlPath = None

def updateValue(value,datatype):  #根据类型进行字符串处理，D日期，P百分数，S字符串不做处理，F或不填按数字处理
    try:
        if value is None:
            return ''
        # if value is None:
        #     if datatype == 'F':
        #         return  "0"
        #     elif datatype == "P":
        #         return  "0.00%"
        #     else:
        #         return  ""
        value = str(value)
        value = value.replace(' ','')
        if datatype == 'D':
            #dateFormat = '[^\d/\d/\d]'
            #'2016年3月1日至2016年3月31日'
            value = re.sub('.*至', '', value)
            value = re.sub('\D$','', value)
            value = re.sub(r'\D', r'/', value)
        elif datatype =='P':
            if type(value) == float :
                value = str(value * 100) + '%'
            else:
                Persentage = '[^\d%|\d.\d%]'
                value = re.sub(Persentage, '', value)
        elif datatype =='S':
            return value
        else:
            floatFormat = '[^\d|\d.\d]'
            value = re.sub(floatFormat, '', value)
            if value == '' or value == '.':
                value = 0
            value = float(value)
        return value
    except ValueError as e:
        print("ErrorValue: " + value + "Type: " + datatype)

def checkData(value,datatype): #值校验，是否符合对应类型
    value = str(value)
    date = '^\d{4}\D\d{1,2}\D\d{1,2}\D?$'
    num ='^[-]?\d*[.]?\d*$'
    persentige= '^[-]?\d*[.]?\d*[%]$'
    if value is None or value == '':
        return False
    flag = 0
    if datatype == 'D':
        flag = re.search(date, value)
    elif datatype == 'P':
        flag = re.search(persentige, value)
    elif datatype == 'S':
        return True
    else:
        flag = re.search(num, value)
    if str(flag) == 'None':
        return False
    else:
        return True
def extractForTable(cfgItem):
    endrow = 0
    beginrow = 0
    # sheet_names = sourceExcel.get_sheet_names()
    sheet = sourceExcel["Sheet1"]

    # sheet_names = destExcel.get_sheet_names()
    destSheet = destExcel["Sheet1"]
    tempcell = findStr(sheet, cfgItem[0].attrib["anchor"], 1)
    destBeginRow = int(cfgItem[1].attrib["beginrow"])
    sList = cfgItem[0].attrib["cols"].split(',')
    dList = cfgItem[1].attrib["cols"].split(',')
    datatype = cfgItem[1].attrib["datatype"].split(',')

    #如果差找不到，首行设置为NA
    if tempcell is None:
        for col in dList:
            destSheet[col + str(destBeginRow)].value = "NA"
        return
    # 判断在源文件中的起始行
    beginrow = tempcell.row + int(cfgItem[0].attrib["skiprows"]) + 1

    # 判断结束行
    # 如果范围已经确定，直接确定结束行
    if cfgItem[0].attrib["range"] != "":
        endrow = beginrow + int(cfgItem[0].attrib["range"]) - 1
    # 范围不确定，根据下一行字符确定结束行
    elif cfgItem[0].attrib["anchorend"] != "":
        endrow = findStr(sheet, cfgItem[0].attrib["anchorend"], beginrow).row
    else:
        # 找不到字符，则直到最后一个不为空行的为止
        limited = 1
        while sheet["A" + str(beginrow + limited)].value != "" and sheet["A" + str(beginrow + limited)].value is not None:
            limited += 1
        endrow = beginrow + limited - 1

    print(cfgItem[0].attrib["anchor"])
    # 粘贴数据
    for row in range(beginrow, endrow + 1):
        for col in range(0, len(sList)):
            try:
                tempvalue = updateValue(sheet[sList[col] + str(row)].value, datatype[col])
            except ValueError as e:
                print(sList[col] + str(row))
            isRight = checkData(tempvalue,datatype[col])
            if not isRight:
                fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
                #destSheet[dList[col] + str(destBeginRow + row - beginrow)].fill = fill
                destSheet[dList[col] + str(destBeginRow + row - beginrow)].value = tempvalue
            else:
                destSheet[dList[col] + str(destBeginRow + row - beginrow)].value = tempvalue


def findStr(sheet, key, startRow):
    for row in sheet.iter_rows(min_row=startRow):
        for cell in row:
            if cell.value is not None and cell.value != "":
                if key in str(cell.value).replace(' ', ''):
                    return cell
def keyValueToDestExcel(value, dCols, dNode):
    sheetD = destExcel['Sheet1']
    datatype = 'F'
    if 'datatype' in dNode.attrib:
        datatype = dNode.attrib['datatype']
    value = updateValue(value, datatype)
    if 'value' in dNode.attrib:
        value = dNode.attrib['value']
    isRight = checkData(value, datatype)
    if not isRight:
        fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        sheetD[dCols].fill = fill
        value = 0
    sheetD[dCols] = value

def extractForKeyValue(cfgItem):
    sheetS = sourceExcel['Sheet1']

    sNode = cfgItem.find('source')
    dNode = cfgItem.find('dest')
    dCols = dNode.attrib['cols']
    anchors = sNode.attrib['anchor'].strip().split(';')
    length = len(anchors)
    index = 0
    beginCol = 'A'
    sCols = sNode.attrib['cols']
    isFind = 0;
    for i in range(sheetS._current_row):
        text = sheetS[beginCol + str(i+1-index)].value
        if text is None:
            continue
        if str(anchors[index]) in text.replace(' ',''):
            beginCol = chr(ord(beginCol)+1)
            index += 1
            if index >= length:
                isFind =1
                value = sheetS[sCols + str(i-index+2)].value
                keyValueToDestExcel(value, dCols, dNode)
                break;
    if isFind ==0:
        keyValueToDestExcel('', dCols, dNode)

def selectType(cfgItems):
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        istable = cfgItem.attrib['istable'].lower()
        if istable == 'true':
            extractForTable(cfgItem)
        else:
            extractForKeyValue(cfgItem)

def chooseExcel(inputFile, outputFile, templateFile, cfgRoot ):
    global sourceExcel
    global destExcel
    if os.path.exists(outputFile):
        os.remove(outputFile)
    open(outputFile, "wb").write(open(templateFile, "rb").read())
    sourceExcel = load_workbook(inputFile)
    destExcel = load_workbook(outputFile)
    selectType(cfgRoot)
    destExcel.save(outputFile)


def main():
    global DATANOTFOUND
    global cdfp
    global sourceExcel
    global destExcel
    configFilePath = r'C:\PyCharm\xml\444.xml'


    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    input = cfgRoot.attrib['input']
    output = cfgRoot.attrib['output']
    templateFilePath = cfgRoot.attrib['template']

    if os.path.isfile(input):
        filename = input.split('\\')[-1]
        outputFile = os.path.join(output, filename)
        chooseExcel(input, outputFile, templateFilePath, cfgRoot)
    elif os.path.isdir(input):
        for parent, dirnames, filenames in os.walk(input, followlinks=True):
            for filename in filenames:
                inputFile = os.path.join(parent, filename)
                outputFile = os.path.join(output, filename)
                chooseExcel(inputFile, outputFile, templateFilePath, cfgRoot)

main()
