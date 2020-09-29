# 每个tab包含两个view用于显示源文件与目标文件
import os
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from Action.Action import Action
import openpyxl
import re
import xml.etree.cElementTree as XETree
import win32clipboard
import win32con
import numpy as np
import sip
import traceback


class Tab(QWidget):
    essentialColor = QColor(255, 230, 153)
    isChanged = None
    logGenerated = pyqtSignal(str)
    signal_changed = pyqtSignal(QWidget, bool)
    clipboardFile = "Config/clipBoard.et"

    def __init__(self, actionCode, acitonName, parameters):
        super().__init__()
        self.actionCode = actionCode
        self.actionName = acitonName
        self.parameters = parameters
        self.initUI()

    def initUI(self):
        pass

    def run(self):
        action = Action(self.actionCode, self.parameters)
        log_str = ""
        for (k, v) in self.parameters.items():
            log_str = log_str + "\n" + k + ":  " + v
        self.logGenerated.emit("开始执行任务：{0}".format(self.actionName))
        outPutFilePath = self.parameters["outputFile"]
        # self.signal_changed.emit(self, True)
        result = action.runAciton()
        fileName = os.path.basename(result).split('.')[0]
        LogPath = ''
        if self.actionCode == 'FirstCheck':
            LogPath = os.path.join(outPutFilePath, 'Error_第一步校验错误_{0}.txt'.format(fileName))
        elif self.actionCode == 'SecondCheck':
            LogPath = os.path.join(outPutFilePath, 'Error_第二步校验错误_{0}.txt'.format(fileName))
        elif self.actionCode == 'ThirdCheck':
            LogPath = os.path.join(os.getcwd(), "FileTranslator", 'Logs', '{0}.txt'.format(fileName))
        elif self.actionCode == 'CompareCNABS':
            LogPath = os.path.join(os.getcwd(), "FileTranslator", 'Logs', '{0}.txt'.format(fileName))
        elif self.actionCode == 'BaseInfoCheck':
            pass
        #LogPath = os.path.join(os.getcwd(), "FileTranslator", 'Logs', '{0}.txt'.format(fileName))
        # result = "FileTranslator/MappingXml/20190720_182549.xml"
        if os.path.exists(LogPath):
            log = open(LogPath)
            for line in log:
                self.logGenerated.emit(line)
            log.close()
        else:
            self.logGenerated.emit("该文件夹内文件无错误，校验已完成！")
        self.logGenerated.emit("执行完成\n\n")

        #root = XETree.parse(result).getroot()


class Sheet(QTableWidget):
    rowsMark = [int]
    mouseReleased = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.__hasload = False
        self.filePath = ""
        self.sheetIndex = None
        self.sheetName = None
        self.rowsMark.append(0)
        self.horizontalScrollBar().setStyleSheet("height:25px;background-color: rgb(222, 222, 222);")
        self.verticalScrollBar().setStyleSheet("width:25px;background-color: rgb(222, 222, 222);")
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)

    def init(self):
        self.setColumnCount(10)
        self.setRowCount(30)
        for i in range(0, 10):
            item = TableItem(value=intToletter(i))
            self.setHorizontalHeaderItem(i, item)
        for i in range(0, 30):
            item = TableItem(value=str(i + 1))
            self.setVerticalHeaderItem(i, item)

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:
        super().mouseReleaseEvent(event)
        self.mouseReleased.emit()
        event.accept()

    # def fillSheetByExcelSheetIndex(self, filepath, sheetIndex):
    #
    #     try:
    #         workbok = openpyxl.load_workbook(filepath)
    #         sheetNames = workbok.sheetnames
    #         self.__fileSheetByWorkBook(workbok[sheetNames[sheetIndex]])
    #         self.filePath = filepath
    #         self.sheetName = sheetNames[sheetIndex]
    #     except Exception as e:
    #         QMessageBox.critical(self, "错误", "解析文件发生错误: \n" + traceback.format_exc(), QMessageBox.Ok)
    #
    # def fillSheetByExcelSheetName(self, filepath, sheetName):
    #     try:
    #         workbok = openpyxl.load_workbook(filepath)
    #         self.__fileSheetByWorkBook(workbok[sheetName])
    #         self.filePath = filepath
    #         self.sheetName = sheetName
    #
    #     except Exception as e:
    #         QMessageBox.critical(self, "错误", "解析文件发生错误：\n" + traceback.format_exc(), QMessageBox.Ok)

    # 设置item位置超过table大小时自动更新大小
    def setItem(self, row: int, column: int, item: QTableWidgetItem) -> None:
        # 例如当前row = 9，实际rowcount = 9，rowcount应该改为10， 并给新hearder设置值
        current_rowCount = self.rowCount()
        current_columnCount = self.columnCount()
        if row > current_rowCount - 1:
            self.setRowCount(row + 1)
            for i in range(current_rowCount, row + 1):
                self.setVerticalHeaderItem(i, TableItem(str(i + 1)))
        if column > current_columnCount - 1:
            self.setColumnCount(column + 1)
            for i in range(current_columnCount, column + 1):
                self.setHorizontalHeaderItem(i, TableItem(intToletter(i)))
        super().setItem(row, column, item)

    def setCellWidget(self, row: int, column: int, widget: QWidget) -> None:
        current_rowCount = self.rowCount()
        current_columnCount = self.columnCount()
        if row > current_rowCount - 1:
            self.setRowCount(row + 1)
            for i in range(current_rowCount, row + 1):
                self.setVerticalHeaderItem(i, TableItem(str(i + 1)))
        if column > current_columnCount - 1:
            self.setColumnCount(column + 1)
            for i in range(current_columnCount, column + 1):
                self.setHorizontalHeaderItem(i, TableItem(intToletter(i)))
        super().setCellWidget(row, column, widget)

    # def __fileSheetByWorkBook(self, sheet):
    #     row_count = sheet.max_row
    #     col_count = sheet.max_column
    #     if self.rowCount() < row_count:
    #         self.setRowCount(row_count)
    #     if self.columnCount() < sheet.max_column:
    #         self.setColumnCount(col_count)
    #     for i in range(0, col_count):
    #         if self.horizontalHeaderItem(i) is None:
    #             item = TableItem(value=intToletter(i))
    #             self.setHorizontalHeaderItem(i, item)
    #     for i in range(0, row_count):
    #         if self.verticalHeaderItem(i) is None:
    #             item = TableItem(value=str(i + 1))
    #             self.setVerticalHeaderItem(i, item)
    #
    #     # 开始初始化表格
    #     for row in sheet.iter_rows(min_row=0):
    #         for cell in row:
    #             item = TableItem("")
    #             if type(cell.column) == str:
    #                 self.setItem(cell.row - 1, letterToint(cell.column), item)
    #             elif type(cell.column) == int:
    #                 self.setItem(cell.row - 1, cell.column - 1, item)
    #             # color = cell.fill.bgColor.rgb
    #             # if color is not None and color != "00000000" and type(color) != openpyxl.styles.colors.RGB:
    #             #     color = HexToRgb(color)
    #             #     item.setBackground(QColor(color["r"], color["g"], color["b"]))
    #             if cell.value is not None and cell.value != "":
    #                 item.setText(str(cell.value))
    #
    #     # 合并单元格
    #     merged_cell = str(sheet.merged_cells).split(" ")
    #     for cell in merged_cell:
    #         if cell != "":
    #             startPoint = cell.split(":")[0]
    #             endPoint = cell.split(":")[1]
    #             startRow = int(re.sub(r"\D", "", startPoint)) - 1
    #             startCol = letterToint(re.sub(r"[^A-Z]", "", startPoint))
    #             endRow = int(re.sub(r"\D", "", endPoint)) - 1
    #             endCol = letterToint(re.sub(r"[^A-Z]", "", endPoint))
    #             self.setSpan(startRow, startCol, endRow - startRow + 1, endCol - startCol + 1)
    #     self.__hasload = True

    # def pasteData(self, startrow, startcol, data, format=True):
    #     i = 0
    #     for row in data:
    #         for j in range(len(row)):
    #             item = self.item(startrow + i, startcol + j)
    #             if item is None:
    #                 item = TableItem("")
    #                 self.setItem(startrow + i, startcol + j, item)
    #             if format:
    #                 item.setFormatValue(row[j])
    #             else:
    #                 item.setText(row[j])
    #         i += 1

    def hasLoad(self):
        return self.__hasload

    def setColumnCount(self, columns: int) -> None:
        current_column = self.columnCount()
        super().setColumnCount(columns)
        if columns > current_column:
            for i in range(current_column, columns):
                self.setColumnWidth(i, 110)

    def save(self):
        workbok = openpyxl.load_workbook(self.filePath)
        sheet = workbok[self.sheetName]
        for col in range(0, self.columnCount()):
            for row in range(0, self.rowCount()):
                cell = self.item(row, col)
                if cell is None:
                    continue
                if cell.text() is not None and cell.text() != "":
                    try:
                        value = cell.text()
                        if cell.dataType == "F":
                            value = float(value)
                        sheet[intToletter(col) + str(row + 1)].value = value
                    except Exception as e:
                        print(intToletter(col) + str(row))
                        QMessageBox.critical(self, "错误", traceback.format_exc(), QMessageBox.Ok)
        workbok.save(self.filePath)


class TableItem(QTableWidgetItem):
    dataType = "S"
    IsSaved = True
    errorColor = QBrush(QColor(255, 114, 116))
    noramlColor = QBrush(QColor(255, 255, 255))
    essentialColor = QColor(255, 230, 153)

    def __int__(self):
        super().__init__()
        self.__init()

    def __init__(self, value):
        super().__init__()
        self.__init()
        self.setText(value)

    def __init(self):
        self.setForeground(QColor(0, 0, 0))

    # def __init__(self, dataType, value):
    #     super().__init__()
    #     self.setFormatValue(value)
    def getDataType(self):
        return self.dataType

    def setDataType(self, dataType):
        self.dataType = dataType


# 列字母转数字
def letterToint(s):
    letterdict = {}
    for i in range(26):
        letterdict[chr(ord('A') + i)] = i + 1
    output = 0
    for i in range(len(s)):
        output = output * 26 + letterdict[s[i]]
    return output - 1


# 数字转列字母
def intToletter(i):
    if type(i) is not int:
        return i
    str = ''
    i += 1
    while (not (i // 26 == 0 and i % 26 == 0)):

        temp = 25

        if (i % 26 == 0):
            str += chr(temp + 65)
        else:
            str += chr(i % 26 - 1 + 65)

        i //= 26
        # print(str)
    # 倒序输出拼写的字符串
    return str[::-1]


def parseChar(string):
    letter = re.sub(r"[^A-Z]", "", string)
    num = int(re.sub(r"\D", "", string)) - 1
    return "{0},{1}".format(str(num), letterToint(letter))


# 十六进制颜色编码转rgb
def HexToRgb(tmp):
    rgb = dict()
    opt = re.findall(r'(.{2})', tmp)
    rgb["r"] = 255 - int(opt[0], 16)
    rgb["g"] = 255 - int(opt[1], 16)
    rgb["b"] = 255 - int(opt[2], 16)
    return rgb


# rgb字符串转Qcolor对象
def getColor(colorString):
    r = int(colorString.split(",")[0])
    g = int(colorString.split(",")[1])
    b = int(colorString.split(",")[2])
    return QColor(r, g, b)


def indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


def killWidget(widget):
    for item in widget.children():
        if len(item.children()) != 0:
            killWidget(item)
        else:
            sip.delete(item)
    sip.delete(widget)


# 窗口居中显示
def center(dialog, width, height):
    screen = QDesktopWidget().screenGeometry()
    # +120 稍微向上平移
    dialog.move((screen.width() - width) / 2, (screen.height() - height) / 2 - 120)



