# -*- coding: utf-8 -*-
"""
Created on Tue Dec 17 13:53:35 2019

@author: HUAWEI
"""

# -*- coding: utf-8 -*-
"""
Created on Sun Dec 15 14:43:23 2019

@author: HUAWEI
"""

#%%
def ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,ErrorInformation):
    ErrorPath=r"C:\Users\HUAWEI\Desktop\Error\TrustError.xlsx"
    wb = openpyxl.load_workbook(ErrorPath) # 读取xlsx文件
    ws=wb['Sheet1']
    maxrow=ws.max_row
    maxrow+=1
#    print(maxrow)
    ws.cell(maxrow,1,TrustId).value
    ws.cell(maxrow,2,TrustCode).value
    ws.cell(maxrow,3,TrustName).value
    ws.cell(maxrow,4,Excelfilepath).value
    ws.cell(maxrow,5,ErrorInformation).value
    wb.save(ErrorPath)
    
    
def Trust(Excelfilepath,TrustCode,TrustName):
#    print(Excelfilepath)
    global TrustId,TrustStartDate,PoolCloseDate,ClosureDate,IssueAmount
    conn = pymssql.connect(host='172.16.7.130\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    try:
        selectId="select TrustId from TrustManagement.Trust where TrustCode='{}'".format(TrustCode)
        b1.execute(selectId)
        TrustId=b1.fetchone()[0]
    except:
        A0='文件TrustCode与系统TrustCoden不匹配请检查!'
        
        ErrorMessage('#',TrustCode,TrustName,Excelfilepath,A0)
        print(TrustName,'文件TrustCode与系统TrustCoden不匹配请检查!')
#        sys.exit(1)
        return

    DataTrust=pd.read_excel(Excelfilepath,sheet_name="Trust")

    #获取产品主体
    ProductSubject1=DataTrust.loc[0][5]
    #获取名称
    TrustName=DataTrust.loc[0][0]
    #获取发行金额
    IssueAmount=DataTrust.loc[0][7]
    #获取设立日
    TrustStartDate=DataTrust.loc[0][12]
#    print(TrustStartDate)
    #获取封包日
    PoolCloseDate=DataTrust.loc[0][11]
#    print(PoolCloseDate)
#    print(type(PoolCloseDate))
    #获取法定到期日
    ClosureDate=DataTrust.loc[0][13]
#    print(ClosureDate)

    selectProductSubject="select ItemCode from TrustManagement.TrustInfoExtension where TrustId={}".format(TrustId)
    b1.execute(selectProductSubject)
    SItemCode=b1.fetchall()
    TrustInfoExtension=[]
    for st in SItemCode:
        TrustInfoExtension.append(st[0])
        
    if 'ProductSubject' not in TrustInfoExtension:
        ProductSubjectInsert="insert into TrustManagement.TrustInfoExtension values({},GETDATE(),NULL,NULL,'ProductSubject',N'{}')".format(TrustId,ProductSubject1)
        b1.execute(ProductSubjectInsert)
        conn.commit()
        T0='{},{}产品主体插入完成!'.format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T0)
        print(T0)
    else:
        T0='{},{}产品主体已存在!'.format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T0)
        print(T0)
        
    
    selectIssueAmount="select IssueAmount from TrustManagement.Trust where TrustId={}".format(TrustId)
    b1.execute(selectIssueAmount)
    SQLIssueAmount=b1.fetchone()[0]
    if SQLIssueAmount==None:
        UpdateIssueAmount="update TrustManagement.Trust set IssueAmount={} where TrustId={}".format(IssueAmount,TrustId)
        b1.execute(UpdateIssueAmount)
        conn.commit()
        T1="{},{},'发行金额更新完成!'".format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T1)
        print(T1) 
    else:
        T1='{},{}发行金额已存在!'.format(TrustId,TrustName)
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,T1)
        print(T1)
        
        
def TrustBond(Excelfilepath,TrustCode):

    Error=0
    try:
        DataTrustBond1=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
    except:
        pass
        print('无TrustBond')
        return
#    try:
#        DataPrincipalSchedule=pd.read_excel(Excelfilepath,sheet_name='PrincipalSchedule')
#        huanben=DataPrincipalSchedule.loc[2][2]
#        print(huanben)
#    except:
#        print('无还本计划!')
#        pass
    
    try:
        DataTrustBond=DataTrustBond1[['TrustBondID','ItemID','ItemCode','ItemValue']]
    except:
        print('TrustBond字段错误!')
        return 
    
    DataTrustBond.dropna(subset=['ItemCode','ItemID'],inplace=True)
#    DataTrustBond.dropna(subset=['ItemID'],inplace=True)
    if DataTrustBond['TrustBondID'].count()>25 and DataTrustBond['TrustBondID'].sum()==0:
        A1="TrustBond表分层标识(TrustBondID)重复'"
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A1)
        print(TrustId,TrustName,'分层标识重复(TrustBondID)')
        Error=1
    AmountC=0
    
    ZClass=0
    for i in DataTrustBond.index:       
        TrustBondId=DataTrustBond.loc[i][0]
        ItemId=int(DataTrustBond.loc[i][1])
        ItemCode=DataTrustBond.loc[i][2]
        ItemValue=DataTrustBond.loc[i][3]
        
        if ItemCode=='SecurityExchangeCode' and ('.' or 'IB') in str(ItemValue):
            A2="{},{},'不能包含英文字符'".format(TrustBondId,ItemCode)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A2)
            print(TrustId,TrustName,TrustBondId,'债券代码不能包含英文字符')
            Error=1
            
        if ItemCode=='SecurityExchangeCode' and pd.isnull(ItemValue)==True:
            Error=1
            AJJ1='证券代码不能为空'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ1)
            print(TrustId,TrustName,AJJ1)
            
        
        if TrustBondId==0 and ItemCode=='ClassType' and ItemValue!='FirstClass':
            A3='优先级层级标识填写错误，请检查!'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A3)
            print(TrustId,TrustName,'优先级层级标识填写错误，请检查!')
            Error=1
            
        if DataTrustBond.TrustBondID.drop_duplicates().count()==1 and ItemCode=='ClassType' and ItemValue !='FirstClass':
            A4='{},债券层级填写错误'.format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A4)
            print(TrustId,TrustName,'{},债券层级填写错误'.format(TrustBondId))
            Error=1
            
        if DataTrustBond.TrustBondID.drop_duplicates().count()==2 and ItemCode=='ClassType' and ItemValue not in ('FirstClass','EquityClass'):
            A5='{},债券层级填写错误'.format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A5)
            print(TrustId,TrustName,'{},债券层级填写错误'.format(TrustBondId))
            Error=1
            
        if ItemCode=='CouponBasis' and type(ItemValue) not in (int,float):
            A6="{},票面利率(CouponBasis)数据类型应该为int、float".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A6)
            print(TrustId,TrustName,A6)
            Error=1
            
            
        if ItemCode=='ClassType' and ItemValue in ('FirstClass','SubClass'):
            ZClass=1

        if ItemCode=='CouponBasis' and ItemValue==0 and ZClass==1:
            print('ZCass:',ZClass)
            Error=1
            AJ='优先级或次优先级票面利率不能为0,请检查!'
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJ)
            print(TrustId,TrustName,'优先级或次优先级票面利率不能为0')
            
        if ItemCode=='CouponBasis' and ItemValue!=0 and ZClass==1:
            ZClass=0
            
        
        
        if ItemCode=='OfferAmount' and type(ItemValue) not in (int,float,np.float,np.float64,np.float32,np.float16):
            
            Error=1
            print(ItemValue)
            print(type(ItemValue),pd.isnull(ItemValue))
            A7="{},发行规模(OfferAmount)数据类型应该为int、float".format(TrustBondId)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A7)
            print(TrustId,TrustName,"{},发行规模(OfferAmount)数据类型应该为int、float且不能为空".format(TrustBondId))
            
            
            
            
        if ItemCode=='ClassType' and ItemValue not in ('FirstClass','SubClass','EquityClass'):
            A8="{},债券类别(ClassType)填写值为[{}]错误!只能填写'FirstClass','SubClass','EquityClass'".format(TrustBondId,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A8)
            print(TrustId,TrustName,"{},债券类别(ClassType)填写值为[{}]错误!只能填写'FirstClass','SubClass','EquityClass'".format(TrustBondId,ItemValue))
            Error=1

        
        if ItemCode=='InterestPaymentType' and ItemValue not in (1,3,6,12):
            A9="{},付息频率(InterestPaymentType)填写值[{}]错误!只能填写1,3,6,12".format(TrustBondId,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A9)
            print(TrustId,TrustName,"{},付息频率(InterestPaymentType)填写值[{}]错误!只能填写1,3,6,12".format(TrustBondId,ItemValue))
            Error=1


        if ItemCode=='InterestRateCalculation' and ItemValue not in ('按天','按月','按半年','按年'):
            A10="{},计息方式(InterestRateCalculation)填写值[{}]错误!只能是'按天','按月','按半年','按年'".foramt(TrustBondId,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A10)
            print(TrustId,TrustName,"{},计息方式(InterestRateCalculation)填写值[{}]错误!只能是'按天','按月','按半年','按年'".foramt(TrustBondId,ItemValue))
            Error=1

                
        if ItemCode=='PaymentConvention' and ItemValue not in ('固定摊还','到期一次性还本付息','到期一次性还本','过手摊还','按期等额本金','按期等额本息','到期一次还本付息','到期一次性兑付','到期获取剩余收益') and pd.isnull(ItemValue)==True:
            
            A11="{},还本付息方式(PaymentConvention)填写值[{}]错误!只能为'固定摊还','一次性还本付息','到期一次性还本','过手摊还','按期等额本金','按期等额本息'".format(TrustBondId,ItemValue)
            
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A11)
            print(TrustId,TrustName,A11)
            Error=1
            
#        if ItemCode=='PaymentConvention' and ItemValue=='固定摊还' and pd.isnull(huanben)==True:
#            AJJ='{},固定摊还应有还本计划'.format(TrustBondId)
#            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,AJJ)
#            print(TrustId,TrustName,AJJ)
#            Error=1
        

        if ItemCode=='CouponPaymentReference' and ItemValue not in ('浮动利率','固定利率'):
            A12="利率形式(CouponPaymentReference)填写值[{}]错误!只能为'浮动利率','固定利率'".format(ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A12)
            print(TrustId,TrustName,"利率形式(CouponPaymentReference)填写值[{}]错误!只能为'浮动利率','固定利率'".format(ItemValue))
            Error=1
            
            
        if ItemCode in ('OriginalCreditRating','ClassName') and ItemValue not in ('A','AA','AAA',np.nan,'NR','A+','AA+','AA-'):
            A13='{},债券评级填写值【{}】错误!'.format(TrustBondId,ItemValue)
            ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A13)
            print(TrustId,TrustName,'{},债券评级填写值【{}】错误!'.format(TrustBondId,ItemValue))
            Error=1
            
        try:
            
            if ItemCode=='OfferAmount':
                Amount=ItemValue
                AmountC=AmountC+Amount
        except:
            pass
            print('数据类型错误')
                
                
    if  IssueAmount-AmountC>1000 or IssueAmount-AmountC<-1000:  #7
        A14='发行金额{}与层级金额{}相加不符'.format(IssueAmount,AmountC)
        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A14)
        
        print(TrustId,TrustName,'发行金额{}与层级金额{}相加不符'.format(IssueAmount,AmountC))
        Error=1
    print(Error)
    
    if Error==0:
        A15='TrustBond表校验通过!'
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,A15)
        print(TrustId,TrustName,'TrustBond表校验通过!')
        TrustBondImport(Excelfilepath,TrustId,TrustName)
        
        
def TrustBondImport(Excelfilepath,TrustId,TrustName):
    
    conn = pymssql.connect(host='172.16.7.130\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    selectBondId="select TrustId from TrustManagement.TrustBond2"
    
    b1.execute(selectBondId)
    TrustBondId=b1.fetchall()
    conn.commit()
    BondId=[]
    for Bond in TrustBondId:
        BondId.append(Bond[0])
    if TrustId not in BondId:
        DataTrustBond=pd.read_excel(Excelfilepath,sheet_name='TrustBond')
        DataTrustBond.dropna(subset=['ItemCode'],inplace=True)
        DataTrustBond.dropna(subset=['ItemID'],inplace=True)
        del DataTrustBond['Unnamed: 5']
        del DataTrustBond['StartDate']
        del DataTrustBond['EndDate']
        for i in DataTrustBond.index:
            TrustBondId=DataTrustBond.loc[i][0]
            ItemId=int(DataTrustBond.loc[i][1])
            ItemCode=DataTrustBond.loc[i][2]
            ItemValue=DataTrustBond.loc[i][3]
    #            print(TrustBondId,ItemId,ItemCode,ItemValue)            
            InsertBond="insert into TrustManagement.TrustBond2(TrustBondId,TrustId,StartDate,EndDate,ItemId,ItemCode,ItemValue) values({},{},GETDATE(),NULL,{},'{}',N'{}')".format(TrustBondId,TrustId,ItemId,ItemCode,ItemValue)
        #    print(sql)
            b1.execute(InsertBond)
        conn.commit()
        BI='TrustBond导入完成!'
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,BI)
        print(TrustId,TrustName,'TrustBond导入完成!')
        
    else:
        BI='数据库已存在TrustBond数据跳过导入!'
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,BI)
        
        print(TrustId,TrustName,BI)
        
def TrustExtensionImport(Excelfilepath):
    conn = pymssql.connect(host='172.16.7.130\mssql', user='sa', password='PasswordGS2017',
                       database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    selectTrustExtension="select TrustId from TrustManagement.TrustExtension1"
    b1.execute(selectTrustExtension)
    STrustExtensionId=b1.fetchall()
    conn.commit()
    TrustExtensionId=[]
    for GTrustExtensionId in STrustExtensionId:
        TrustExtensionId.append(GTrustExtensionId[0])
    if TrustId not in TrustExtensionId:
        TrustExtension=pd.read_excel(Excelfilepath,sheet_name='TrustExtension',header=None)
        TrustExtension=TrustExtension[[2,3]]
        TrustExtension=TrustExtension[1:]
        
        for tindex in TrustExtension.index:
            TrustExtension_ItemCode=TrustExtension.loc[tindex][2]
            TrustExtension_ItemValue=TrustExtension.loc[tindex][3] 
            Itemid_sql="select ItemId from TrustManagement.Item where ItemCode='{}'".format(TrustExtension_ItemCode)
            b1.execute(Itemid_sql)
            ItemId=b1.fetchone()[0]
    
            TrustExtensionInsert="insert into TrustManagement.TrustExtension1(TrustId,StartDate,EndDate,ItemId,ItemCode,ItemValue) values({},GETDATE(),NULL,{},'{}','{}')".format(TrustId,ItemId,TrustExtension_ItemCode,TrustExtension_ItemValue)
                
            b1.execute(TrustExtensionInsert)
        conn.commit()
        EI='TrustExtension导入完成!'
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,EI)
        print(TrustId,TrustName,EI)
        
    else:
        EI='TrustExtension表数据库已存在跳过上传!'
#        ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,EI)
        print(TrustId,TrustName,EI)
            
    
    
      
        
if __name__=="__main__":
    import xlrd,xlwt
    import pandas as pd,numpy as np
    import numpy as np
    import os
    import time
    import sys
    import datetime
    import pymssql
    import openpyxl
    
    
    path=r'\\172.16.6.143\StudentsProducts\1210网商花呗受托报告\花呗管理人投后报告汇总\基础表\3\光大花呗8期2'
    
    for i in os.listdir(path):
        
        if '基础表.xlsx' in i:
            
            split=i.split('_')
            TrustCode=split[0]
            TrustName=split[1]
            Excelfilepath=os.path.join(path,i)
            Trust(Excelfilepath,TrustCode,TrustName)
            TrustBond(Excelfilepath,TrustCode)
    
    
#    path=r'\\172.16.6.143\StudentsProducts\1210网商花呗受托报告\花呗管理人投后报告汇总\基础表\1'
#    
#    for filej in os.listdir(path):
#        Pfilej=os.path.join(path,filej)
#        for file in os.listdir(Pfilej):
#            if '基础表.xlsx' in file:
#                
#                split=file.split("_")
#                
#                TrustCode=split[0]
#                TrustName=split[1]
##                try:
#                    
#                Excelfilepath=os.path.join(Pfilej,file)
#                Trust(Excelfilepath,TrustCode,TrustName)
#                TrustBond(Excelfilepath,TrustCode)
#                except:
#                    pass
#                    print(TrustName,'错误!')
#                        TrustExtension(Excelfilepath)
    
    

    

#            
            
        
            

        
        
    
    
    
    
    
    #%%
    
    #     #单个文件夹校验导入
    # path=r'\\172.16.6.143\StudentsProducts\1210网商花呗受托报告\花呗管理人投后报告汇总\基础表\补充'
    # for file in os.listdir(path):
    #     if '基础表' in file:
    #         split=file.split(";")
    #         TrustCode=split[0]
    #         TrustName=split[1]
    #         Excelfilepath=os.path.join(path,file)
    #         Trust(Excelfilepath,TrustCode,TrustName)
    #         TrustBond(Excelfilepath)
    #         TrustExtension(Excelfilepath)
    #         PrincipalSchedule(Excelfilepath)
    #         print('\n')
    #
    # #%%
    #
    #批量校验上传
    path=r'\\172.16.6.143\StudentsProducts\1210网商花呗受托报告\花呗管理人投后报告汇总\基础表\1'
    for p_file in os.listdir(path):
        file_path=os.path.join(path,p_file)
        os.chdir(file_path)
        for P1_file in os.listdir(file_path):
            p1_filepath=os.path.join(file_path,P1_file)
            try:               
                for excelfile in os.listdir(p1_filepath):
#                    if '.txt' in excelfile:
#                        TrustCode=excelfile.split('.')[0]
                    
                    if '基础表.xlsx' in excelfile:
                        print(excelfile)
                        split=excelfile.split("_")
                        
                        FileTrustCode=split[0]
                        TrustName=split[1]
                        
                        Excelfilepath=os.path.join(p1_filepath,excelfile)
                        Trust(Excelfilepath,FileTrustCode,TrustName)
                        TrustBond(Excelfilepath)
#                        TrustExtension(Excelfilepath)
                    
            except:
                print(TrustId,TrustName,'出错!')
                print('\n')
                pass