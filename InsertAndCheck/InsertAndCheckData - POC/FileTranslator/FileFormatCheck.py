# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import re
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import pandas
import openpyxl.styles as sty

logtxtFilePath = None
errtxtFilePath = None
def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def writeErr(msg):
    if not os.path.exists(errtxtFilePath):
        f = open(errtxtFilePath, "w")
    # print(msg)
    with open(errtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def is_number(s):
    try:
        float(s.replace(',', ''))
        return True
    except ValueError:
        return False

def markCellError(sheet, ctag, errtype):
    color = 'E93936' #datatype
    if errtype == 1:#empty
        color = 'AA2927'
    #sheet[ctag].fill=sty.PatternFill(fill_type='solid', fgColor=color)

#return isValid, isNA, errorType(1:empty, 2:datatype)
def checkCellValue(sheet, ctag, cvalue, cdtype):
    if ((cdtype == 'mumber') and isinstance(cvalue, float)
       ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return 1, 0, 0

    # 去除单元格内容中的空格换行等
    cvalue = str(cvalue).replace(' ', '').replace(',', '').replace('.00', '').replace('\t', '').replace('\n', '')

    if cvalue == 'NA' or cvalue == '-' or cvalue == '':
        return 1, 1, 0

    # if cvalue == '':
    #     markCellError(sheet, ctag, 1)
    #     return 0, 0, 1

    dtcheck = 0
    if cdtype == 'int':
        dtcheck = 1 if cvalue.isdigit() else 0
    elif cdtype == 'number':
        dtcheck = 1 if is_number(cvalue) else 0
    elif cdtype == 'date':
        reg = r"(\d{4}[-/]\d{1,2}[^a-z]([-/]\d{1,2}[^a-z])?)|((\d{1,2}[^a-z][-/])?\d{1,2}[^a-z][-/]\d{4})|(\d{4}年\d{1,2}月(\d{1,2}日)?)|^(\d{5,6})$|^(\d{8})$"
        dtcheck = 1 if re.search(reg, cvalue) is not None else 0
    elif cdtype == 'rate':
        if '%' in cvalue or '%' in sheet[ctag].number_format:
            dtcheck = 1 if is_number(cvalue.rstrip('%')) else 0
        else:
            dtcheck = 0
    else:  # string
        dtcheck = 1

    if dtcheck == 0:
        markCellError(sheet, ctag, 2)
        return 0, 0, 2

    return 1, 0, 0

def getDate(value):
    model1 = r"(\d{4}年\d{1,2}月(\d{1,2}日)?)"
    model2 = r"(\d{4}[-/]\d{1,2}[^a-z]([-/]\d{1,2}[^a-z])?)|((\d{1,2}[^a-z][-/])?\d{1,2}[^a-z][-/]\d{4})"
    model3 = r"^(\d{5,6})$|^(\d{8})$"
    #value = '2016年3月31日'
    if re.search(model1, value) is not None:
        value = re.sub('\D$', '', value)
        value = re.sub(r'\D', r'-', value)
        value = pandas.to_datetime(value)
    elif re.search(model2, value) is not None:
        value = pandas.to_datetime(value)
    elif re.search(model3, value) is not None:
        value = pandas.Timedelta(str(value) + 'D')
        value = pandas.to_datetime('1899-12-30') + value
    print(value)
    return value.strftime('%Y-%m-%d')

def specificCellsExtract(cfgItem, sheet):
    checkmsg = []

    for cell in cfgItem:
        ctag = cell.tag
        cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        #crequired = cell.attrib['required'] if 'required' in cell.attrib else '0'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

        isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)

        if isValid != 1:
            etype = "不能为空" if errorType == 1 else "格式错误"
            checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, etype))

    return checkmsg

def specificColsExtract(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 1
    breakRowsLoop = 0
    while rStart <= rEnd:  # rows loop
        for cell in cNode: # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            #crequired = cell.attrib['required'] if 'required' in cell.attrib else '0'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break #break cells loop

            isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)
            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            if isValid != 1:
                etype = "不能为空" if errorType == 1 else "格式错误"
                checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, etype))

        if (isFirstRow == 1 and firstRowAllNA == 1) or breakRowsLoop == 1:
            rStart = rEnd + 1 #break rows loop
        else:
            isFirstRow = 0
            rStart += 1

    return checkmsg

def CompareDateExtract(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 0
    breakRowsLoop = 0
    previousValue = ''
    while rStart <= rEnd:  # rows loop
        for cell in cNode:  # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'

            cvalue = str(sheet[ctag].value) if sheet[ctag].value != None else ''
            if str(cvalue).replace(' ', '').replace('\t', '').replace('\n','') == 'NA':
                firstRowAllNA = 1
                break  # break cells loop
            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n',
                                                                                       '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break  # break cells loop
            if previousValue == '':
                previousValue = cvalue
            if getDate(previousValue) > getDate(cvalue):
                checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, "日期不合理，日期小余上期日期！"))
            else:
                previousValue = cvalue

        if (isFirstRow == 1 or breakRowsLoop == 1) and firstRowAllNA == 1:
            rStart = rEnd + 1  # break rows loop
        else:
            isFirstRow = 0
            rStart += 1

    return checkmsg

def checkFileFormat(wb, cfgItems):
    sheet = wb['Sheet1']
    ckResult = []
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']
        itemtype = cfgItem.attrib['type'] if 'type' in cfgItem.attrib else ''

        itemCkResult = []
        if itemtype == 'SpecificCells':
            itemCkResult = specificCellsExtract(cfgItem, sheet)
        elif itemtype == 'CompareDate':
            itemCkResult = CompareDateExtract(cfgItem, sheet)
        else:
            itemCkResult = specificColsExtract(cfgItem, sheet)

        if len(itemCkResult) > 0:
            ckResult.append("【{0}】".format(itemdesc))
            ckResult.extend(itemCkResult)

    ckResultLen = len(ckResult)
    if ckResultLen > 0:
        logSheet = wb.create_sheet("格式检查")
        i = 1
        while i <= ckResultLen:
            logSheet["A{0}".format(i)] = ckResult[i - 1]
            i += 1

    return ckResultLen

def main(configFilePath, dateId):
    global errtxtFilePath
    global logtxtFilePath
    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    scriptFolder = cfgRoot.attrib['sourcefolder']
    destFolder = cfgRoot.attrib['destfolder']
    scriptFolderPath = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolderPath, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolderPath, 'Logs',
                                  '{0}.txt'.format(dateId))
    errtxtFilePath = os.path.join(destFolder,
                                  'Error_第一步校验错误_{0}.txt'.format(dateId))

    for dirPath, dirNames, fileNames in os.walk(scriptFolder):
        for fileName in fileNames:
            if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告'):
                print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
                continue

            sourceFilePath = os.path.join(dirPath, fileName)

            filePath = "\n{0}".format(sourceFilePath)
            writeLog(filePath)

            fileNameAry = fileName.split(';')
            if len(fileNameAry) != 6:
                msg = "【文件名错误】文件名称命名不规范"
                writeLog(msg)
                writeErr(filePath)
                writeErr(msg)
                continue

            # initialize key pamerters
            #trustID = 0
            paymentPeriodID = 0

            #trustID = fileNameAry[1]
            paymentPeriodID = fileNameAry[3]
            
            if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
                msg = "【错误】文件名中的TrustCode或报告期数设置有误"
                writeLog(msg)
                writeErr(filePath)
                writeErr(msg)
                continue

            excelwb = load_workbook(sourceFilePath)
            if '格式检查' in excelwb.sheetnames:
                excelwb.remove(excelwb['格式检查'])
                excelwb.save(sourceFilePath)

            hasError = checkFileFormat(excelwb, cfgRoot)
            if hasError > 0:
                excelwb.save(sourceFilePath)
                writeLog('【有格式错误】详情见文档[格式检查]sheet')
                writeErr(filePath)
                writeErr('【有格式错误】详情见文档[格式检查]sheet')
