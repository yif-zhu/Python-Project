# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
from openpyxl import Workbook
import pyodbc


logtxtFilePath = None
errtxtFilePath = None
dbConnectionStr = None

##### Helper Methods #####
def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def execSQLCmdFetchAll(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        rows = cursor.execute(sql).fetchall()
        cnxn.commit()
        return rows
    except Exception as ex:
        writeLog(str(ex))
        raise ex
    finally:
        cnxn.close()

def runDBDataValidation(type, poolType, Importime, assetType, TrustId):

    sql = "exec DV.InternalLogicCheckAdmin {0},{1},N'{2}',N'{3}','{4}'".format(type, poolType, Importime, assetType, TrustId)
    msg = execSQLCmdFetchAll(sql)
    return msg

def main(configFilePath, dateId):
    global logtxtFilePath
    global errtxtFilePath
    global dbConnectionStr
    global trustID
    global paymentPeriodID

    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    type = cfgRoot.attrib['type']
    poolType = cfgRoot.attrib['poolType']
    ImportTime = cfgRoot.attrib['ImportTime']
    assetType = cfgRoot.attrib['AssetType']
    TrustId = cfgRoot.attrib['TrustId']
    destFolder =cfgRoot.attrib['destfolder']
    dbConnectionStr = cfgRoot.attrib['dbconnstr']

    scriptFolderPath = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolderPath, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolderPath, 'Logs', '{0}.txt'.format(dateId))
    errtxtFilePath = os.path.join(destFolder, 'Error_第五步逻辑校验结果_{0}.xlsx'.format(dateId))

    dbCheckResult = runDBDataValidation(type, poolType, ImportTime, assetType, TrustId)

    if len(dbCheckResult) > 0:
        wb = Workbook()  # 新建工作簿
        ws1 = wb.active
        wb.save(errtxtFilePath)
        writeLog("第五步入库逻辑校验完成，错误详情请查看:{0}".format(errtxtFilePath))
        excelwb = load_workbook(errtxtFilePath)
        logSheet = excelwb['Sheet']
        logSheet["A{0}".format(1)] = dbCheckResult[0].cursor_description[0][0]
        logSheet["B{0}".format(1)] = dbCheckResult[0].cursor_description[1][0]
        logSheet["C{0}".format(1)] = dbCheckResult[0].cursor_description[2][0]
        logSheet["D{0}".format(1)] = dbCheckResult[0].cursor_description[3][0]
        logSheet["E{0}".format(1)] = dbCheckResult[0].cursor_description[4][0]
        logSheet["F{0}".format(1)] = dbCheckResult[0].cursor_description[5][0]
        logSheet["G{0}".format(1)] = dbCheckResult[0].cursor_description[6][0]
        i = 2
        for r in dbCheckResult:
            logSheet["A{0}".format(i)] = r[0]
            logSheet["B{0}".format(i)] = r[1]
            logSheet["C{0}".format(i)] = r[2]
            logSheet["D{0}".format(i)] = r[3]
            logSheet["E{0}".format(i)] = r[4]
            logSheet["F{0}".format(i)] = r[5]
            logSheet["G{0}".format(i)] = r[6]
            i += 1
        excelwb.save(errtxtFilePath)
    else:
        writeLog("第五步入库逻辑校验完成，无错误！！！")

if __name__ == "__main__":
    filePath = "C:/PyCharm/InsertAndCheckData/FileTranslator/ThirdLogicalCheck_Trustee.xml"
    dateId = 2313121
    main(filePath, dateId)