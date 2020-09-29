# -*- coding: utf-8 -*-
import pandas as pd
import os
import re
import numpy as np
import datetime
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import traceback

logtxtFilePath = None

#mapping文件添加map节点
def addNodeForMapXml(MapPath, sourcecell, destCell, comment, sourcecolor, destcolor):
    tree = XETree.parse(MapPath)
    root = tree.getroot()
    node = root.find("Mapping")
    MapNode = XETree.Element('map')  # 创建节点,单个文件的mapping文件
    MapNode.set("destcolor", destcolor)
    MapNode.set("sourcecolor", sourcecolor)
    MapNode.set("comment", comment)
    MapNode.set("destCell", destCell)
    MapNode.set("sourcecell", sourcecell)
    node.append(MapNode)
    indent(node)
    tree.write(MapPath, encoding='utf-8', xml_declaration=True)
    return

def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

#插入格式检查表，填入问题
def writeErrToSheet(sourceFilePath, ErrorMsg, mapPath):
    excelwb = load_workbook(sourceFilePath)
    if len(ErrorMsg) > 0:
        logSheet = excelwb.create_sheet("格式检查")
        i = 1
        while i <= len(ErrorMsg):
            if len(ErrorMsg[i - 1].split(':')) == 3:
                sourceCell = ErrorMsg[i - 1].split(':')[0]
                destCell = "A{0}".format(i)
                errorType = ErrorMsg[i - 1].split(':')[1]
                comment = ErrorMsg[i - 1].split(':')[2]
                sourceColor = "255,255,255"
                destColor = ''
                if errorType == 1:
                    destColor = "170,41,39"
                elif errorType == 2:
                    destColor = "233,57,54"
                else:
                    destColor = "255,230,153"
                addNodeForMapXml(mapPath, sourceCell, destCell, comment, sourceColor, destColor)
                logSheet["A{0}".format(i)] = ErrorMsg[i - 1]
                i += 1
            else:
                logSheet["A{0}".format(i)] = ErrorMsg[i - 1]
                i += 1
    else:
        logSheet = excelwb.create_sheet("格式检查")
        logSheet["A1"] = '检验完成，无错误！！！'

    excelwb.save(sourceFilePath)

#根据标题确定列
def ErrLocation(colName):  # 定位错误的具体列
    if colName == 'PaymentPeriodID':
        return 'A'
    elif colName == 'Amount':
        return 'G'
    elif colName == 'BucketSequenceNo':
        return 'E'
    elif colName == 'AmountPercentage':
        return 'H'
    elif colName == 'CountPercentage':
        return 'J'
    elif colName == 'Count':
        return 'I'

def is_Number(value):
    num = '^[-]?\d*[.]?\d*$'
    flag = re.search(num, value)
    if str(flag) == 'None':
        return False
    else:
        return True

#校验函数，校验池分布excel是否符合规则
def CheckData(sourceFilePath, mapPath):
    ErrorMsg = []  # 错误信息记录
    file = pd.read_excel(sourceFilePath)
    splitname = sourceFilePath.split('_')[2]

    # PaymentPeriodID列检测
    try:
        if 'PaymentPeriodID' not in file.columns:
            ErrorMsg.append("A1:0:PaymentPeriodID字段缺失，请勿更改模板列标题")
        if 'Amount' not in file.columns:
            ErrorMsg.append("G1:0:Amount字段缺失，请勿更改模板列标题")
        if 'BucketSequenceNo' not in file.columns:
            ErrorMsg.append("E1:0:BucketSequenceNo字段缺失，请勿更改模板列标题")
        if 'AmountPercentage' not in file.columns:
            ErrorMsg.append("H1:0:AmountPercentage字段缺失，请勿更改模板列标题")
        if 'CountPercentage' not in file.columns:
            ErrorMsg.append("J1:0:CountPercentage字段缺失，请勿更改模板列标题")
        if 'Count' not in file.columns:
            ErrorMsg.append("I1:0:Count字段缺失，请勿更改模板列标题")
        if 'DistributionType' not in file.columns:
            ErrorMsg.append("B1:0:DistributionType字段缺失，请勿更改模板列标题")

        if len(ErrorMsg) > 0:  # 如果有字段缺失，先补全字段后继续校验
            writeErrToSheet(sourceFilePath, ErrorMsg, mapPath)
            return len(ErrorMsg)

        for i in range(file['PaymentPeriodID'].index.values.size):
            value = int(file['PaymentPeriodID'][i]) if file['PaymentPeriodID'][i] == file['PaymentPeriodID'][i] else -1   #用于防止数据是NaN的情况
            if value != 0:
                val = int(file['PaymentPeriodID'].index.values[i])
                ErrorMsg.append("A{0}:0:PaymentPeriodID期数填写错误，应为0期".format(val + 2))

        # 检测BucketSequenceNo列是否有重复

        def checkUnique(file):
            if file['BucketSequenceNo'].is_unique == False:
                cells = ''
                for i in range(file['BucketSequenceNo'].index.values.size):
                    val = int(file['BucketSequenceNo'].index.values[i])
                    cells = cells + ErrLocation('BucketSequenceNo') + str(val + 2) + ';'
                ErrorMsg.append("{0}:0:BucketSequenceNo排序数不唯一，请修改".format(cells.rstrip(';')))

        pd.DataFrame(file.groupby('DistributionType').apply(checkUnique))

        # 检测AmountPercentage/CountPercentage列数值设置是否大于或小于100
        def Verification(file):
            ErrorNum = 0
            for i in range(file['AmountPercentage'].index.values.size):
                val = int(file['AmountPercentage'].index.values[i])
                flag = is_Number(str(file['AmountPercentage'][val]))
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('AmountPercentage') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-AmountPercentage值：{1}格式错误，请修改".format(cell.rstrip(';'),
                                                                                         file['AmountPercentage'][val]))

                val = int(file['CountPercentage'].index.values[i])
                flag = is_Number(str(file['CountPercentage'][val]))
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('CountPercentage') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-CountPercentage值：{1}格式错误，请修改".format(cell.rstrip(';'),
                                                                                        file['CountPercentage'][val]))
                val = int(file['Count'].index.values[i])
                flag = is_Number(str(file['Count'][val]))
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('Count') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-Count值：{1}格式错误，请修改".format(cell.rstrip(';'), file['Count'][val]))

                val = int(file['Amount'].index.values[i])
                flag = is_Number(str(file['Amount'][val]))
                if flag is False:
                    ErrorNum += 1
                    cell = ErrLocation('Amount') + str(val + 2)
                    ErrorMsg.append("{0}:0:列数值设置有误-Amount值：{1}格式错误，请修改".format(cell.rstrip(';'), file['Amount'][val]))

            if ErrorNum == 0:
                file['Count'] = file['Count'].astype(float)
                file['Amount'] = file['Amount'].astype(float)

                AmountPercentage_c = file['AmountPercentage'].sum()
                CountPercentage_c = file['CountPercentage'].sum()
                Count = file['Count'].sum()
                Amount = file['Amount'].sum()

                if (AmountPercentage_c > 100.5 or AmountPercentage_c < 99.5) and CountPercentage_c != 0 and AmountPercentage_c != 0:
                    cells = ''
                    for i in range(file['AmountPercentage'].index.values.size):
                        val = int(file['AmountPercentage'].index.values[i])
                        cells = cells + ErrLocation('AmountPercentage') + str(val + 2) + ';'
                    ErrorMsg.append(
                        "{0}:0:列数值设置有误-AmountPercentage分布总计值：{1}不等于100(忽略精度)".format(cells.rstrip(';'), AmountPercentage_c))

                if (CountPercentage_c >= 100.5 or CountPercentage_c <= 99.5) and CountPercentage_c != 0 and AmountPercentage_c != 0:
                    cells = ''
                    for i in range(file['CountPercentage'].index.values.size):
                        val = int(file['CountPercentage'].index.values[i])
                        cells = cells + ErrLocation('CountPercentage') + str(val + 2) + ';'
                    ErrorMsg.append(
                        "{0}:0:列数值设置有误-CountPercentage分布总计值：{1}不等于100(忽略精度)".format(cells.rstrip(';'), CountPercentage_c))

                # 检测Count\Amount是否录反
                if Count > Amount:
                    cells = ''
                    for i in range(file['Count'].index.values.size):
                        val = int(file['Count'].index.values[i])
                        cells = cells + ErrLocation('Count') + str(val + 2) + ';'
                    for i in range(file['Amount'].index.values.size):
                        val = int(file['Amount'].index.values[i])
                        cells = cells + ErrLocation('Amount') + str(val + 2) + ';'
                    ErrorMsg.append(
                        "{0}:0:Count合计值：{1}不能大于Amount合计值：{2}".format(cells.rstrip(';'), Count, Amount))

        file = file.replace(np.nan, 0)
        pd.DataFrame(file.groupby('DistributionType').apply(Verification))
        writeErrToSheet(sourceFilePath, ErrorMsg, mapPath)
        return len(ErrorMsg)
    except Exception as e:
        ErrorMsg.append("A0:0:文件错误，请确认使用模板，使用规定池分布模板，或表格中填值不能为空，请填入NA！！！")
        writeErrToSheet(sourceFilePath, ErrorMsg, mapPath)
        writeLog(traceback.format_exc())
        return -1

#给xml增加换行符
def indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

#新建mamppingxml， multiply 用于判断是单个文件还是文件夹
def createXml(xmlPath, inputfile, outputfile, multiply):
    if multiply.lower() == 'false':
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "false")
            root.set("inputFile", inputfile)
            root.set("outputFile", outputfile)
            root.set("inputSheetIndex", "0")
            root.set("outputSheetIndex", "1")
            root.set("inputSheetName", r"贷款分布")
            root.set("outputSheetName", r"格式检查")
            tree = XETree.ElementTree(root)  # 创建文档
            Mapping1 = XETree.Element('Mapping') #创建子节点
            Mapping1.set("description", r'源文件，目标文件对应情况')
            root.append(Mapping1)
            indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
    else:
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "true")
            tree = XETree.ElementTree(root)  # 创建文档
            # indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)

def main(configFilePath, dateId):
    global logtxtFilePath
    scriptFolder = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolder, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolder, 'Logs',
                                  '{0}.txt'.format(dateId)) #获取log文件路径，并新建

    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    sourceFileForder = cfgRoot.attrib['sourcefolder']

    dir_path = scriptFolder + '\\MappingXml\\'  # mapping文件存放路径
    mappingPath = dir_path + dateId + '.xml'
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)
    #sourceFileForder = r"C:/PyCharm\pdf-docx\source\Example\池分布"
    for dirPath, dirNames, fileNames in os.walk(sourceFileForder):
        config = 1
        createXml(mappingPath, '', '', 'true')
        for fileName in fileNames:
            if not fileName.endswith('池分布.xlsx'):
                msg = "【跳过】文件名称不符合，已跳过文件{0}".format(fileName)
                writeLog(msg)
                continue

            sourceFilePath = os.path.join(dirPath, fileName)

            msg = "\n{0}".format(sourceFilePath)
            writeLog(msg)
            fileNameAry = fileName.split('_')
            if len(fileNameAry) != 4:
                msg = "【文件名错误】文件名称命名不规范"
                writeLog(msg)
                continue

            mulPath = dir_path + dateId + '_' + str(config) + '.xml'
            createXml(mulPath, sourceFilePath, sourceFilePath, 'false')
            config += 1
            tree = XETree.parse(mappingPath)
            root = tree.getroot()
            MapPath = XETree.Element('filename')  # 创建节点,单个文件的mapping文件
            MapPath.set("path", mulPath)
            root.append(MapPath)
            indent(root)
            tree.write(mappingPath, encoding='utf-8', xml_declaration=True) #创建mappingxml文件

            excelwb = load_workbook(sourceFilePath)
            if '格式检查' in excelwb.sheetnames:
                excelwb.remove(excelwb['格式检查'])
                excelwb.save(sourceFilePath)

            hasError = CheckData(sourceFilePath, mulPath)
            if hasError > 0 or hasError < 0:
                writeLog(sourceFilePath)
                writeLog('【有格式错误】详情见文档[格式检查]sheet')
            else:
                writeLog(sourceFilePath)
                writeLog('通过校验，无错误！！！')

