import os
import os.path
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
sourceExcel = None

#列字母转数字 A转1
def letterToint(s):
    letterdict = {}
    for i in range(26):
        letterdict[chr(ord('A') + i)] = i + 1
    output = 0
    for i in range(len(s)):
        output = output * 26 + letterdict[s[i]]
    return output

#数字转列字母 1转A
def intToletter(i):
    if type(i) is not int:
        return i
    str = ''
    while (not (i // 26 == 0 and i % 26 == 0)):

        temp = 25

        if (i % 26 == 0):
            str += chr(temp + 65)
        else:
            str += chr(i % 26 - 1 + 65)

        i //= 26
        # print(str)
    # 倒序输出拼写的字符串
    return str[::-1]
#列表匹配
def extractForTable(cfgItem):
    sheet = sourceExcel["Sheet1"]
    tempcell = findStr(sheet, cfgItem[0].attrib["anchor"], 1, 1)

    #如果差找不到，首行设置为NA
    if tempcell is None:
        return 0
    else:
        return 1

#查找匹配的cell单元格
def findStr(sheet, key, startRow, startCol):
    for col in range(startCol, sheet.max_column):
        for row in range(startRow, sheet.max_row):
            cell = sheet[intToletter(col) + str(row)]
            if cell is None:
                return None
            if cell.value is not None and cell.value != "":
                if key in str(cell.value).replace(' ', ''):
                    return cell

#根据key查找cell单元格
def extractForKeyValue(cfgItem):
    sheetS = sourceExcel['Sheet1']

    sNode = cfgItem.find('source')
    anchors = sNode.attrib['anchor'].strip().split(';')
    length = len(anchors)
    beginCol = 1
    beginRow = 1
    isFind = 0;
    tempcell = None
    for i in range(length):
        tempcell = findStr(sheetS, anchors[i], beginRow, beginCol)
        if tempcell is not None:
            beginRow = tempcell.row
            beginCol = letterToint(tempcell.column) + 1
            if i == length - 1:
                isFind = 1
                break
    if isFind == 1:
        return 1
    if isFind == 0:
        return 0

#根据类型选择列表匹配，还是值匹配
def selectType(cfgItems):
    count = 0
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        istable = cfgItem.attrib['istable'].lower()
        isfind = 0
        if istable == 'true':
            isfind = extractForTable(cfgItem)
        else:
            isfind =extractForKeyValue(cfgItem)
        count += isfind
    return count

#给定源文件，目标文件进行运行
def chooseExcel(inputFile, cfgRoot):
    global sourceExcel
    sourceExcel = load_workbook(inputFile)
    count = selectType(cfgRoot)
    return count

#比较各配置文件中匹配的个数，返回匹配最多的一个xml文件
def compareMatchCount(matchCount):
    xmlIndex = 0
    xmlCount = 0
    for index in range(len(matchCount)):
        matchIndex = int(matchCount[index].split(':')[0])
        count = int(matchCount[index].split(':')[1])
        if count > xmlCount:
            xmlIndex = matchIndex
            xmlCount = count
    return xmlIndex

def main(input):
    dir_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    inputFile = ''
    matchCount = []
    if os.path.isfile(input):
        inputFile = input
    elif os.path.isdir(input):
        for parent, dirnames, filenames in os.walk(input, followlinks=True):
            for filename in filenames:
                suffix = filename.split('.')[-1]
                if suffix == 'xlsx':
                    inputFile = os.path.join(parent, filename)
                    break

    configFilePath = os.path.join(dir_path, 'Config\\config.xml')   #总xml记录文件路径
    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    xmlNode = cfgRoot.find('xmlFIle')
    for index in range(len(xmlNode)):
        xmlPath = xmlNode[index].text
        xmlTree = XETree.parse(xmlPath)
        xmlRoot = xmlTree.getroot()
        count = chooseExcel(inputFile, xmlRoot)
        matchStr = "{0}:{1}".format(index, count)
        matchCount.append(matchStr)
    matchIndex = compareMatchCount(matchCount)
    Name = xmlNode[matchIndex].attrib['Name']
    return Name
if __name__ == "__main__":
    path = 'C:\\PyCharm\\pdf-docx\\source\\task1'
    name = main(path)
    print(name)