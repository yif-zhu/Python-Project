# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import re
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import openpyxl.styles as sty

explainStr = dict()
explainStr["0"] = "首先请注意模板中绿框填写要求，填0还是填NA；其次对百分比，日期格式数据，请按要求填写"
explainStr["1"] = "首先请注意模板中绿框填写要求，填0还是填NA；第2期的期初余额减去第2期的本金还款等于第2期的期末余额，第1期的期末余额等于第2期的期初余额，以此类推，所以本金并不会越还越多，请确定是否取数取错地方了"
explainStr["2"] = "首先请注意模板中绿框填写要求，填0还是填NA；第1期的期初余额减去第1期的本金还款等于第0期的期末余额，第1期的期末余额等于第2期的期初余额，以此类推，所以本金并不会越还越多，请确定是否取数取错地方了"
explainStr["3"] = "首先请注意款模板中绿框填写要求，填0还是填NA；第2期的期初余额减去第2期的本金还等于第2期的期末余额，第1期的期末余额等于第2期的期初余额，以此类推。如果发现所有期都对不上请告知组长，寻找解决方法"
explainStr["4"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为在本金未还清之前，都得付利息，可联系实际思考一下。首先确定是否录错，如确实未录错（备注表请下载）那就备注：未录错，未修改，XXX层利息还款确实为XXX，发给所在组的组长"
explainStr["5"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为在本金未还清之前，都得付利息，至少有一层是有利息的。首先确定是否录错，如确实未录错（备注表请下载）那就备注：未录错，未修改，当期所有层的利息还款确实为XXX，发给所在组的组长"
explainStr["6"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为受托报告披露的即为整个资产池情况，资产池中每一笔资产都将披露，所以笔数占比应该加总为100%。如确实未录错（备注表请下载）那就备注：未录错，未修改，笔数占比之和确实不等于100%，为XXX，发给所在组的组长"
explainStr["7"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为受托报告披露的即为整个资产池情况，资产池中每一笔资产都将披露，所以会认为逾期分布金额合计值等于当期资产池总余额。如确实未录错（备注表请下载）那就备注：未录错，未修改，差异值为XXX，发给所在组的组长"
explainStr["8"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为受托报告披露的即为整个资产池情况，资产池中每一笔资产都将披露，所以金额占比应该加总为100%。如确实未录错（备注表请下载）那就备注：未录错，未修改，笔数占比之和确实不等于100%，为XXX，发给所在组的组长"
explainStr["9"] = "首先请注意模板中绿框填写要求，填0还是填NA；现金流归集表是指立足于当期对于未来期回款的预测（比如当期是201909，那么现金流归集表指的是201910以及往后月份的一个回款预测），那么还本金额不可能越还越多，所以期初本金余额应小于应收本金余额（除最后一期等于），如遇特殊情况，请告知组长，寻找解决方法"
explainStr["10"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为收入账分项相加等于总计，但有很多不等的情况，或者同一个数据在多个表披露，但是不等（这种情况请及时询问），遇见这样的，首先保证和PDF一致，未录错，一定不要漏填（一定注意上期转存、转存下期等字段数据一定要录入，因为加上上期转存的大合计并不是当期的实际收入，需要录入后，系统自动减去）（特殊资产类型需要加上手续费等作为收入账数据，请联系实际思考信用卡），其次根据实际情况备注（备注表请下载）发给所在组的组长"
explainStr["11"] ="首先请注意模板中绿框填写要求，填0还是填NA；一般认为本金账分项相加等于总计，但有很多不等的情况，或者同一个数据在多个表披露，但是不等（这种情况请及时询问），遇见这样的，首先保证和PDF一致，未录错，一定不要漏填（一定注意上期转存、转存下期等字段数据，，因为加上上期转存的大合计并不是当期的实际收入，需要录入后，系统自动减去）（特殊资产类型需要加上手续费等作为收入账数据，请联系实际思考信用卡），其次根据实际情况备注（备注表请下载）发给所在组的组长"
explainStr["12"] = "首先请注意模板中绿框填写要求，填0还是填NA；由于在所有资产类型中，房贷的期限是最长的，为30年，所以任何资产类型的期限特征都不可能大于12*30个月，如遇特殊情况，请告知组长，寻找解决方法"
explainStr["13"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为受托报告会披露资产池的整体表现情况，所以会总数应该等于当期资产池余额，但是实际上大多不相等，这是因为存在未披露的数据，这种情况下，首先需要保证未录错，其次如果能计算出差值和披露的哪个数据能对上更好，然后根据实际情况备注（备注表请下载）发给所有组的组长"
explainStr["14"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为累计违约金额顶多为初始资产池余额，比如最开始的资产池未100元，那违约也只能违约这100元，不可能违约101元，但是如果是循环购买产品会不断增加资产池金额，所以累计违约金额会超出最开始的100元。如果遇见这样的，首先保证和PDF一致，未录错，根据实际情况备注（备注表请下载）发给所在组的组长"




logtxtFilePath = None
class cObj(object):
    def __init__(self, cell, desc, val):
        self.cell = cell
        self.desc = desc
        self.val = val
#文件无错误，更新文件index
def updateIndexForMapXml(wb, MapPath):
    logSheet = wb.create_sheet("log")
    logSheet["A1"] = r"校验完成，该文件通过校验，无错误！"
#mapping文件添加map节点
def addNodeForMapXml(MapPath, sourcecell, destCell, comment, sourcecolor, destcolor):
    tree = XETree.parse(MapPath)
    root = tree.getroot()
    node = root.find("Mapping")
    MapNode = XETree.Element('map')  # 创建节点,单个文件的mapping文件
    MapNode.set("destcolor", destcolor)
    MapNode.set("sourcecolor", sourcecolor)
    MapNode.set("comment", comment)
    MapNode.set("destCell", destCell)
    MapNode.set("sourcecell", sourcecell)
    node.append(MapNode)
    indent(node)
    tree.write(MapPath, encoding='utf-8', xml_declaration=True)
    return

def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def is_number(s):
    try:
        float(s.replace(',', ''))
        return True
    except ValueError:
        return False

def getNumberVal(val):
    if isinstance(val, int) or isinstance(val, float):
        return val
    val = val.replace(' ', '')
    if val =='NA':
        return None
    elif val == '-' or val == '':
        return 0.0
    else:
        return float(val.replace('\t', '').replace('\n', '').replace(',', '').replace('%', '').replace(' ', ''))
    #return 0.0 if val == 'NA' or val =='-' or val == '' else float(val.replace('\t', '').replace('\n', '').replace(',', '').replace('%', '').replace(' ', ''))

def markCellError(sheet, ctag, errtype):
    color = 'E93936' #datatype
    if errtype == 1:#empty
        color = 'AA2927'
    #sheet[ctag].fill=sty.PatternFill(fill_type='solid', fgColor=color)

def checkCellValue(sheet, ctag, cvalue, cdtype):
    if ((cdtype == 'mumber' or cdtype == 'rate') and isinstance(cvalue, float)
        ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return 1, 0, 0

    cvalue = str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '')

    if cvalue == 'NA' or cvalue == '-':
        return 1, 1, 0

    if cvalue == '':
        markCellError(sheet, ctag, 1)
        return 0, 0, 1

    dtcheck = 0
    if cdtype == 'int':
        dtcheck = 1 if cvalue.isdigit() else 0
    elif cdtype == 'number':
        dtcheck = 1 if is_number(cvalue) else 0
    elif cdtype == 'date':
        reg = r"(\d{4}[-/]\d{1,2}([-/]\d{1,2})?)|((\d{1,2}[-/])?\d{1,2}[-/]\d{4})|(\d{4}年\d{1,2}月(\d{1,2}日)?)|^(\d{5,6})$|^(\d{8})$"
        dtcheck = 1 if re.search(reg, cvalue) is not None else 0
    elif cdtype == 'rate':
        dtcheck = 1 if is_number(cvalue.rstrip('%')) else 0
    else:  # string
        dtcheck = 1

    if dtcheck == 0:
        markCellError(sheet, ctag, 2)
        return 0, 0, 2

    return 1, 0, 0

def specificCellsCheck(cfgItem, sheet):
    checkmsg = []
    r = []
    for cell in cfgItem:
        ctag = cell.tag
        cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

        isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)

        if isValid != 1:
            etype = "不能为空" if errorType == 1 else "格式错误"
            checkmsg.append("{0}:{3}:[{1}] - {2}:0".format(ctag, cdesc, etype, errorType))

        r.append(cObj(ctag, cdesc, cvalue))

    return checkmsg, r

def specificColsCheck(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 1
    breakRowsLoop = 0
    rs = []
    while rStart <= rEnd:  # rows loop
        r = []
        for cell in cNode: # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break #break cells loop

            isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)
            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            if isValid != 1:
                etype = "不能为空" if errorType == 1 else "格式错误"
                checkmsg.append("{0}:{3}:[{1}] - {2}:0".format(ctag, cdesc, etype, errorType))

            r.append(cObj(ctag, cdesc, cvalue))

        if (isFirstRow == 1 and firstRowAllNA == 1) or breakRowsLoop == 1:
            rStart = rEnd + 1 #break rows loop
        else:
            isFirstRow = 0
            rStart += 1
            rs.append(r)

    return checkmsg, rs

def writeSheetLog(wb, res, mapPath):
    logSheet = wb.create_sheet("格式检查")
    i = 1
    logSheet["A1"] = "错误信息："
    logSheet["B1"] = "解释信息："
    while i <= len(res):
        if len(res[i-1].split(':')) == 4:
            sourceCell = res[i-1].split(':')[0]
            destCell = "A{0}".format(i+1)
            explainCell = "B{0}".format(i+1)
            errorType = res[i-1].split(':')[1]
            comment = res[i-1].split(':')[2]
            explainType = res[i-1].split(':')[3]
            sourceColor = "255,230,153"
            destColor = ''
            if errorType == 1:
                destColor = "170,41,39"
            elif errorType == 2:
                destColor = "233,57,54"
            else:
                destColor = "255,230,153"
            addNodeForMapXml(mapPath, sourceCell, destCell, comment, sourceColor, destColor)
            logSheet[destCell] = str(sourceCell + '：' + comment)
            logSheet[explainCell] = explainStr[explainType]
            i += 1
        else:
            logSheet["A{0}".format(i+1)] = res[i - 1]
            i += 1

def getValidationResult(ds):
    ckResult = []

    bondsAmt = 0.0
    cells = ''
    for d in ds[0]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestPaied = getNumberVal(d[3].val)
        pendAmt = getNumberVal(d[4].val)
        if pInterestPaied is not None:
            bondsAmt += pInterestPaied
            cells += str(d[3].cell) + ';'

        if pendAmt is not None and pstartAmt is not None and pendAmt > pstartAmt:
            ckResult.append("{2};{3}:{4}:【收益分配记录】债券期末余额 [{0}] 大于期初余额{1},数据对应关系错误:1".format(pendAmt, pstartAmt, d[4].cell, d[1].cell, 0))
        if pPrincipalAmt is not None and pstartAmt is not None and pPrincipalAmt > pstartAmt:
            ckResult.append("{2};{3}:{4}:【收益分配记录】债券本金金额 [{0}] 大于期初余额{1}，数据对应关系错误:2".format(pPrincipalAmt, pstartAmt, d[2].cell, d[1].cell, 0))

        if pstartAmt is not None and pPrincipalAmt is not None and pendAmt is not None and round(pstartAmt - pPrincipalAmt, 6) != pendAmt:
            ckResult.append("{3};{4};{5}:{6}:【收益分配记录】期初余额 [{0}] 减去本金金额 [{1}] 不等于期末余额 [{2}] 错误:3".format(pstartAmt,
                                                                                                   pPrincipalAmt, pendAmt, d[1].cell, d[2].cell, d[4].cell, 0))
        if pInterestPaied is not None and pPrincipalAmt is not None and pPrincipalAmt > 0 and pInterestPaied == 0:
            ckResult.append(
                "{0};{1}:{2}:【收益分配记录】本金还款金额存在，利息还款金额缺失，错误:4".format(d[2].cell, d[3].cell, 0))
        # if pInterestPaied is not None and pendAmt is not None and pInterestPaied >= pendAmt:
        #     ckResult.append(
        #         "{0};{1}:{2}:【收益分配记录】利息还款金额大于等于剩余金额，错误".format(d[3].cell, d[4].cell, 0))

    bondsAmt = round(bondsAmt, 4)
    if bondsAmt == 0 and cells != '':
        ckResult.append(
            "{0}:{1}:【收益分配记录】当期数据中，必须有一层利息还款金额不为0，错误:5".format(cells[:-1], 0))
    t2 = 0.00
    t3 = 0.00
    t4 = 0.00
    cells2 = ''
    cells3 = ''
    cells4 = ''
    for d in ds[1]:
        loanCount = getNumberVal(d[2].val)  #笔数占比
        Amount = getNumberVal(d[3].val)  #金额
        amtCount = getNumberVal(d[4].val)  #金额占比
        if loanCount is not None:
            t2 += loanCount
            cells2 += str(d[2].cell) + ';'
        if Amount is not None:
            t3 += Amount
            cells3 += str(d[3].cell) + ';'
        if amtCount is not None:
            t4 += amtCount
            cells4 += str(d[4].cell) + ';'
    t2 = round(t2, 6)
    t4 = round(t4, 6)
    if t2 != 0 and (abs(t2 - 100) > 0.02 and abs(t2 - 1) > 0.0002):
        ckResult.append("{0}:{1}:【资产池整体表现情况】笔数占比之和不等于100% :6".format(cells2[:-1], 0))
    assetAmt = getNumberVal(ds[5][2].val)
    if assetAmt is not None and assetAmt >= 0 and abs(assetAmt - t3) > 10000:
        ckResult.append("{2}{3}:{4}:【逾期分布金额合计值 】 [{0}] 同当期资产池总余额 [{1}] 相差在10000之上:7".format(t3,
                                                                                                     assetAmt, cells3, ds[5][2].cell, 0))
    if t4 != 0 and (abs(t4 - 100) > 0.02 and abs(t4 - 1) > 0.0002):
        ckResult.append("{0}:{1}:【资产池整体表现情况】金额占比之和不等于100%:8".format(cells4[:-1], 0))
    lenCash = len(ds[2])
    nowCount = 1
    for d in ds[2]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestAmt = getNumberVal(d[3].val)

        if nowCount == lenCash:
            if pstartAmt is not None and pPrincipalAmt is not None and pstartAmt != pPrincipalAmt:
                ckResult.append("{2};{3}:{4}:【现金流归集表】当期期初本金金额 [{0}] 不等于应收本金金额 [{1}] 数据对应关系错误:9".format(pstartAmt,
                                                                                                pPrincipalAmt, d[1].cell, d[2].cell, 0))
        else:
            if pstartAmt is not None and pPrincipalAmt is not None and pstartAmt <= pPrincipalAmt:
                ckResult.append("{2};{3}:{4}:【现金流归集表】当期期初本金金额 [{0}] 小于等于应收本金金额 [{1}] 数据对应关系错误:9".format(pstartAmt,
                                                                                                pPrincipalAmt, d[1].cell, d[2].cell, 0))

        nowCount+=1
    assetPoolCF = ds[3]
    totalInterest = 0.00
    cellsInterest = ''
    totalPrincipal = 0.00
    cellsPrincipal = ''
    if len(assetPoolCF) > 0:
        for i in range(8):
            val = getNumberVal(assetPoolCF[i].val)
            cellsInterest += str(assetPoolCF[i].cell) + ';'
            if val is not None:
                totalInterest += val
        for i in range(8, 16):
            val = getNumberVal(assetPoolCF[i].val)
            cellsPrincipal += str(assetPoolCF[i].cell) + ';'
            if val is not None:
                totalPrincipal += val

    assetPoolST = ds[4]
    totalInterest = round(totalInterest, 6)
    totalPrincipal = round(totalPrincipal, 6)
    if len(assetPoolST) > 0:
        val = getNumberVal(assetPoolST[0].val)
        if val is not None and val != totalInterest:  # 收入合计
            ckResult.append("{4}{2}:{3}:【资产池情况】{0} [{1}] 收入账户加总值{5}不相等:10".format(assetPoolST[0].desc, val, assetPoolST[0].cell, 0, cellsInterest, totalInterest))
        val = getNumberVal(assetPoolST[1].val)
        if val is not None and val != totalPrincipal:  # 本金合计
            ckResult.append("{4}{2}:{3}:【资产池情况】{0} [{1}] 本金账户加总值{5}不相等:11".format(assetPoolST[1].desc, val, assetPoolST[1].cell, 0, cellsPrincipal, totalPrincipal))

    for i in range(5, 10):
       termVal = getNumberVal(ds[5][i].val)
       if termVal is not None and termVal > 10950.0:
           ckResult.append("{1}:{2}:期限特征本期期末数值{0} 大于30年最大天数12*30，错误:12".format(termVal, ds[5][i].cell, 0))



    defaultAmount = getNumberVal(ds[6][0].val)
    assetAmt = getNumberVal(ds[5][2].val)
    if defaultAmount is not None and assetAmt is not None and defaultAmount != 0 and defaultAmount > assetAmt:
        ckResult.append("{2};{4}:{3}:【资产池累计违金额】{0}超出当期资产池余额{1}:13".format(defaultAmount, assetAmt, ds[6][0].cell, 0, ds[5][2].cell))

    #累计违约率有可能会降低，所以暂时不验证是否逐期递增，只验证是否每期的数字都在0~1以内
    for d in ds[7]:
        cdr = getNumberVal(d[1].val)
        if cdr is not None and cdr != 0:
            if (str(d[1].val).endswith('%') and cdr > 100) or (not str(d[1].val).endswith('%') and cdr > 1) or (str(d[1].val).endswith('%') and cdr < 0) or (not str(d[1].val).endswith('%') and cdr < 0):
                ckResult.append("{1}:{2}:【累计违约率】违约率数值 [{0}] 超出合理范围:14".format(cdr, d[1].cell, 0))

    return ckResult

def checkFileFormat(wb, cfgItems, MapPath):
    sheet = wb['Sheet1']
    ckResult = []
    tp = [[], [], [], [], [], [], [], []]
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']
        itemtype = cfgItem.attrib['type'] if 'type' in cfgItem.attrib else ''

        itemCkResult = []
        r = []
        if itemtype == 'SpecificCells':
            itemCkResult, r = specificCellsCheck(cfgItem, sheet)
        else:
            itemCkResult, r = specificColsCheck(cfgItem, sheet)

        tp[i] = r
        if len(itemCkResult) > 0:
            ckResult.append("【{0}】".format(itemdesc))
            ckResult.extend(itemCkResult)

    if len(ckResult) == 0:
        ckResult = getValidationResult(tp)
    if len(ckResult) > 0:
        writeSheetLog(wb, ckResult, MapPath)
    if len(ckResult) == 0:
        updateIndexForMapXml(wb, MapPath)
    return len(ckResult)
#给xml增加换行符
def indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i
#新建mamppingxml， multiply 用于判断是单个文件还是文件夹
def createXml(xmlPath, inputfile, outputfile, multiply):
    if multiply.lower() == 'false':
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "false")
            root.set("inputFile", inputfile)
            root.set("outputFile", outputfile)
            root.set("inputSheetIndex", "0")
            root.set("outputSheetIndex", "1")
            root.set("inputSheetName", "Sheet1")
            root.set("outputSheetName", r"格式检查")
            tree = XETree.ElementTree(root)  # 创建文档
            Mapping1 = XETree.Element('Mapping') #创建子节点
            Mapping1.set("description", r'源文件，目标文件对应情况')
            root.append(Mapping1)
            indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
    else:
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "true")
            tree = XETree.ElementTree(root)  # 创建文档
            # indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
#开始函数，解析xml获取源文件，目标文件
def main(configFilePath, dateId):
    global logtxtFilePath

    scriptFolder = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolder, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolder, 'Logs',
                                  '{0}.txt'.format(dateId))
    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    sourceFolderPath = cfgRoot.attrib['sourcefolder']
    dir_path = scriptFolder + '\\MappingXml\\'  # mapping文件存放路径
    mappingPath = dir_path + dateId + '.xml'
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)

    if os.path.isfile(sourceFolderPath):
        fileName = os.path.basename(sourceFolderPath)
        if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告数据提取;'):
            msg = "【跳过】文件名称不符合，已跳过文件{0}".format(fileName)
            print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
            writeLog(msg)

        msg = "\n{0}".format(sourceFolderPath)
        writeLog(msg)

        fileNameAry = fileName.split(';')
        if len(fileNameAry) != 4:
            msg = "【文件名错误】文件名称命名不规范"
            writeLog(msg)

        paymentPeriodID = 0

        paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
        if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
            msg = "【错误】文件名中的TrustCode或报告期数设置有误"
            writeLog(msg)

        createXml(mappingPath, sourceFolderPath, sourceFolderPath, 'false')

        excelwb = load_workbook(sourceFolderPath)
        if '格式检查' in excelwb.sheetnames:
            excelwb.remove(excelwb['格式检查'])
            excelwb.save(sourceFolderPath)

        if 'log' in excelwb.sheetnames:
            excelwb.remove(excelwb['log'])
            excelwb.save(sourceFolderPath)

        hasError = checkFileFormat(excelwb, cfgRoot, mappingPath)
        if hasError > 0:
            excelwb.save(sourceFolderPath)
            writeLog('【有格式错误】详情见文档[格式检查]sheet')

    elif os.path.isdir(sourceFolderPath):
        for dirPath, dirNames, fileNames in os.walk(sourceFolderPath):
            config = 1
            createXml(mappingPath, '', '', 'true')
            for fileName in fileNames:
                if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告数据提取;'):
                    msg = "【跳过】文件名称不符合，已跳过文件{0}".format(fileName)
                    print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
                    writeLog(msg)
                    continue

                sourceFilePath = os.path.join(dirPath, fileName)

                msg = "\n{0}".format(sourceFilePath)
                writeLog(msg)

                fileNameAry = fileName.split(';')
                if len(fileNameAry) != 4:
                    msg = "【文件名错误】文件名称命名不规范"
                    writeLog(msg)
                    continue

                paymentPeriodID = 0

                paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
                if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
                    msg = "【错误】文件名中的TrustCode或报告期数设置有误"
                    writeLog(msg)
                    continue

                mulPath = dir_path + dateId + '_' + str(config) + '.xml'
                createXml(mulPath, sourceFilePath, sourceFilePath, 'false')
                config += 1
                tree = XETree.parse(mappingPath)
                root = tree.getroot()
                MapPath = XETree.Element('filename')  # 创建节点,单个文件的mapping文件
                MapPath.set("path", mulPath)
                root.append(MapPath)
                indent(root)
                tree.write(mappingPath, encoding='utf-8', xml_declaration=True)

                excelwb = load_workbook(sourceFilePath)
                if '格式检查' in excelwb.sheetnames:
                    excelwb.remove(excelwb['格式检查'])
                    excelwb.save(sourceFilePath)
                lenSheet = len(excelwb.sheetnames)
                for i in range(lenSheet):
                    if 'log' in excelwb.sheetnames[lenSheet - i-1]:
                        sheetName = excelwb.sheetnames[lenSheet - i - 1]
                        excelwb.remove(excelwb[sheetName])
                excelwb.save(sourceFilePath)

                hasError = checkFileFormat(excelwb, cfgRoot, mulPath)
                if hasError > 0:
                    excelwb.save(sourceFilePath)
                    writeLog('【有格式错误】详情见文档[格式检查]sheet')
                elif hasError == 0:
                    excelwb.save(sourceFilePath)
                    writeLog('通过校验，无错误！')




