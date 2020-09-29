import os
import re
import os.path
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
sourceExcel = None
destExcel = None
xmlPath = None
#mapping文件添加map节点
def addNodeForMapXml(MapPath, sourcecell, destCell, comment, sourcecolor, destcolor):
    tree = XETree.parse(MapPath)
    root = tree.getroot()
    node = root.find("Mapping")
    MapNode = XETree.Element('map')  # 创建节点,单个文件的mapping文件
    MapNode.set("destcolor", destcolor)
    MapNode.set("sourcecolor", sourcecolor)
    MapNode.set("comment", comment)
    MapNode.set("destCell", destCell)
    MapNode.set("sourcecell", sourcecell)
    node.append(MapNode)
    indent(node)
    tree.write(MapPath, encoding='utf-8', xml_declaration=True)
    return
#值的可用性处理
def updateValue(value,datatype):  #根据类型进行字符串处理，D日期，P百分数，S字符串不做处理，F或不填按数字处理
    value = str(value)
    if value is None:
        if datatype == 'F':
            return "0"
        elif datatype == "P":
            return "0.00%"
        else:
            return ""
    if datatype == 'D':
        # dateFormat = '[^\d/\d/\d]'
        # '2016年3月1日至2016年3月31日'
        value = re.sub('.*至', '', value)
        value = re.sub('\D$', '', value)
        value = re.sub(r'\D', r'/', value)
    elif datatype =='P':
        if type(value) == float :
            value = str(value * 100) + '%'
        else:
            Persentage = '[^\d%|\d.\d%]'
            value = re.sub(Persentage, '', value)
    elif datatype =='S':
        return value
    else:
        floatFormat = '[^\d|\d.\d]'
        value = re.sub(floatFormat, '', value)
        lenth = len(value.split('.')) - 2
        value = value.replace('.', '', lenth)
        if value == '' or value == '.':
            value = '0'
        value = float(value)
    return value
#值类型校验
def checkData(value,datatype): #值校验，是否符合对应类型
    value = str(value)
    date = '^\d{4}\D\d{1,2}\D\d{1,2}\D?$'
    num ='^[-]?\d*[.]?\d*$'
    persentige= '^[-]?\d*[.]?\d*[%]$'
    if value is None or value == '':
        return False
    flag = 0
    if datatype == 'D':
        flag = re.search(date, value)
    elif datatype == 'P':
        flag = re.search(persentige, value)
    elif datatype == 'S':
        return True
    else:
        flag = re.search(num, value)

    if str(flag) == 'None':
        return False
    else:
        return True
#列字母转数字 A转1
def letterToint(s):
    letterdict = {}
    for i in range(26):
        letterdict[chr(ord('A') + i)] = i + 1
    output = 0
    for i in range(len(s)):
        output = output * 26 + letterdict[s[i]]
    return output
#数字转列字母 1转A
def intToletter(i):
    if type(i) is not int:
        return i
    str = ''
    while (not (i // 26 == 0 and i % 26 == 0)):

        temp = 25

        if (i % 26 == 0):
            str += chr(temp + 65)
        else:
            str += chr(i % 26 - 1 + 65)

        i //= 26
        # print(str)
    # 倒序输出拼写的字符串
    return str[::-1]
#列表匹配
def extractForTable(cfgItem, MapPath):
    endrow = 0
    beginrow = 0
    # sheet_names = sourceExcel.get_sheet_names()
    sheet = sourceExcel["Sheet1"]

    # sheet_names = destExcel.get_sheet_names()
    destSheet = destExcel["Sheet1"]
    tempcell = findStr(sheet, cfgItem[0].attrib["anchor"], 1, 1)
    destBeginRow = int(cfgItem[1].attrib["beginrow"])
    sList = cfgItem[0].attrib["cols"].split(',')
    dList = cfgItem[1].attrib["cols"].split(',')
    datatype = cfgItem[1].attrib["datatype"].split(',')

    #如果差找不到，首行设置为NA
    if tempcell is None:
        for col in dList:
            destSheet[col + str(destBeginRow)].value = "NA"
        return
    # 判断在源文件中的起始行
    beginrow = tempcell.row + int(cfgItem[0].attrib["skiprows"]) + 1

    # 判断结束行
    # 如果范围已经确定，直接确定结束行
    if cfgItem[0].attrib["range"] != "":
        endrow = beginrow + int(cfgItem[0].attrib["range"]) - 1
    # 范围不确定，根据下一行字符确定结束行
    elif cfgItem[0].attrib["anchorend"] != "":
        endrow = findStr(sheet, cfgItem[0].attrib["anchorend"], beginrow, 1).row
    else:
        # 找不到字符，则直到最后一个不为空行的为止
        limited = 1
        while sheet["A" + str(beginrow + limited)].value != "" and sheet["A" + str(beginrow + limited)].value is not None:
            limited += 1
        endrow = beginrow + limited - 1

    # 粘贴数据
    transposition = 'false'
    if 'transposition' in cfgItem[0].attrib:
        transposition = cfgItem[0].attrib['transposition'].lower()
    if transposition == 'false':
        for row in range(beginrow, endrow + 1):
            for col in range(0, len(sList)):
                sCols = sList[col] + str(row)
                dCols = dList[col] + str(destBeginRow + row - beginrow)
                tempvalue = updateValue(sheet[sCols].value, datatype[col])
                isRight = checkData(tempvalue,datatype[col])
                if not isRight:
                    ErrorStr = ''
                    if datatype[col] == 'F':
                        ErrorStr += '（类型错误，为浮点数）'
                    elif datatype[col] == 'P':
                        ErrorStr += '（类型错误，为百分数,格式为xx%）'
                    elif datatype[col] == 'D':
                        ErrorStr += '（类型错误，为日期，格式为YYYY/MM/DD）'
                    fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
                    destSheet[dCols].fill = fill
                    destSheet[dCols].value = str(sheet[sCols].value) + ErrorStr
                    addNodeForMapXml(MapPath, sCols, dCols, '', '85, 170, 255', '255, 114, 116')
                else:
                    destSheet[dCols].value = tempvalue
                    addNodeForMapXml(MapPath, sCols, dCols, '', '85, 170, 255', '255, 230, 153')
    else:
        for col in range(0, len(sList)):
            for row in range(beginrow, endrow + 1):
                sCols = sList[col] + str(row)
                dCols = dList[(row-beginrow) if (row-beginrow) < len(sList) else (len(sList) - 1)] + str(destBeginRow + col)
                #dCols = dList[col] + str(destBeginRow + row - beginrow)
                tempvalue = updateValue(sheet[sCols].value, datatype[(row-beginrow) if (row-beginrow) < len(sList) else (len(sList) - 1)])
                isRight = checkData(tempvalue, datatype[(row-beginrow) if (row-beginrow) < len(sList) else (len(sList) - 1)])
                if not isRight:
                    fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
                    destSheet[dCols].fill = fill
                    destSheet[dCols].value = sheet[sCols].value
                    addNodeForMapXml(MapPath, sCols, dCols, '', '255, 230, 153', '255, 114, 116')
                else:
                    destSheet[dCols].value = tempvalue
                    addNodeForMapXml(MapPath, sCols, dCols, '', '85, 170, 255', '255, 230, 153')
#查找匹配的cell单元格
def findStr(sheet, key, startRow, startCol):
    for col in range(startCol, sheet.max_column):
        for row in range(startRow, sheet.max_row):
            cell = sheet[intToletter(col) + str(row)]
            if cell is None:
                return None
            if cell.value is not None and cell.value != "":
                if key in str(cell.value).replace(' ', ''):
                    return cell
#插入map表，插入数据到目标文件
def keyValueToDestExcel(tempcell, dCols, dNode, isFind, MapPath):
    sheetD = destExcel['Sheet1']
    datatype = 'F'
    value = '0.0'
    if 'datatype' in dNode.attrib:
        datatype = dNode.attrib['datatype']
    if isFind == 1:
        value = tempcell.value
    value = updateValue(value, datatype)
    isRight = checkData(value, datatype)
    if not isRight or isFind == 0:
        fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        sheetD[dCols].fill = fill
        value = '0.00'
        if isFind == 0:
            value = 'NA'
        if tempcell is not None:
            sCols = tempcell.column + str(tempcell.row)
            addNodeForMapXml(MapPath, sCols, dCols, '', '85, 170, 255', '255, 114, 116')
    else:
        sCols = tempcell.column + str(tempcell.row)
        addNodeForMapXml(MapPath, sCols, dCols, '', '85, 170, 255', '255, 230, 153')
    sheetD[dCols] = value

#根据key查找cell单元格
def extractForKeyValue(cfgItem, MapPath):
    sheetS = sourceExcel['Sheet1']

    sNode = cfgItem.find('source')
    dNode = cfgItem.find('dest')
    dCols = dNode.attrib['cols']
    anchors = sNode.attrib['anchor'].strip().split(';')
    length = len(anchors)
    beginCol = 1
    beginRow = 1
    sCols = sNode.attrib['cols']
    isFind = 0;
    tempcell = None
    for i in range(length):
        tempcell = findStr(sheetS, anchors[i], beginRow, beginCol)
        if tempcell is not None:
            beginRow = tempcell.row
            beginCol = letterToint(tempcell.column) + 1
            if i == length - 1:
                isFind = 1
                break
    if isFind == 1:
        destCell = sheetS[sCols + str(tempcell.row)]
        keyValueToDestExcel(destCell, dCols, dNode, isFind, MapPath)
    if isFind == 0:
        keyValueToDestExcel(tempcell, dCols, dNode, isFind, MapPath)
#根据类型选择列表匹配，还是值匹配
def selectType(cfgItems, MapPath):
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        istable = cfgItem.attrib['istable'].lower()
        if istable == 'true':
            extractForTable(cfgItem, MapPath)
        else:
            extractForKeyValue(cfgItem, MapPath)
#给定源文件，目标文件进行运行
def chooseExcel(inputFile, outputFile, templateFile, cfgRoot, MapPath):
    global sourceExcel
    global destExcel
    if os.path.exists(outputFile):
        os.remove(outputFile)
    open(outputFile, "wb").write(open(templateFile, "rb").read())
    sourceExcel = load_workbook(inputFile)
    destExcel = load_workbook(outputFile)
    selectType(cfgRoot, MapPath)
    destExcel.save(outputFile)
#给xml增加换行符
def indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i
#新建mamppingxml， multiply 用于判断是单个文件还是文件夹
def createXml(xmlPath, inputfile, outputfile, multiply):
    if multiply.lower() == 'false':
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "false")
            root.set("inputFile", inputfile)
            root.set("outputFile", outputfile)
            root.set("inputSheetIndex", "0")
            root.set("outputSheetIndex", "0")
            root.set("inputSheetName", "Sheet1")
            root.set("outputSheetName", r"Sheet1")
            tree = XETree.ElementTree(root)  # 创建文档
            Mapping1 = XETree.Element('Mapping') #创建子节点
            Mapping1.set("description", r'源文件，目标文件对应情况')
            root.append(Mapping1)
            indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
    else:
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "true")
            tree = XETree.ElementTree(root)  # 创建文档
            # indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
#开始函数，解析xml获取源文件，目标文件
def main(configFilePath, dateId):
    global DATANOTFOUND
    global cdfp
    global xmlPath

    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    input = cfgRoot.attrib['input']
    output = cfgRoot.attrib['output']
    templateFilePath = cfgRoot.attrib['template']
    dir_path = os.path.dirname(os.path.abspath(__file__)) + '\\MappingXml\\'  #mapping文件存放路径
    mappingPath = dir_path + dateId + '.xml'
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)

    if os.path.isfile(input):    #判断是文件还是文件夹
        filename = os.path.basename(input)
        suffix = filename.split('.')[-1]
        if suffix == 'xlsx':
            outputFile = os.path.join(output, filename)
            createXml(mappingPath, input, outputFile, 'false')
            chooseExcel(input, outputFile, templateFilePath, cfgRoot, mappingPath)
    elif os.path.isdir(input):
        for parent, dirnames, filenames in os.walk(input, followlinks=True):
            config = 1
            createXml(mappingPath, '', '', 'true')
            for filename in filenames:
                suffix = filename.split('.')[-1]
                if suffix == 'xlsx':
                    inputFile = os.path.join(parent, filename)
                    outputFile = os.path.join(output, filename)
                    mulPath = dir_path + dateId + '_' + str(config) + '.xml'
                    createXml(mulPath, inputFile, outputFile, 'false')
                    config += 1
                    tree = XETree.parse(mappingPath)
                    root = tree.getroot()
                    #node = root.find("result")
                    MapPath = XETree.Element('filename')  # 创建节点,单个文件的mapping文件
                    MapPath.set("path", mulPath)
                    root.append(MapPath)
                    indent(root)
                    tree.write(mappingPath, encoding='utf-8', xml_declaration=True)
                    chooseExcel(inputFile, outputFile, templateFilePath, cfgRoot, mulPath)


