# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import re
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import openpyxl.styles as sty


explainStr = dict()
explainStr["0"] = "首先请注意模板中绿框填写要求，填0还是填NA；其次对百分比，日期格式数据，请按要求填写"
explainStr["1"] = "首先请注意模板中绿框填写要求，填0还是填NA；第2期的期初余额减去第2期的本金还款等于第2期的期末余额，第1期的期末余额等于第2期的期初余额，以此类推，所以本金并不会越还越多，请确定是否取数取错地方了"
explainStr["2"] = "首先请注意模板中绿框填写要求，填0还是填NA；第1期的期初余额减去第1期的本金还款等于第0期的期末余额，第1期的期末余额等于第2期的期初余额，以此类推，所以本金并不会越还越多，请确定是否取数取错地方了"
explainStr["3"] = "首先请注意款模板中绿框填写要求，填0还是填NA；第2期的期初余额减去第2期的本金还等于第2期的期末余额，第1期的期末余额等于第2期的期初余额，以此类推。如果发现所有期都对不上请告知组长，寻找解决方法"
explainStr["4"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为在本金未还清之前，都得付利息，可联系实际思考一下。首先确定是否录错，如确实未录错（备注表请下载）那就备注：未录错，未修改，XXX层利息还款确实为XXX，发给所在组的组长"
explainStr["5"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为在本金未还清之前，都得付利息，至少有一层是有利息的。首先确定是否录错，如确实未录错（备注表请下载）那就备注：未录错，未修改，当期所有层的利息还款确实为XXX，发给所在组的组长"
explainStr["6"] = "首先请注意模板中绿框填写要求，填0还是填NA；现金流归集表是指立足于当期对于未来期回款的预测（比如当期是201909，那么现金流归集表指的是201910以及往后月份的一个回款预测），那么还本金额不可能越还越多，所以期初本金余额应小于等于应收本金余额，如遇特殊情况，请告知组长，寻找解决方法"
explainStr["7"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为分项相加等于总计（不良资产大部分不区分收入账和本金账，钱全放一个账户），但有很多不等的情况，或者同一个数据在多个表披露，但是不等（这种情况请及时询问），遇见这样的，首先保证和PDF一致，未录错，一定不要漏填（一定注意上期转存、转存下期等字段数据一定要录入，因为加上上期转存的大合计并不是当期的实际收入，需要录入后，系统自动减去）（特殊资产类型需要加上手续费等作为收入账数据，请联系实际思考信用卡），其次根据实际情况备注（备注表请下载）发给所在组的组长"
explainStr["8"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为合计值等于分项加总，如遇特殊情况，首先保证未录错，其次根据实际情况备注（备注表请下载）发给所在组的组长"
explainStr["9"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为合计值等于分项加总，如遇特殊情况，首先保证未录错，其次根据实际情况备注（备注表请下载）发给所在组的组长"
explainStr["10"] = "首先请注意模板中绿框填写要求，填0还是填NA；一般认为合计值等于分项加总，如遇特殊情况，首先保证未录错，其次根据实际情况备注（备注表请下载）发给所在组的组长"
explainStr["11"] ="首先请注意模板中绿框填写要求，填0还是填NA；一般认为分项相加等于总计，如确实不等，先保证和PDF一致，未录错，一定不要漏填，其次根据实际情况备注（备注表请下载）发给所在组的组长"



logtxtFilePath = None
class cObj(object):
    def __init__(self, cell, desc, val):
        self.cell = cell
        self.desc = desc
        self.val = val
#文件无错误，更新文件index
def updateIndexForMapXml(wb, MapPath):
    logSheet = wb.create_sheet("log")
    logSheet["A1"] = r"校验完成，该文件通过校验，无错误！"
#mapping文件添加map节点
def addNodeForMapXml(MapPath, sourcecell, destCell, comment, sourcecolor, destcolor):
    tree = XETree.parse(MapPath)
    root = tree.getroot()
    node = root.find("Mapping")
    MapNode = XETree.Element('map')  # 创建节点,单个文件的mapping文件
    MapNode.set("destcolor", destcolor)
    MapNode.set("sourcecolor", sourcecolor)
    MapNode.set("comment", comment)
    MapNode.set("destCell", destCell)
    MapNode.set("sourcecell", sourcecell)
    node.append(MapNode)
    indent(node)
    tree.write(MapPath, encoding='utf-8', xml_declaration=True)
    return

def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def is_number(s):
    try:
        float(s.replace(',', ''))
        return True
    except ValueError:
        return False

def getNumberVal(val):
    if isinstance(val, int) or isinstance(val, float):
        return val
    val = val.replace(' ', '')
    if val == 'NA':
        return None
    elif val == '-' or val == '':
        return 0.0
    else:
        return float(val.replace('\t', '').replace('\n', '').replace(',', '').replace('%', '').replace(' ', ''))

def markCellError(sheet, ctag, errtype):
    color = 'E93936' #datatype
    if errtype == 1:#empty
        color = 'AA2927'
    sheet[ctag].fill=sty.PatternFill(fill_type='solid', fgColor=color)

def checkCellValue(sheet, ctag, cvalue, cdtype):
    if ((cdtype == 'mumber' or cdtype == 'rate') and isinstance(cvalue, float)
        ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return 1, 0, 0

    cvalue = str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '')

    if cvalue == 'NA' or cvalue == '-':
        return 1, 1, 0

    if cvalue == '':
        markCellError(sheet, ctag, 1)
        return 0, 0, 1

    dtcheck = 0
    if cdtype == 'int':
        dtcheck = 1 if cvalue.isdigit() else 0
    elif cdtype == 'number':
        dtcheck = 1 if is_number(cvalue) else 0
    elif cdtype == 'date':
        reg = r"(\d{4}[-/]\d{1,2}([-/]\d{1,2})?)|((\d{1,2}[-/])?\d{1,2}[-/]\d{4})|(\d{4}年\d{1,2}月(\d{1,2}日)?)|^(\d{5,6})$|^(\d{8})$"
        dtcheck = 1 if re.search(reg, cvalue) is not None else 0
    elif cdtype == 'rate':
        dtcheck = 1 if is_number(cvalue.rstrip('%')) else 0
    else:  # string
        dtcheck = 1

    if dtcheck == 0:
        markCellError(sheet, ctag, 2)
        return 0, 0, 2

    return 1, 0, 0

def specificCellsCheck(cfgItem, sheet):
    checkmsg = []
    r = []
    for cell in cfgItem:
        ctag = cell.tag
        cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

        isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)

        if isValid != 1:
            etype = "不能为空" if errorType == 1 else "格式错误"
            checkmsg.append("{0}:{3}:[{1}] - {2}:0".format(ctag, cdesc, etype, errorType))

        r.append(cObj(ctag, cdesc, cvalue))

    return checkmsg, r

def specificColsCheck(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 1
    breakRowsLoop = 0
    rs = []
    while rStart <= rEnd:  # rows loop
        r = []
        for cell in cNode: # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break #break cells loop

            isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)
            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            if isValid != 1:
                etype = "不能为空" if errorType == 1 else "格式错误"
                checkmsg.append("{0}:{3}:[{1}] - {2}:0".format(ctag, cdesc, etype, errorType))

            r.append(cObj(ctag, cdesc, cvalue))

        if (isFirstRow == 1 and firstRowAllNA == 1) or breakRowsLoop == 1:
            rStart = rEnd + 1 #break rows loop
        else:
            isFirstRow = 0
            rStart += 1
            rs.append(r)

    return checkmsg, rs

def writeSheetLog(wb, res, mapPath):
    logSheet = wb.create_sheet("格式检查")
    i = 1
    logSheet["A1"] = "错误信息："
    logSheet["B1"] = "解释信息："
    while i <= len(res):
        if len(res[i-1].split(':')) == 4:
            sourceCell = res[i-1].split(':')[0]
            destCell = "A{0}".format(i+1)
            explainCell = "B{0}".format(i+1)
            errorType = res[i-1].split(':')[1]
            comment = res[i-1].split(':')[2]
            explainType = res[i-1].split(':')[3]
            sourceColor = "255,230,153"
            destColor = ''
            if errorType == 1:
                destColor = "170,41,39"
            elif errorType == 2:
                destColor = "233,57,54"
            else:
                destColor = "255,230,153"
            addNodeForMapXml(mapPath, sourceCell, destCell, comment, sourceColor, destColor)
            logSheet[destCell] = str(sourceCell + '：' + comment)
            logSheet[explainCell] = explainStr[explainType]
            i += 1
        else:
            logSheet["A{0}".format(i+1)] = res[i - 1]
            i += 1

def getValidationResult(ds):
    ckResult = []

    bondsAmt = 0.0
    cells = ''
    for d in ds[0]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestPaied = getNumberVal(d[3].val)
        pendAmt = getNumberVal(d[4].val)

        if pInterestPaied is not None:
            bondsAmt += pInterestPaied
            cells += str(d[3].cell) + ';'
        if pendAmt is not None and pstartAmt is not None and pendAmt > pstartAmt:
            ckResult.append("{2};{3}:{4}:【收益分配记录】债券期末余额 [{0}] 大于期初余额{1},数据对应关系错误:1".format(pendAmt, pstartAmt, d[4].cell, d[1].cell, 0))
        if pPrincipalAmt is not None and pstartAmt is not None and pPrincipalAmt > pstartAmt:
            ckResult.append("{2};{3}:{4}:【收益分配记录】债券本金金额 [{0}] 大于期初余额{1}，数据对应关系错误:2".format(pPrincipalAmt, pstartAmt, d[2].cell, d[1].cell, 0))

        if pstartAmt is not None and pPrincipalAmt is not None and pendAmt is not None and round(pstartAmt - pPrincipalAmt, 4) != pendAmt:
            ckResult.append("{3};{4};{5}:{6}:【收益分配记录】期初余额 [{0}] 减去本金金额 [{1}] 不等于期末余额 [{2}] 错误:3".format(pstartAmt,
                                                                                                   pPrincipalAmt, pendAmt, d[1].cell, d[2].cell, d[4].cell, 0))
        if pInterestPaied is not None and pPrincipalAmt is not None and pPrincipalAmt > 0 and pInterestPaied == 0:
            ckResult.append(
                "{0};{1}:{2}:【收益分配记录】本金还款金额存在，利息还款金额缺失，错误:4".format(d[2].cell, d[3].cell, 0))
        # if pInterestPaied is not None and pendAmt is not None and pInterestPaied >= pendAmt:
        #     ckResult.append(
        #         "{0};{1}:{2}:【收益分配记录】利息还款金额大于等于剩余金额，错误".format(d[3].cell, d[4].cell, 0))

    bondsAmt = round(bondsAmt, 4)
    if bondsAmt == 0 and cells != '':
        ckResult.append(
            "{0}:{1}:【收益分配记录】当期数据中，必须有一层利息还款金额不为0，错误:5".format(cells[:-1], 0))

    lenCash = len(ds[2])
    nowCount = 1
    for d in ds[2]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestAmt = getNumberVal(d[3].val)

        if nowCount == lenCash:
            if pstartAmt is not None and pPrincipalAmt is not None and pstartAmt != pPrincipalAmt:
                ckResult.append("{2};{3}:{4}:【现金流归集表】当期期初本金金额 [{0}] 不等于应收本金金额 [{1}] 数据对应关系错误:6".format(pstartAmt,
                                                                                                       pPrincipalAmt, d[1].cell, d[2].cell, 0))
        else:
            if pstartAmt is not None and pPrincipalAmt is not None and pstartAmt <= pPrincipalAmt:
                ckResult.append("{2};{3}:{4}:【现金流归集表】当期期初本金金额 [{0}] 小于等于应收本金金额 [{1}] 数据对应关系错误:6".format(pstartAmt,
                                                                                                        pPrincipalAmt, d[1].cell, d[2].cell, 0))
        nowCount += 1

    CurrentAmount = 0.00  #本期回收金额加总值
    CurrentCells = ''
    CumulativeAmount = 0.00     #累计回收金额加总值
    CumulativeCells = ''
    EstimatedAmount = 0.00      #逾期回收金额加总值
    EstimatedCells = ''
    RecoveryRate = 0.00         #回收率加总值
    RecoveryCells = ''
    for i in range(8):
        count1 = getNumberVal(ds[3][i][1].val)
        count2 = getNumberVal(ds[3][i][2].val)
        count3 = getNumberVal(ds[3][i][3].val)
        rate = getNumberVal(ds[3][i][4].val)

        CurrentCells += str(ds[3][i][1].cell) + ';'
        if count1 is not None:
            CurrentAmount += count1

        CumulativeCells += str(ds[3][i][2].cell) + ';'
        if count2 is not None:
            CumulativeAmount += count2

        EstimatedCells += str(ds[3][i][3].cell) + ';'
        if count3 is not None:
            EstimatedAmount += count3

        RecoveryCells += str(ds[3][i][4].cell) + ';'
        if rate is not None:
            RecoveryRate += rate


    CurrenTotal = getNumberVal(ds[3][8][1].val)
    CumulativeTotal = getNumberVal(ds[3][8][2].val)
    EstimatedTotal = getNumberVal(ds[3][8][3].val)
    RecoveryTotal = getNumberVal(ds[3][8][4].val)
    if CurrenTotal is not None and abs(CurrenTotal - round(CurrentAmount, 4))>1:
        ckResult.append("{2}{3}:{4}:现金流流入情况本期回收金额加总值：{0} 不等于合计值：{1}，错误:7".format(CurrentAmount,
                                                                               CurrenTotal, CurrentCells, ds[3][8][1].cell, 0))
    if CumulativeTotal is not None and abs(CumulativeTotal - round(CumulativeAmount, 4))>1:
        ckResult.append("{2}{3}:{4}:现金流流入情况累计回收金额加总值：{0} 不等于合计值：{1}，错误:8".format(CumulativeAmount,
                                                                               CumulativeTotal, CumulativeCells, ds[3][8][2].cell, 0))
    if EstimatedTotal is not None and abs(EstimatedTotal - round(EstimatedAmount, 4))>1:
        ckResult.append("{2}{3}:{4}:现金流流入情况预计回收金额加总值：{0} 不等于合计值：{1}，错误:9".format(EstimatedAmount,
                                                                               EstimatedTotal, EstimatedCells, ds[3][8][3].cell, 0))
    if RecoveryTotal is not None and RecoveryTotal != round(RecoveryRate, 4):
        ckResult.append("{2}{3}:{4}:现金流流入情况回收率加总值：{0} 不等于合计值：{1}，错误:10".format(RecoveryRate,
                                                                               RecoveryTotal, RecoveryCells, ds[3][8][4].cell, 0))
    totalAmount = 0.00   #现金流流出情况费用加总值
    totalCells = ''
    for i in range(13):
        termVal = getNumberVal(ds[4][i].val)
        totalCells += str(ds[4][i].cell) + ';'
        if termVal is not None:
            totalAmount += termVal

    FeeTotal = getNumberVal(ds[4][13].val)
    if FeeTotal is not None and abs(FeeTotal - round(totalAmount, 4)) > 1:
        ckResult.append("{2}{3}:{4}:现金流流出情况费用加总值：{0} 不等于总计值：{1}，错误:11".format(totalAmount, FeeTotal, totalCells, ds[4][13].cell, 0))

    outFeeTotal = 0.0
    inFeeTotal = getNumberVal(ds[3][8][1].val)
    outFeeCells = ''
    for i in range(13, 23):
        termVal = getNumberVal(ds[4][i].val)
        outFeeCells += str(ds[4][i].cell) + ';'
        if termVal is not None:
            outFeeTotal += termVal

    Balance = getNumberVal(ds[5][1].val)
    outFeeCells += str(ds[5][1].cell) + ';'
    if Balance is not None:
        outFeeTotal += Balance
    # if inFeeTotal is not None and abs(inFeeTotal - round(outFeeTotal, 4))>1:
    #     ckResult.append(
    #         "{2}{3}:{4}:现金流流出金额加总值：{0} 不等于现金流流入总计值：{1}，错误".format(outFeeTotal, inFeeTotal, outFeeCells, ds[3][8][1].cell, 0))

    return ckResult

def checkFileFormat(wb, cfgItems, MapPath):
    sheet = wb['Sheet1']
    ckResult = []
    tp = [[], [], [], [], [], [], [], []]
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']
        itemtype = cfgItem.attrib['type'] if 'type' in cfgItem.attrib else ''

        itemCkResult = []
        r = []
        if itemtype == 'SpecificCells':
            itemCkResult, r = specificCellsCheck(cfgItem, sheet)
        else:
            itemCkResult, r = specificColsCheck(cfgItem, sheet)

        tp[i] = r
        if len(itemCkResult) > 0:
            ckResult.append("【{0}】".format(itemdesc))
            ckResult.extend(itemCkResult)

    if len(ckResult) == 0:
        ckResult = getValidationResult(tp)
    if len(ckResult) > 0:
        writeSheetLog(wb, ckResult, MapPath)
    if len(ckResult) == 0:
        updateIndexForMapXml(wb, MapPath)
    return len(ckResult)
#给xml增加换行符
def indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i
#新建mamppingxml， multiply 用于判断是单个文件还是文件夹
def createXml(xmlPath, inputfile, outputfile, multiply):
    if multiply.lower() == 'false':
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "false")
            root.set("inputFile", inputfile)
            root.set("outputFile", outputfile)
            root.set("inputSheetIndex", "0")
            root.set("outputSheetIndex", "1")
            root.set("inputSheetName", "Sheet1")
            root.set("outputSheetName", r"格式检查")
            tree = XETree.ElementTree(root)  # 创建文档
            Mapping1 = XETree.Element('Mapping') #创建子节点
            Mapping1.set("description", r'源文件，目标文件对应情况')
            root.append(Mapping1)
            indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
    else:
        if not os.path.exists(xmlPath):
            # open(configFilePath, "wb").write(bytes("", encoding="utf-8"))
            root = XETree.Element('result')  # 创建节点
            root.set("multiply", "true")
            tree = XETree.ElementTree(root)  # 创建文档
            # indent(root)  # 增加换行符
            tree.write(xmlPath, encoding='utf-8', xml_declaration=True)
#开始函数，解析xml获取源文件，目标文件
def main(configFilePath, dateId):
    global logtxtFilePath

    scriptFolder = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolder, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolder, 'Logs',
                                  '{0}.txt'.format(dateId))
    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    sourceFolderPath = cfgRoot.attrib['sourcefolder']
    dir_path = scriptFolder + '\\MappingXml\\'  # mapping文件存放路径
    mappingPath = dir_path + dateId + '.xml'
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)

    if os.path.isfile(sourceFolderPath):
        fileName = os.path.basename(sourceFolderPath)
        if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告数据提取;'):
            msg = "【跳过】文件名称不符合，已跳过文件{0}".format(fileName)
            print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
            writeLog(msg)

        msg = "\n{0}".format(sourceFolderPath)
        writeLog(msg)

        fileNameAry = fileName.split(';')
        if len(fileNameAry) != 4:
            msg = "【文件名错误】文件名称命名不规范"
            writeLog(msg)

        paymentPeriodID = 0

        paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
        if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
            msg = "【错误】文件名中的TrustCode或报告期数设置有误"
            writeLog(msg)

        createXml(mappingPath, sourceFolderPath, sourceFolderPath, 'false')

        excelwb = load_workbook(sourceFolderPath)
        if '格式检查' in excelwb.sheetnames:
            excelwb.remove(excelwb['格式检查'])
            excelwb.save(sourceFolderPath)

        if 'log' in excelwb.sheetnames:
            excelwb.remove(excelwb['log'])
            excelwb.save(sourceFolderPath)

        hasError = checkFileFormat(excelwb, cfgRoot, mappingPath)
        if hasError > 0:
            excelwb.save(sourceFolderPath)
            writeLog('【有格式错误】详情见文档[格式检查]sheet')

    elif os.path.isdir(sourceFolderPath):
        for dirPath, dirNames, fileNames in os.walk(sourceFolderPath):
            config = 1
            createXml(mappingPath, '', '', 'true')
            for fileName in fileNames:
                if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告数据提取;'):
                    msg = "【跳过】文件名称不符合，已跳过文件{0}".format(fileName)
                    print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
                    writeLog(msg)
                    continue

                sourceFilePath = os.path.join(dirPath, fileName)

                msg = "\n{0}".format(sourceFilePath)
                writeLog(msg)

                fileNameAry = fileName.split(';')
                if len(fileNameAry) != 4:
                    msg = "【文件名错误】文件名称命名不规范"
                    writeLog(msg)
                    continue

                paymentPeriodID = 0

                paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
                if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
                    msg = "【错误】文件名中的TrustCode或报告期数设置有误"
                    writeLog(msg)
                    continue

                mulPath = dir_path + dateId + '_' + str(config) + '.xml'
                createXml(mulPath, sourceFilePath, sourceFilePath, 'false')
                config += 1
                tree = XETree.parse(mappingPath)
                root = tree.getroot()
                MapPath = XETree.Element('filename')  # 创建节点,单个文件的mapping文件
                MapPath.set("path", mulPath)
                root.append(MapPath)
                indent(root)
                tree.write(mappingPath, encoding='utf-8', xml_declaration=True)

                excelwb = load_workbook(sourceFilePath)
                if '格式检查' in excelwb.sheetnames:
                    excelwb.remove(excelwb['格式检查'])
                    excelwb.save(sourceFilePath)

                if 'log' in excelwb.sheetnames:
                    excelwb.remove(excelwb['log'])
                    excelwb.save(sourceFilePath)

                hasError = checkFileFormat(excelwb, cfgRoot, mulPath)
                if hasError > 0:
                    excelwb.save(sourceFilePath)
                    writeLog('【有格式错误】详情见文档[格式检查]sheet')
                elif hasError == 0:
                    excelwb.save(sourceFilePath)
                    writeLog('通过校验，无错误！')




