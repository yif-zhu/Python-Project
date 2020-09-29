# -*- coding: utf-8 -*-
"""
Created on Wed Jan 15 14:38:26 2020

@author: HUAWEI
"""


#%%
dbConnectionStr = None
errtxtFilePath = None
Error=0
Excelfilepath=None
def ErrorMessage(TrustId,TrustCode,TrustName,Excelfilepath,ErrorInformation):
    wb = openpyxl.load_workbook(errtxtFilePath) #读取xlsx文件
    ws=wb['Sheet']
    maxrow=ws.max_row
    maxrow+=1
    print(maxrow)
    ws.cell(maxrow,1,TrustId).value
    ws.cell(maxrow,2,TrustCode).value
    ws.cell(maxrow,3,TrustName).value
    ws.cell(maxrow,4,Excelfilepath).value
    ws.cell(maxrow,5,ErrorInformation).value
    wb.save(errtxtFilePath)


def PoolDistributions(Excelfilepath,FileTrustCode,TrustName):
    global TrustId,Error,errtxtFilePath
    Error=0
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    try:
        selectId="select TrustId from TrustManagement.Trust where TrustCode='{}'".format(FileTrustCode)
        b1.execute(selectId)
        TrustId=b1.fetchone()[0]
    except:
        Error=1
        C0='文件TrustCode【{}】与系统TrustCoden不匹配请检查!'.format(FileTrustCode)
        ErrorMessage('#',FileTrustCode,TrustName,Excelfilepath,C0)
        print(TrustName,C0)
        return
    
    data=pd.read_excel(Excelfilepath,header=1)
    
    PoolDistributions_columns=['PaymentPeriodID','资产池分布类型','DistributionType','DatabaseItem','BucketSequenceNo','Bucket','Amount','AmountPercentage','Count','CountPercentage']
    
    for columnsP in PoolDistributions_columns:
        if columnsP not in data.columns:
            Error=1
            print(TrustId,TrustName,'表格字段错误不能含有中文或字段缺失!')
            C9='表格字段错误不能含有中文或字段缺失!'
            Error=1
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C9)
            return
            

    
    data.dropna(subset=['Amount','Bucket','AmountPercentage'],inplace=True)
    data=data[['PaymentPeriodID','DistributionType','DatabaseItem','BucketSequenceNo','Bucket','Amount','AmountPercentage','Count','CountPercentage']]
    DatabaseItem=data['DatabaseItem'].drop_duplicates()
    for o in DatabaseItem:
        AmountPercentageSum=0
        CountPercentageSum=0
        BucketSequenceNolist=[]
        for i in data.index:
            DatabaseItem=data.loc[i][2]
            if DatabaseItem==o:
                try:
                    PaymentPeriodID=data.loc[i][0].astype(float)
                    BucketSequenceNo=data.loc[i][3].astype(float)
                    BucketSequenceNolist.append(BucketSequenceNo)
                    AmountPercentage=data.loc[i][6].astype(float)
                    CountPercentage=data.loc[i][8].astype(float)
                    Amount=data.loc[i][5].astype(float)
                    Count=data.loc[i][7].astype(float)

                    DistributionType=data.loc[i][1]
                    Bucket=data.loc[i][4]


                except:
                    Error=1
                    C1="{},'BucketSequenceNo'or'AmountPercentage'or'CountPercentage'or'Amount','Count',数据类型错误应为数值类型".format(o)
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C1)
                    print(TrustName,C1)
                    return

				#判断分布是否带单位

                if DistributionType=='ApprovalAmount' and '元'  not in Bucket and '万元'  not in str(Bucket):
                    Error=1
                    DWJ1='合同金额分布未带单位!请检查!'
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,DWJ1)
                    print('合同金额分布未带单位!请检查!')
                        
                if DistributionType=='CurrentPrincipalBalance' and '元'  not in Bucket and '万元'  not in str(Bucket):
                    Error=1
                    DWJ2='剩余本金分布未带单位!请检查!'
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,DWJ2)
                    print('剩余本金分布未带单位!请检查!')
                    
                if DistributionType=='LoanTerm' and '月' not in str(Bucket) and '天' not in str(Bucket) and '年' not in str(Bucket):
                    Error=1
                    DWJ3='合同期限分布未带单位!请检查!'
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,DWJ3)
                    print('合同期限分布未带单位!请检查!')
                    
                if DistributionType=='Seasoning' and '月' not in str(Bucket) and '天' not in str(Bucket) and '年' not in str(Bucket):
                    Error=1
                    DWJ4='账龄分布未带单位!请检查!'
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,DWJ4)
                    print('账龄分布未带单位!请检查!')
                    
                if DistributionType=='RemainingTerm' and '月' not in str(Bucket) and '天' not in str(Bucket) and '年' not in str(Bucket):
                    Error=1
                    DWJ5='剩余期限分布未带单位!请检查!'
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,DWJ5)
                    print('剩余期限分布未带单位!请检查!')
                    
                if DistributionType=='CustomerAge' and '岁' not in Bucket:
                    Error=1
                    DWJ6='年龄分布未带单位!请检查!'
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,DWJ6)
                    print('年龄分布未带单位!请检查!')
                    
                if DistributionType=='CustomerAnnualIncome' and '元'  not in Bucket and '万元'  not in str(Bucket):
                    Error=1
                    DWJ7='收入分布未带单位!请检查!'
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,DWJ7)
                    print('收入分布未带单位!请检查!')

                if PaymentPeriodID!=0:
                    Error=1
                    C8='PaymentPeriodID填写值【】错误!只能填【0】'.format(PaymentPeriodID)
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C8)
                    print('PaymentPeriodID填写值【】错误!只能填【0】'.format(PaymentPeriodID))
                    
                AmountPercentageSum+=AmountPercentage
                CountPercentageSum+=CountPercentage
                BucketSequenceNoset=set(BucketSequenceNolist)
                if len(BucketSequenceNolist)!=len(BucketSequenceNoset):
                    Error=1
                    C2="{},'BucketSequenceNo值重复'".format(o)
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C2)
                    print(TrustId,TrustName,C2)
                    
                    BucketSequenceNolist.clear()
                if Amount<Count:
                    Error=1
                    C3="{},Amount【{}】不应大于Count【{}】,可能数据录反请检查!".format(o,Amount,Count)
                    ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C3)
                    print(TrustId,TrustName,C3)
                    
                    
        if AmountPercentageSum-1>0.1 or AmountPercentageSum-1<-0.1:
            Error=1
            C4="{},AmountPercentage列相加【{}】不等于1(忽略精度影响+-0.1)请检查!".format(o,AmountPercentageSum)
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C4)
            print(TrustId,TrustName,C4)
            
        if CountPercentageSum-1>0.1 or CountPercentageSum-1<-0.1:
            Error=1
            C5="{},CountPercentage列相加【{}】不等于1(忽略精度影响+-0.1)请检查!".format(o,CountPercentageSum)
            ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C5)
            print(TrustId,TrustName,C5)
            
    print(Error)
    if Error==0:
        print(TrustId,TrustName,'池分布数据校验通过!')
        PoolDistributionsImport(Excelfilepath,FileTrustCode)

            

def PoolDistributionsImport(Excelfilepath,FileTrustCode):
    conn = pymssql.connect(host='172.16.6.143\mssql', user='sa', password='PasswordGS2017',
                           database='PortfolioManagement', charset='utf8')
    b1=conn.cursor()
    
    selectPoolDistributionsId="select TrustId from dbo.PoolDistributions1"
    b1.execute(selectPoolDistributionsId)
    CTrustId=b1.fetchall()
#    conn.commit()
    PoolDistributions1Id=[]
    for CTrustId in CTrustId:
        PoolDistributions1Id.append(CTrustId[0])
    if TrustId not in PoolDistributions1Id:
        DataPoolDistributions=pd.read_excel(Excelfilepath,header=1)
        DataPoolDistributionsImport=DataPoolDistributions[['PaymentPeriodID','DistributionType','BucketSequenceNo','Bucket','Amount','AmountPercentage','Count','CountPercentage']]
        DataPoolDistributionsImport=DataPoolDistributionsImport.dropna(subset=['Bucket','Amount'])
        
        for Cindex in DataPoolDistributionsImport.index:
            PaymentPeriodID=DataPoolDistributionsImport.loc[Cindex][0]
            DistributionTypeCode=DataPoolDistributionsImport.loc[Cindex][1]
            BucketSequenceNo=DataPoolDistributionsImport.loc[Cindex][2]
            Bucket=DataPoolDistributionsImport.loc[Cindex][3]
            Amount=DataPoolDistributionsImport.loc[Cindex][4]
            AmountPercentage=DataPoolDistributionsImport.loc[Cindex][5]
            Count=DataPoolDistributionsImport.loc[Cindex][6]
            CountPercentage=DataPoolDistributionsImport.loc[Cindex][7]

            if AmountPercentage<1:
                AmountPercentage=AmountPercentage*100

            if CountPercentage<1:
                CountPercentage=CountPercentage*100
            
            PoolDistributions1Insert="insert into dbo.PoolDistributions1(TrustId,PaymentPeriodID,DistributionTypeCode,BucketSequenceNo,Bucket,Amount,AmountPercentage,Count,CountPercentage) values({},{},'{}',{},N'{}',{},{},{},{})".format(TrustId,PaymentPeriodID,DistributionTypeCode,BucketSequenceNo,Bucket,Amount,AmountPercentage,Count,CountPercentage)
            b1.execute(PoolDistributions1Insert)
            conn.commit()
        b1.close()
        conn.close()
        C6='池分布导入完成!'
        ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C6)
        print("{},{},{},池分布导入完成!".format(TrustId,FileTrustCode,TrustName)) 
    else:
        C7="数据库池分布数据已存在跳过上传!"
        ErrorMessage(TrustId,FileTrustCode,TrustName,Excelfilepath,C7)
        print(TrustId,TrustName,'数据库池分布数据已存在跳过上传!')
		
def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        writeLog(str(ex))
        writeErr("\n【{0}】".format(sourceFilePath))
        writeErr(str(ex))
        print(str(ex))
        # raise ex
    finally:
        cnxn.close()

def InsertVerificationLog(userId, CheckType, IsSucess, filePath, result):
    sql = "exec TaskCollection.dbo.usp_InsertVerificationLog N'{0}',N'{1}',N'{2}',N'{3}',N'{4}'".format(userId,CheckType,IsSucess, filePath, result)
    execSQLCmd(sql)

def InsertTrusteeCheckByTrustId(userId, TrustId,ckResultLen):
    sql = "exec TaskCollection.dbo.usp_InsertTrusteeCheckByTrustId N'{0}',N'{1}',N'{2}',N'{3}',N'{4}'".format(userId,15, TrustId, 0, ckResultLen)
    execSQLCmd(sql)

if __name__=="__main__":
    import xlrd,xlwt
    import pandas as pd,numpy as np
    import numpy as np
    import os
    import time
    import sys
    import datetime
    import pymssql
    import pyodbc
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl import Workbook
    
    #path=r'\\172.16.6.143\StudentsProducts\1206新产品和受托报告\说明书\个人住房贷款\20组 工元宜居2019年第七期个人住房抵押贷款资产支持证券说明书'
    filepath = str(sys.argv[1])
    dateId = str(sys.argv[2])
    userId = str(sys.argv[3])
    dbConnectionStr="DRIVER={SQL Server};SERVER=172.16.6.143\MSSQL;DATABASE=TaskCollection;UID=sa;PWD=PasswordGS2017"

    errtxtFilePath = os.path.join('E:\Client\TrusteeTaskCollection\TrusteeError','Error_池分布校验结果_{0}.xlsx'.format(dateId))
    if not os.path.exists(errtxtFilePath):
        wb = Workbook()  # 新建工作簿
        ws1 = wb.active
        wb.save(errtxtFilePath)
        wb = openpyxl.load_workbook(errtxtFilePath) # 读取xlsx文件
        ws=wb['Sheet']
        maxrow=ws.max_row
        ws.cell(maxrow,1,'TrustId').value
        ws.cell(maxrow,2,'TrustCode').value
        ws.cell(maxrow,3,'TrustName').value
        ws.cell(maxrow,4,'Path').value
        ws.cell(maxrow,5,'ErrorInformation').value
        wb.save(errtxtFilePath)

    #循环遍查找指定文件夹
    ErrorList=[]
    for root, dirs, files in os.walk(filepath):
        for name in files:
            if not name.endswith('池分布.xlsx'):
                continue
            Excelfilepath = os.path.join(root,name)
            if ';' in name:
                SplitName=name.split(';')
                TrustCode=SplitName[0]
                TrustName=SplitName[1]
                PoolDistributions(Excelfilepath,TrustCode,TrustName)
                data=pd.read_excel(errtxtFilePath)
                ckResultLen=len(data)
                InsertTrusteeCheckByTrustId(userId, TrustId,ckResultLen)
                ErrorList.append(Error)
            else:
                Error=1
                ErrorList.append(Error)
                FileError='文件名称格式错误!应为[TrustCode;TrustName;池分布.xlsx]'
                ErrorMessage('*','*','*',Excelfilepath,FileError)
                print(name,"--文件名称格式错误!应为[TrustCode;TrustName;池分布.xlsx]")		

    if sum(ErrorList)==0:
        IsSucess=1
        result='计划说明书池分布校验通过!'
        InsertVerificationLog(userId, 18, IsSucess, errtxtFilePath, result)
    else:
        IsSucess=0
        result = '计划说明书池分布校验错误，详细错误请下载查看!'
        InsertVerificationLog(userId, 18, IsSucess, errtxtFilePath, result)

