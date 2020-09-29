# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import re
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import pandas
import pandas as pd
import openpyxl.styles as sty
logtxtFilePath = None
errtxtFilePath = None
def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def writeErr(msg):
    if not os.path.exists(errtxtFilePath):
        f = open(errtxtFilePath, "w")
    # print(msg)
    with open(errtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def is_number(s):
    try:
        float(s.replace(',', ''))
        return True
    except ValueError:
        return False

def markCellError(sheet, ctag, errtype):
    color = 'E93936' #datatype
    if errtype == 1:#empty
        color = 'AA2927'
    #sheet[ctag].fill=sty.PatternFill(fill_type='solid', fgColor=color)

#return isValid, isNA, errorType(1:empty, 2:datatype)
def checkCellValue(sheet, ctag, cvalue, cdtype):
    if ((cdtype == 'mumber' or cdtype == 'rate') and isinstance(cvalue, float)
       ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return 1, 0, 0

    # 去除单元格内容中的空格换行等
    cvalue = str(cvalue).replace(' ', '').replace(',', '').replace('.00', '').replace('\t', '').replace('\n', '')

    if cvalue == 'NA' or cvalue == '-' or cvalue == '':
        return 1, 1, 0

    # if cvalue == '':
    #     markCellError(sheet, ctag, 1)
    #     return 0, 0, 1

    dtcheck = 0
    if cdtype == 'int':
        dtcheck = 1 if cvalue.isdigit() else 0
    elif cdtype == 'number':
        dtcheck = 1 if is_number(cvalue) else 0
    elif cdtype == 'date':
        reg = r"(\d{4}[-/]\d{1,2}[^a-z]([-/]\d{1,2}[^a-z])?)|((\d{1,2}[^a-z][-/])?\d{1,2}[^a-z][-/]\d{4})|(\d{4}年\d{1,2}月(\d{1,2}日)?)|^(\d{5,6})$|^(\d{8})$"
        dtcheck = 1 if re.search(reg, cvalue) is not None else 0
    elif cdtype == 'rate':
        dtcheck = 1 if is_number(cvalue.rstrip('%')) else 0
    else:  # string
        dtcheck = 1

    if dtcheck == 0:
        markCellError(sheet, ctag, 2)
        return 0, 0, 2

    return 1, 0, 0

def getDate(value):
    model1 = r"(\d{4}年\d{1,2}月(\d{1,2}日)?)"
    model2 = r"(\d{4}[-/]\d{1,2}[^a-z]([-/]\d{1,2}[^a-z])?)|((\d{1,2}[^a-z][-/])?\d{1,2}[^a-z][-/]\d{4})"
    model3 = r"^(\d{5,6})$|^(\d{8})$"
    #value = '2016年3月31日'
    if re.search(model1, value) is not None:
        value = re.sub('\D$', '', value)
        value = re.sub(r'\D', r'-', value)
        value = pandas.to_datetime(value)
    elif re.search(model2, value) is not None:
        value = pandas.to_datetime(value)
    elif re.search(model3, value) is not None:
        value = pandas.Timedelta(str(value) + 'D')
        value = pandas.to_datetime('1899-12-30') + value
    print(value)
    return value.strftime('%Y-%m-%d')

def specificCellsExtract(cfgItem, sheet):
    checkmsg = []

    for cell in cfgItem:
        ctag = cell.tag
        cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        #crequired = cell.attrib['required'] if 'required' in cell.attrib else '0'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

        isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)

        if isValid != 1:
            etype = "不能为空" if errorType == 1 else "格式错误"
            checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, etype))

    return checkmsg

def specificColsExtract(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 1
    breakRowsLoop = 0
    while rStart <= rEnd:  # rows loop
        for cell in cNode: # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            #crequired = cell.attrib['required'] if 'required' in cell.attrib else '0'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break #break cells loop

            isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)
            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            if isValid != 1:
                etype = "不能为空" if errorType == 1 else "格式错误"
                checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, etype))

        if (isFirstRow == 1 and firstRowAllNA == 1) or breakRowsLoop == 1:
            rStart = rEnd + 1 #break rows loop
        else:
            isFirstRow = 0
            rStart += 1

    return checkmsg

def CompareDateExtract(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 0
    breakRowsLoop = 0
    previousValue = ''
    while rStart <= rEnd:  # rows loop
        for cell in cNode:  # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'

            cvalue = str(sheet[ctag].value) if sheet[ctag].value != None else ''
            if str(cvalue).replace(' ', '').replace('\t', '').replace('\n','') == 'NA':
                firstRowAllNA = 1
                break  # break cells loop
            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n',
                                                                                       '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break  # break cells loop
            if previousValue == '':
                previousValue = cvalue
            if getDate(previousValue) > getDate(cvalue):
                checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, "日期不合理，日期小余上期日期！"))
            else:
                previousValue = cvalue

        if (isFirstRow == 1 or breakRowsLoop == 1) and firstRowAllNA == 1:
            rStart = rEnd + 1  # break rows loop
        else:
            isFirstRow = 0
            rStart += 1

    return checkmsg

# 受托报告中池分布校验
def PoolDistributions(Excelfilepath, paymentPeriodID):
    #global TrustId, Error, errtxtFilePath
    ErrorMsg = []
    Error = 0

    data = pd.read_excel(Excelfilepath, sheet_name='Sheet2', header=1)

    PoolDistributions_columns = ['PaymentPeriodID', '资产池分布类型', 'DistributionType', 'DatabaseItem', 'BucketSequenceNo',
                                 'Bucket', 'Amount', 'AmountPercentage', 'Count', 'CountPercentage']

    for columnsP in PoolDistributions_columns:
        if columnsP not in data.columns:
            Error = 1
            #C9 = '表格字段错误不能含有中文或字段缺失!'
            ErrorMsg.append("{0}：[{1}] - {2}".format('A2-J2', '列名', "列名不能更改，请查看，补充完整"))
            Error = 1
            ##ErrorMessage(TrustId, FileTrustCode, TrustName, Excelfilepath, C9)
            return

    data.dropna(subset=['Amount', 'Bucket', 'AmountPercentage'], inplace=True)  #空值替换成NONE，对应列中的
    data = data[['PaymentPeriodID', 'DistributionType', 'DatabaseItem', 'BucketSequenceNo', 'Bucket', 'Amount',
                 'AmountPercentage', 'Count', 'CountPercentage']]
    DatabaseItem = data['DatabaseItem'].drop_duplicates() #列去重
    for o in DatabaseItem:
        AmountPercentageSum = 0
        CountPercentageSum = 0
        BucketSequenceNolist = []
        for i in data.index:
            DatabaseItem = data.loc[i][2]
            if DatabaseItem == o:
                try:
                    PaymentPeriodID = data.loc[i][0].astype(int)
                    BucketSequenceNo = data.loc[i][3].astype(float)
                    BucketSequenceNolist.append(BucketSequenceNo)
                    AmountPercentage = data.loc[i][6].astype(float)
                    CountPercentage = data.loc[i][8].astype(float)
                    Amount = data.loc[i][5].astype(float)
                    Count = data.loc[i][7].astype(float)

                    DistributionType = data.loc[i][1]
                    Bucket = data.loc[i][4]


                except:
                    ErrorMsg.append("E{0}G{0}H{0}I{0}J{0}：[{1}] - {2}".format(i+3, DatabaseItem, "数据类型错误应为数值类型,请修正！"))
                    return

                # 判断分布是否带单位

                if DistributionType == 'ApprovalAmount' and '元' not in Bucket and '万' not in str(Bucket):
                    ErrorMsg.append("F{0}：[{1}] - {2}".format(i + 3, DatabaseItem, "合同金额分布未带（元或万）单位!请检查!"))


                if DistributionType == 'CurrentPrincipalBalance' and '元' not in Bucket and '万' not in str(Bucket):
                    ErrorMsg.append("F{0}：[{1}] - {2}".format(i + 3, DatabaseItem, "剩余本金分布未带（元或万）单位!请检查!"))


                if DistributionType == 'LoanTerm' and '月' not in str(Bucket) and '天' not in str(Bucket) and '年' not in str(Bucket) and '期' not in str(Bucket):
                    ErrorMsg.append("F{0}：[{1}] - {2}".format(i + 3, DatabaseItem, "合同期限分布未带（天或月或年或期）单位!请检查!"))

                if DistributionType == 'Seasoning' and '月' not in str(Bucket) and '天' not in str(Bucket) and '年' not in str(Bucket):

                    ErrorMsg.append("F{0}：[{1}] - {2}".format(i + 3, DatabaseItem, "账龄分布未带（天或月或年）单位!请检查!"))

                if DistributionType == 'RemainingTerm' and '月' not in str(Bucket) and '天' not in str(Bucket) and '年' not in str(Bucket) and '期' not in str(Bucket):
                    ErrorMsg.append("F{0}：[{1}] - {2}".format(i + 3, DatabaseItem, "剩余期限分布未带（天或月或年或期）单位!请检查!"))

                if DistributionType == 'CustomerAge' and '岁' not in Bucket:
                    ErrorMsg.append("F{0}：[{1}] - {2}".format(i + 3, DatabaseItem, "年龄分布未带（岁）单位!请检查!"))

                if DistributionType == 'CustomerAnnualIncome' and '元' not in Bucket and '万' not in str(Bucket):
                    ErrorMsg.append("F{0}：[{1}] - {2}".format(i + 3, DatabaseItem, "收入分布未带（元或万）单位!请检查!"))

                if PaymentPeriodID != int(paymentPeriodID):
                    ErrorMsg.append("A{0}：[{1}] - PaymentPeriodID填写值：{2}错误,应为：{3}".format(i + 3, DatabaseItem, PaymentPeriodID, paymentPeriodID))



                AmountPercentageSum += AmountPercentage
                CountPercentageSum += CountPercentage
                BucketSequenceNoset = set(BucketSequenceNolist)
                if len(BucketSequenceNolist) != len(BucketSequenceNoset):
                    ErrorMsg.append("E{0}：[{1}] - BucketSequenceNo值重复".format(i + 3, DatabaseItem))
                    BucketSequenceNolist.clear()
        #         if Amount < Count:
        #             ErrorMsg.append("G{0}<I{0}：[{1}] - Amount【{2}】不应小于Count【{3}】,可能数据录反请检查!".format(i + 3, DatabaseItem, Amount, Count))
        #
        #
        # if AmountPercentageSum - 1 > 0.1 or AmountPercentageSum - 1 < -0.1:
        #     ErrorMsg.append("H列：[{0}] - AmountPercentage列相加【{1}】不等于1(忽略精度影响+-0.1)请检查!".format(o, AmountPercentageSum))
        #
        # if CountPercentageSum - 1 > 0.1 or CountPercentageSum - 1 < -0.1:
        #     ErrorMsg.append("J列：[{0}] - CountPercentage列相加【{1}】不等于1(忽略精度影响+-0.1)请检查!".format(o, CountPercentageSum))


    return ErrorMsg

def checkFileFormat(wb, cfgItems,sourceFilePath, paymentPeriodID):
    sheet = wb['Sheet1']
    ckResult = []
    ErrorMsg = []
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']
        itemtype = cfgItem.attrib['type'] if 'type' in cfgItem.attrib else ''

        itemCkResult = []
        if itemtype == 'SpecificCells':
            itemCkResult = specificCellsExtract(cfgItem, sheet)
        elif itemtype == 'CompareDate':
            itemCkResult = CompareDateExtract(cfgItem, sheet)
        else:
            itemCkResult = specificColsExtract(cfgItem, sheet)

        if len(itemCkResult) > 0:
            ckResult.append("【{0}】".format(itemdesc))
            ckResult.extend(itemCkResult)
    msg = []
    msg = PoolDistributions(sourceFilePath, paymentPeriodID)
    ErrorMsg.extend(msg)

    ckResultLen = len(ckResult)
    ErrorMsgLen = len(ErrorMsg)
    if ckResultLen > 0:
        logSheet = wb.create_sheet("格式检查")
        i = 1
        while i <= ckResultLen:
            logSheet["A{0}".format(i)] = ckResult[i - 1]
            i += 1

    if ErrorMsgLen > 0:
        if ckResultLen > 0:
            logSheet = wb["格式检查"]
        else:
            logSheet = wb.create_sheet("格式检查")
        i = ckResultLen + 2
        j = 1
        logSheet["A{0}".format(i + j)] = 'Sheet2表错误如下：'
        while j <= ErrorMsgLen:
            logSheet["A{0}".format(i + j + 1)] = ErrorMsg[j - 1]
            j += 1

    return (ckResultLen + ErrorMsgLen)


def main(configFilePath, dateId):
    global errtxtFilePath
    global logtxtFilePath
    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    scriptFolder = cfgRoot.attrib['sourcefolder']
    destFolder = cfgRoot.attrib['destfolder']
    scriptFolderPath = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolder, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolderPath, 'Logs',
                                  '{0}.txt'.format(dateId))
    errtxtFilePath = os.path.join(destFolder,
                                  'Error_第一步校验错误_{0}.txt'.format(dateId))

    for dirPath, dirNames, fileNames in os.walk(scriptFolder):
        for fileName in fileNames:
            if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告'):
                print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
                continue

            sourceFilePath = os.path.join(dirPath, fileName)

            filePath = "\n{0}".format(sourceFilePath)
            writeLog(filePath)

            fileNameAry = fileName.split(';')
            if len(fileNameAry) != 4:
                msg = "【文件名错误】文件名称命名不规范"
                writeLog(msg)
                writeErr(filePath)
                writeErr(msg)
                continue

            paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
            if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
                msg = "【错误】文件名中的TrustCode或报告期数设置有误"
                writeLog(msg)
                writeErr(filePath)
                writeErr(msg)
                continue

            excelwb = load_workbook(sourceFilePath)
            if '格式检查' in excelwb.sheetnames:
                excelwb.remove(excelwb['格式检查'])
                excelwb.save(sourceFilePath)

            hasError = checkFileFormat(excelwb, cfgRoot,sourceFilePath, paymentPeriodID)
            if hasError > 0:
                excelwb.save(sourceFilePath)
                writeLog('【有格式错误】详情见文档[格式检查]sheet')
                writeErr(filePath)
                writeErr('【有格式错误】详情见文档[格式检查]sheet')
