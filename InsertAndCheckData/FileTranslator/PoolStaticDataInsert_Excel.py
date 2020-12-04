# _*_ coding:utf-8 _*_

import sys
import os, re
import os.path
import datetime
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import pyodbc
import calendar
import pandas as pd


logtxtFilePath = None
errtxtFilePath = None
dbConnectionStr = None
AssetTypeDf = None

##### Helper Methods #####
def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        writeLog(str(ex))
        print(str(ex))
        # raise ex
    finally:
        cnxn.close()

def execSQLCmdFetchAll(sql):
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        rows = cursor.execute(sql).fetchall()
        return rows
    except Exception as ex:
        writeLog(str(ex))
        raise ex
    finally:
        cnxn.close()

def deleteOldData(fileName):
    sql = "delete from  PortfolioManagement.DvImport.StaticPoolData where FileNames = N'{0}'".format(fileName)
    execSQLCmd(sql)

def getNumber(value,Type):
    try:
        if Type == 'string':
            return value
        elif Type == 'date':
            model1 = r"(\d{4}年\d{1,2}月(\d{1,2}日)?)"
            model2 = r"(\d{4}[-/]\d{1,2}[^a-z]([-/]\d{1,2}[^a-z])?)|((\d{1,2}[^a-z][-/])?\d{1,2}[^a-z][-/]\d{4})"
            model3 = r"^(\d{5,6})$"
            if re.search(model1, str(value)) is not None:
                value = re.sub('\D$', '', value)
                value = re.sub(r'\D', r'-', value)
                value = pd.to_datetime(value)
            elif re.search(model3, str(value)) is not None:
                if re.search(model2, str(value)) is not None:
                    value = pd.to_datetime(value)
                else:
                    value = pd.to_datetime(value, format=('%Y%m'))
            elif value == '':
                return ''
            else:
                value = pd.to_datetime(value)
            year, month = str(value).split('-')[0], str(value).split('-')[1]  # 获取月末日期
            end = calendar.monthrange(int(year), int(month))[1]
            value = pd.to_datetime('%s-%s-%s' % (year, month, end))
            return value.strftime('%Y%m%d')
        cvalue = str(value).replace(' ', '').replace(',', '').replace('.00', '').replace('\t', '').replace('，', '')
        if cvalue == 'NA' or cvalue == '-' or cvalue == '':
            return 0
        if Type == 'int':
            return int(cvalue)
        elif Type == 'float':
            return float(cvalue)
    except Exception as ex:
        writeLog(str(ex))
        print(str(ex))
#获取数据库资产类型进行比较
def getAssetType():
    sql = "select ItemCode,ItemTitle  FROM [PortfolioManagement].[DV].[Item] where CategoryCode = 'AssetType'"
    msg = execSQLCmdFetchAll(sql)
    L = [(a, b) for a, b in msg]
    df = pd.DataFrame(L, columns=['ItemCode', 'ItemTitle'])
    return df

def concatSql(filePath, cfgRoot):
    writeLog('文件：{0}，开始读取'.format(filePath))
    fileName = os.path.basename(filePath)
    deleteOldData(fileName)  # 删除同名文件的旧数据
    excelwb = load_workbook(filePath)
    sheet = excelwb._sheets[0]
    addition = []
    addition.append(fileName)
    for cell in cfgRoot[0]: #静态池的几个额外信息，获取
        ctag = cell.tag
        cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''
        cvalue = getNumber(cvalue, cdtype)
        if cdesc == '发起机构':
            if '股份有限公司' not in cvalue and '有限责任公司' not in cvalue:
                writeLog('文件：{0}，发起机构填写错误，必须包含‘股份有限公司’或‘有限责任公司’！！！'.format(filePath))
                return
        if cdesc == '资产类型':
            if cvalue not in AssetTypeDf['ItemTitle'].values:
                writeLog('文件：{0}，资产类型填写错误，必须为数据库已存在资产类型！！！'.format(filePath))
                return
        addition.append(cvalue)

    statement = cfgRoot[1].attrib['stat']
    cNode = cfgRoot[1].find('cols')
    rNode = cfgRoot[1].find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    tmpl = ""
    RowNum = 0   #用来计数，当800行时，执行一次插入语句，因为插入语句必须少有1000条
    isEmpty = 0
    while rStart <= rEnd:  # rows loop
        rvalues = ''
        for info in addition:
            rvalues += "N'{0}',".format(info)
        for cell in cNode:  # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''
            crv = getNumber(cvalue, cdtype)

            if crv == '' and cemptybreak == '1':
                rvalues = ''
                isEmpty = 1
                break  # break cells loop
            rvalues += "N'{0}',".format(crv)

        if (isEmpty== 1):
            rStart = rEnd + 1  # break rows loop
        else:
            if rvalues != "":
                tmpl += "({0}),".format(rvalues.rstrip(','))
            rStart += 1
            RowNum += 1
        if RowNum >= 800:
            RowNum = 0
            execSQLCmd("{0}{1}".format(statement, tmpl.rstrip(',')))
            tmpl = ""
    print("{0}{1}".format(statement, tmpl.rstrip(',')))
    if tmpl != '':
        execSQLCmd("{0}{1}".format(statement, tmpl.rstrip(',')))

    writeLog('文件：{0}，读取入库完成'.format(filePath))

def main(configFilePath, dateId):
    global logtxtFilePath
    global errtxtFilePath
    global dbConnectionStr
    global trustID
    global paymentPeriodID
    global AssetTypeDf

    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    audioType = int(cfgRoot.attrib['audioType'])
    DirPath = cfgRoot.attrib['DirPath']
    filePath = cfgRoot.attrib['filePath']
    dbConnectionStr = 'DRIVER={SQL Server};SERVER=172.16.6.143\MSSQL;DATABASE=PortfolioManagement;UID=sa;PWD=PasswordGS2017'
    #日志记录文件
    scriptFolderPath = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolderPath, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)
    logtxtFilePath = os.path.join(scriptFolderPath, 'Logs', '{0}.txt'.format(dateId))
    AssetTypeDf = getAssetType()

    if audioType == 1:  #当是文件夹时
        for dirPath, dirNames, fileNames in os.walk(DirPath):
            for fileName in fileNames:
                if not fileName.endswith('.xlsx') and not fileName.endswith('.xls'):
                    print("已跳过文件{0}".format(fileName))
                    continue
                sourceFilePath = os.path.join(dirPath, fileName)
                concatSql(sourceFilePath, cfgRoot)
    else:       #当选择文件时
        concatSql(filePath, cfgRoot)





