# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import pyodbc
import pandas as pd
import math

logtxtFilePath = None
errtxtFilePath = None
dbConnectionStr = None
sourceFilePath = None
trustID = 0
paymentPeriodID = 0
##### Helper Methods #####
def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def writeErr(msg):
    if not os.path.exists(errtxtFilePath):
        f = open(errtxtFilePath, "w")
    # print(msg)
    with open(errtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def execSQLCmd(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        cursor.execute(sql)
        cnxn.commit()
    except Exception as ex:
        writeLog(str(ex))
        writeErr("\n【{0}】".format(sourceFilePath))
        writeErr(str(ex))
        print(str(ex))
        # raise ex
    finally:
        cnxn.close()

def execSQLCmdFetchOne(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        row = cursor.execute(sql).fetchone()
        return row
    except Exception as ex:
        writeLog(str(ex))
        writeErr("\n【{0}】".format(sourceFilePath))
        writeErr(str(ex))
        raise ex
    finally:
        cnxn.close()

def execSQLCmdFetchAll(sql):
    # print(sql)
    cnxn = pyodbc.connect(dbConnectionStr)
    try:
        cursor = cnxn.cursor()
        rows = cursor.execute(sql).fetchall()
        return rows
    except Exception as ex:
        writeLog(str(ex))
        writeErr("\n【{0}】".format(sourceFilePath))
        writeErr(str(ex))
    finally:
        cnxn.close()

def getTrustID(trustCode, assetType):
    sql = "select TrustId from DV.view_Products where TrustCode = N'{0}' and AssetType = N'{1}'".format(trustCode, assetType)
    try:
        tid = execSQLCmdFetchOne(sql).TrustId
        return tid
    except Exception as ex:
        writeErr("\n【{0}】".format(sourceFilePath))
        writeErr(str(ex))
        return 0

def cleanPoolDistribution(TrustId, PaymentPeriod):
    sql = "delete from dbo.PoolDistributions1 where TrustId = N'{0}' and PaymentPeriodID = N'{1}' ".format(TrustId, PaymentPeriod)
    try:
        execSQLCmd(sql)
    except Exception as ex:
        writeErr(str(ex))
        return 0

def cleanOldData():
    sql = "exec DVImport.ClearLastTimeImport_other {0}, {1}".format(trustID, paymentPeriodID)
    execSQLCmd(sql)

def runDBDataValidation():
    sql = "exec DVImport.CheckNonDataValidation {0}, {1}".format(trustID, paymentPeriodID)
    msg = execSQLCmdFetchAll(sql)
    return msg

##### Extract Implemention Methods #####
#return: cellValue, isNA
def getCellValue(cvalue, cdtype):
    if ((cdtype == 'mumber' or cdtype == 'rate') and isinstance(cvalue, float)
       ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return str(cvalue), 0

    cvalue = str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '')
    if cvalue == 'NA' or cvalue == '-':
        return cvalue, 1

    return cvalue, 0

def specificCellsExtract(cfgItem, sheet):
    statement = cfgItem.attrib['stat']
    itemdesc = cfgItem.attrib['desc']
    tmpl = "({0},{1}".format(trustID, paymentPeriodID)

    isAllNA = 1
    for cell in cfgItem:
        ctag = cell.tag
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

        crv, isna = getCellValue(cvalue, cdtype)
        if isna == 0:
            isAllNA = 0
        tmpl += ",N'{0}'".format(crv)

    if isAllNA == 1:
        writeLog("【无数据提示】 [{0}] 数据全部NA".format(itemdesc))

    return "{0}{1})".format(statement, tmpl)

def specificColsExtract(cfgItem, sheet):
    statement = cfgItem.attrib['stat']
    itemdesc = cfgItem.attrib['desc']
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    tmpl = ""
    isFirstRow = 1
    firstRowAllNA = 1
    # breakRowsLoop = 0
    while rStart <= rEnd:  # rows loop
        rvalues = ''

        for cell in cNode:  # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            crv, isNA = getCellValue(cvalue, cdtype)

            if isFirstRow != 1 and crv == '' and cemptybreak == '1':
                rvalues = ''
                break  # break cells loop

            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            rvalues += ",N'{0}'".format(crv)

        if (isFirstRow == 1 and firstRowAllNA == 1):
            rStart = rEnd + 1  # break rows loop
        else:
            isFirstRow = 0
            if rvalues != "":
                tmpl += "({0},{1}{2}),".format(trustID, paymentPeriodID, rvalues)
            rStart += 1

    if tmpl == "":
        writeLog("【无数据提示】 [{0}] 数据全部NA".format(itemdesc))
        return ''

    return "{0}{1}".format(statement, tmpl.rstrip(','))

def PoolDistributionsImport(Excelfilepath, TrustId, paymentPeriodID):
    cleanPoolDistribution(TrustId, paymentPeriodID)
    statement = "insert into dbo.PoolDistributions1(TrustId,PaymentPeriodID,DistributionTypeCode,BucketSequenceNo,Bucket" \
             ",Amount,AmountPercentage,Count,CountPercentage,DatabaseItem) values "
    sqlStr = ""
    DataPoolDistributions = pd.read_excel(Excelfilepath, sheet_name='Sheet2', header=1)
    DataPoolDistributionsImport = DataPoolDistributions[
        ['PaymentPeriodID', 'DistributionType', 'BucketSequenceNo', 'Bucket', 'Amount', 'AmountPercentage', 'Count',
         'CountPercentage', 'DatabaseItem']]
    DataPoolDistributionsImport = DataPoolDistributionsImport.dropna(subset=['Bucket', 'Amount'])
    for Cindex in DataPoolDistributionsImport.index:
        PaymentPeriodID = DataPoolDistributionsImport.loc[Cindex][0]
        DistributionTypeCode = DataPoolDistributionsImport.loc[Cindex][1]
        BucketSequenceNo = DataPoolDistributionsImport.loc[Cindex][2]
        Bucket = DataPoolDistributionsImport.loc[Cindex][3]
        Amount = 'null' if math.isnan(DataPoolDistributionsImport.loc[Cindex][4]) else \
        DataPoolDistributionsImport.loc[Cindex][4]
        AmountPercentage = 'null' if math.isnan(DataPoolDistributionsImport.loc[Cindex][5]) else \
        DataPoolDistributionsImport.loc[Cindex][5]
        Count = 'null' if math.isnan(DataPoolDistributionsImport.loc[Cindex][6]) else \
        DataPoolDistributionsImport.loc[Cindex][6]
        CountPercentage = 'null' if math.isnan(DataPoolDistributionsImport.loc[Cindex][7]) else \
        DataPoolDistributionsImport.loc[Cindex][7]
        DatabaseItem = DataPoolDistributionsImport.loc[Cindex][8]

        if AmountPercentage != 'null' and AmountPercentage <= 1:
            AmountPercentage = AmountPercentage * 100

        if CountPercentage != 'null' and CountPercentage <= 1:
            CountPercentage = CountPercentage * 100

        sqlStr += "({0},{1},'{2}',{3},N'{4}',{5},{6},{7},{8},N'{9}'),".format(TrustId, PaymentPeriodID,
            DistributionTypeCode, BucketSequenceNo, Bucket, Amount, AmountPercentage, Count, CountPercentage,DatabaseItem)

    if sqlStr == "":
        writeLog("【无数据提示】 [Sheet2池分布数据] 数据全部NA")
    else:
        writeLog("【Shhet2已提取】开始提交 [sheet2池分布数据] 至数据库")
        sqlStr = "{0}{1}".format(statement, sqlStr.rstrip(','))
        execSQLCmd(sqlStr)

    #return "{0}{1}".format(statement, sqlStr.rstrip(','))

def extractExcel(wb, cfgItems):
    sheet = wb['Sheet1']

    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']
        itemtype = cfgItem.attrib['type'] if 'type' in cfgItem.attrib else ''

        if not 'stat' in cfgItem.attrib:
            msg = '【程序配置错误】config.xml中[{0}]节点中未配置statement，节点无法操作!'.format(itemdesc)
            writeLog(msg)
            continue

        exesql = ""
        if itemtype == 'SpecificCells':
            exesql = specificCellsExtract(cfgItem, sheet)
        else: #'SpecificCols':
            exesql = specificColsExtract(cfgItem, sheet)

        if exesql != '':
            writeLog("【{0}-已提取】开始提交 [{1}] 至数据库".format(i, itemdesc))
            execSQLCmd(exesql)
        else:
            writeLog("【{0}-未提取】未获取不提交 [{1}] ".format(i, itemdesc))

def writeLogToTable(trustID, paymentPeriodID, sourceFilePath):
    dt = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sql = "insert into DVImport.ToolImportLog values({0},{1},N'{2}','{3}')".format(trustID, paymentPeriodID, sourceFilePath, dt)
    execSQLCmd(sql)

def main(configFilePath, dateId):
    global logtxtFilePath
    global errtxtFilePath
    global dbConnectionStr
    global sourceFilePath
    global trustID
    global paymentPeriodID

    mappingTree = XETree.parse(configFilePath)
    cfgRoot = mappingTree.getroot()
    sourceFolder = cfgRoot.attrib['sourcefolder']
    assetType = cfgRoot.attrib['AssetType']
    destFolder =cfgRoot.attrib['destfolder']
    dbConnectionStr = cfgRoot.attrib['dbconnstr']

    scriptFolderPath = os.path.dirname(os.path.abspath(__file__))
    log_Path = os.path.join(scriptFolderPath, "Logs")
    if not os.path.exists(log_Path):
        os.mkdir(log_Path)

    logtxtFilePath = os.path.join(scriptFolderPath, 'Logs',
                                  '{0}.txt'.format(dateId))
    errtxtFilePath = os.path.join(destFolder,
                                  'Error_第二步校验错误_{0}.txt'.format(dateId))
    trustID = 0
    paymentPeriodID = 0
    sourceFilePath = ''
    ################### Script Main ###################
    for dirPath, dirNames, fileNames in os.walk(sourceFolder):
        for fileName in fileNames:
            if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告'):
                print("已跳过文件{0}".format(fileName))
                continue

            sourceFilePath = os.path.join(dirPath, fileName)
            writeLog("\n【{0}】".format(sourceFilePath))

            fileNameAry = fileName.split(';')
            if len(fileNameAry) != 4:
                writeErr("\n【{0}】".format(sourceFilePath))
                writeErr("【错误】文件名称命名不规范")
                writeLog("【错误】文件名称命名不规范")
                continue

            # initialize key pamerters
            trustID = 0
            paymentPeriodID = 0
            trustID = getTrustID(fileNameAry[1], assetType)
            paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
            if trustID == '' or trustID == 0 or not paymentPeriodID.isdigit() or paymentPeriodID == 0:
                writeErr("\n【{0}】".format(sourceFilePath))
                writeLog("【错误】系统中未能定位到产品，文件名称中TrustCode或报告期数设置有误")
                writeErr("【错误】系统中未能定位到产品，文件名称中TrustCode或报告期数设置有误")
                continue

            excelwb = load_workbook(sourceFilePath)
            if '格式检查' in excelwb.sheetnames:
                writeErr("\n【{0}】".format(sourceFilePath))
                writeLog("【错误】第一步校验中的错误尚未处理并重运行第一步校验检查")
                writeErr("【错误】第一步校验中的错误尚未处理并重运行第一步校验检查")
                continue

            if '正确性验证' in excelwb.sheetnames:
                excelwb.remove(excelwb['正确性验证'])
                excelwb.save(sourceFilePath)


            cleanOldData()
            extractExcel(excelwb, cfgRoot)   #sheet1 内容上传
            PoolDistributionsImport(sourceFilePath, trustID, paymentPeriodID)  # sheet2内容上传
            writeLogToTable(trustID, paymentPeriodID, sourceFilePath)
            dbCheckResult = runDBDataValidation()

            if len(dbCheckResult) > 0:
                writeErr("\n【{0}】".format(sourceFilePath))
                writeErr("【数据准确性校验未通过】错误信息见文件[正确性验证]sheet")
                writeLog("【数据准确性校验未通过】错误信息见文件[正确性验证]sheet")
                logSheet = excelwb.create_sheet("正确性验证")
                logSheet["A{0}".format(1)] = "注：以下为校验有误的数据，仅作参考"
                i = 2
                for r in dbCheckResult:
                    logSheet["A{0}".format(i)] = r.Result
                    logSheet["B{0}".format(i)] = r.Message
                    i += 1

                excelwb.save(sourceFilePath)


