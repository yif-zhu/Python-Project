# _*_ coding:utf-8 _*_

import sys
import os
import os.path
import datetime
import re
import xml.etree.ElementTree as XETree
from openpyxl import load_workbook
import openpyxl.styles as sty

class cObj(object):
    def __init__(self, desc, val):
        self.desc = desc
        self.val = val

def writeLog(msg):
    if not os.path.exists(logtxtFilePath):
        f = open(logtxtFilePath, "w")
    print(msg)
    with open(logtxtFilePath, "a") as f:
        ts = datetime.datetime.now().strftime('[%H:%M:%S]')
        f.write('{0}:  {1}\n'.format(ts, msg))

def is_number(s):
    try:
        float(s.replace(',', ''))
        return True
    except ValueError:
        return False

def getNumberVal(val):
    if isinstance(val, int) or isinstance(val, float):
        return val
    return None if val == 'NA' or val =='-' or val == '' else float(val.replace('\t', '').replace('\n', '').replace(',', '').replace('%', '').replace(' ', ''))

def markCellError(sheet, ctag, errtype):
    color = 'E93936' #datatype
    if errtype == 1:#empty
        color = 'AA2927'
    sheet[ctag].fill=sty.PatternFill(fill_type='solid', fgColor=color)

def checkCellValue(sheet, ctag, cvalue, cdtype):
    if ((cdtype == 'mumber' or cdtype == 'rate') and isinstance(cvalue, float)
        ) or (cdtype == 'int' and isinstance(cvalue, int)):
        return 1, 0, 0

    cvalue = str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '')

    if cvalue == 'NA' or cvalue == '-':
        return 1, 1, 0

    if cvalue == '':
        markCellError(sheet, ctag, 1)
        return 0, 0, 1

    dtcheck = 0
    if cdtype == 'int':
        dtcheck = 1 if cvalue.isdigit() else 0
    elif cdtype == 'number':
        dtcheck = 1 if is_number(cvalue) else 0
    elif cdtype == 'date':
        reg = r"(\d{4}[-/]\d{1,2}([-/]\d{1,2})?)|((\d{1,2}[-/])?\d{1,2}[-/]\d{4})|(\d{4}年\d{1,2}月(\d{1,2}日)?)|(\d{6})|(\d{8})"
        dtcheck = 1 if re.search(reg, cvalue) is not None else 0
    elif cdtype == 'rate':
        dtcheck = 1 if is_number(cvalue.rstrip('%')) else 0
    else:  # string
        dtcheck = 1

    if dtcheck == 0:
        markCellError(sheet, ctag, 2)
        return 0, 0, 2

    return 1, 0, 0

def specificCellsCheck(cfgItem, sheet):
    checkmsg = []
    r = []
    for cell in cfgItem:
        ctag = cell.tag
        cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
        cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
        cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

        isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)

        if isValid != 1:
            etype = "不能为空" if errorType == 1 else "格式错误"
            checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, etype))

        r.append(cObj(cdesc, cvalue))

    return checkmsg, r

def specificColsCheck(cfgItem, sheet):
    checkmsg = []
    cNode = cfgItem.find('cols')
    rNode = cfgItem.find('rows')

    rStart = int(rNode.attrib['start'])
    rEnd = int(rNode.attrib['end'])
    if 'useacturalend' in rNode.attrib and rNode.attrib['useacturalend'] == '1':
        rEnd = sheet.max_row if sheet.max_row > rEnd else rEnd

    isFirstRow = 1
    firstRowAllNA = 1
    breakRowsLoop = 0
    rs = []
    while rStart <= rEnd:  # rows loop
        r = []
        for cell in cNode: # cells loop
            ctag = "{0}{1}".format(cell.tag, rStart)
            cdesc = cell.attrib['desc'] if 'desc' in cell.attrib else ''
            cdtype = cell.attrib['dtype'] if 'dtype' in cell.attrib else 'string'
            cnagroup = cell.attrib['nagroup'] if 'nagroup' in cell.attrib else '0'
            cemptybreak = cell.attrib['emptybreak'] if 'emptybreak' in cell.attrib else '0'
            cvalue = sheet[ctag].value if sheet[ctag].value != None else ''

            if isFirstRow != 1 and str(cvalue).replace(' ', '').replace('\t', '').replace('\n', '') == '' and cemptybreak == '1':
                breakRowsLoop = 1
                break #break cells loop

            isValid, isNA, errorType = checkCellValue(sheet, ctag, cvalue, cdtype)
            if isFirstRow == 1 and cnagroup == '1' and isNA == 0:
                firstRowAllNA = 0

            if isValid != 1:
                etype = "不能为空" if errorType == 1 else "格式错误"
                checkmsg.append("{0}：[{1}] - {2}".format(ctag, cdesc, etype))

            r.append(cObj(cdesc, cvalue))

        if (isFirstRow == 1 and firstRowAllNA == 1) or breakRowsLoop == 1:
            rStart = rEnd + 1 #break rows loop
        else:
            isFirstRow = 0
            rStart += 1
            rs.append(r)

    return checkmsg, rs

def writeSheetLog(wb, res):
    logSheet = wb.create_sheet("格式检查")
    i = 1
    while i <= len(res):
        logSheet["A{0}".format(i)] = res[i - 1]
        i += 1

def getValidationResult(ds):
    ckResult = []

    bondsAmt = 0
    for d in ds[0]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestPaied = getNumberVal(d[3].val)
        pendAmt = getNumberVal(d[4].val)

        if pInterestPaied < 100000:
            ckResult.append("【收益分配记录】债券利息还款 [{0}] 小于10万,数据范围错误".format(pInterestPaied))
        if pendAmt < 100000:
            ckResult.append("【收益分配记录】债券期末金额 [{0}] 小于10万,数据范围错误".format(pendAmt))

        if pendAmt is not None and pstartAmt is not None and pendAmt > pstartAmt:
            ckResult.append("【收益分配记录】债券期末余额 [{0}] 大于期初余额{1},数据对应关系错误".format(pendAmt, pstartAmt))
        if pPrincipalAmt is not None and pstartAmt is not None and pPrincipalAmt > pstartAmt:
            ckResult.append("【收益分配记录】债券本金金额 [{0}] 大于期初余额{1}，数据对应关系错误".format(pPrincipalAmt, pstartAmt))

        if pstartAmt is not None and pPrincipalAmt is not None and pendAmt is not None and pstartAmt - pPrincipalAmt != pendAmt:
            ckResult.append("【收益分配记录】期初余额 [{0}] 减去本金金额 [{1}] 不等于期末余额 [{2}] 错误".format(pstartAmt, pPrincipalAmt, pendAmt))
        if pendAmt is not None:
            bondsAmt += pendAmt

    if bondsAmt is not None and bondsAmt != 0:
        assetAmt = getNumberVal(ds[5][2].val)
        if assetAmt is not None and assetAmt != 0 and assetAmt < 100000:
            ckResult.append("【笔数与金额特征】资产池总金额 [{0}] 小于10万,数据范围错误".format(assetAmt))
        if assetAmt is not None and assetAmt != 0 and abs(assetAmt - bondsAmt) / assetAmt > 0.05:
            ckResult.append("【收益分配记录 与 笔数与金额特征】各债券期初余额之和 [{0}] 同资产池统计的总金额 [{1}] 相差在5%之上错误".format(bondsAmt, assetAmt))

    t2 = 0
    t4 = 0
    for d in ds[1]:
        loanCount = getNumberVal(d[2].val)
        amtCount = getNumberVal(d[4].val)
        if loanCount is not None:
            t2 += loanCount
        if amtCount is not None:
            t4 += amtCount
    if t2 != 0 and (t2 != 100 or t2 != 1):
        ckResult.append("【资产池整体表现情况】笔数占比之和不等于100%")
    if t4 != 0 and (t4 != 100 or t4 != 1):
        ckResult.append("【资产池整体表现情况】金额占比之和不等于100%")

    pstartLast = None
    pPrincipalLast = None
    pInterestlast = None
    for d in ds[2]:
        pstartAmt = getNumberVal(d[1].val)
        pPrincipalAmt = getNumberVal(d[2].val)
        pInterestAmt = getNumberVal(d[3].val)

        if pPrincipalAmt < 100000:
            ckResult.append("【现金流归集表】应收本金金额 [{0}] 小于10万,数据范围错误".format(pPrincipalAmt))
        if pInterestAmt < 100000:
            ckResult.append("【现金流归集表】应收利息金额 [{0}] 小于10万,数据范围错误".format(pInterestAmt))

        if pstartAmt is not None and pstartAmt is not None and pstartAmt < pPrincipalAmt:
            ckResult.append("【现金流归集表】当期期初本金金额 [{0}] 小于应收本金金额 [{1}] 数据对应关系错误".format(pstartAmt, pPrincipalAmt))

        '''
        if pstartLast is not None and pstartAmt is not None and pstartLast < pstartAmt:
            ckResult.append("【现金流归集表】当期期初本金金额 [{0}] 大于上期期初本金金额 [{1}] 错误".format(pstartAmt, pstartLast))        
        if pPrincipalLast is not None and pPrincipalAmt is not None and pPrincipalLast < pPrincipalAmt:
            ckResult.append("【现金流归集表】当期应收本金金额 [{0}] 大于上期应收本金金额 [{1}] 错误".format(pPrincipalAmt, pPrincipalLast))
        if pInterestlast is not None and pInterestAmt is not None and pInterestlast < pInterestAmt:
            ckResult.append("【现金流归集表】当期应收利息金额 [{0}] 大于上期应收利息金额 [{1}] 错误".format(pInterestAmt, pInterestlast))
        '''
        pstartLast = pstartAmt
        pPrincipalLast = pPrincipalAmt
        pInterestlast = pInterestAmt

    assetPoolCF = ds[3]
    if len(assetPoolCF) > 0:
        val = getNumberVal(assetPoolCF[0].val)
        if val is not None and val != 0 and val < 100000:#利息-正常回收
            ckResult.append("【资产池现金流详情】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolCF[0].desc, val))
        val = getNumberVal(assetPoolCF[8].val)
        if val is not None and val != 0 and val < 100000:#本金-正常回收
            ckResult.append("【资产池现金流详情】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolCF[8].desc, val))

    assetPoolST = ds[4]
    if len(assetPoolST) > 0:
        val = getNumberVal(assetPoolST[0].val)
        if val is not None and val != 0 and val < 100000:  # 收入合计
            ckResult.append("【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[0].desc, val))
        val = getNumberVal(assetPoolST[1].val)
        if val is not None and val != 0 and val < 100000:  # 本金合计
            ckResult.append("【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[1].desc, val))
        val = getNumberVal(assetPoolST[2].val)
        if val is not None and val != 0 and val < 10000:  # 税费支出-税收
            ckResult.append("【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[2].desc, val))
        val = getNumberVal(assetPoolST[3].val)
        if val is not None and val != 0 and val < 10000:  # 税费支出-资产服务报酬
            ckResult.append("【资产池情况】{0} [{1}] 数值范围错误，请确认是否进行正确的单位换算".format(assetPoolST[3].desc, val))

    maxPAmt = getNumberVal(ds[5][3].val)
    avgPAmt = getNumberVal(ds[5][4].val)
    if maxPAmt is not None and avgPAmt is not None and maxPAmt < avgPAmt:
        ckResult.append("【笔数与金额特征】最高本金余额 [{0}] 小于平均本金余额 [{1}] 数据对应关系错误".format(maxPAmt, avgPAmt))

    maxRT = getNumberVal(ds[5][8].val)
    minRT = getNumberVal(ds[5][9].val)
    avgRT = getNumberVal(ds[5][7].val)
    if maxRT is not None and minRT is not None and maxRT < minRT:
        ckResult.append("【笔数与金额特征】最长剩余期限 [{0}] 小于最短剩余期限 [{1}] 数据对应关系错误".format(maxRT, minRT))
    if maxRT is not None and minRT is not None and avgRT is not None and (maxRT < avgRT or minRT > avgRT):
        ckResult.append("【笔数与金额特征】平均剩余期限 [{2}] 未介于最长剩余期限 [{0}] 小于最短剩余期限 [{1}] 之间,数据对应关系错误".format(maxRT, minRT, avgRT))

    maxRT = getNumberVal(ds[5][11].val)
    minRT = getNumberVal(ds[5][12].val)
    avgRT = getNumberVal(ds[5][10].val)
    if maxRT is not None and minRT is not None and maxRT < minRT:
        ckResult.append("【笔数于金额特征】最高利率 [{0}] 小于最低利率 [{1}] 数据对应关系错误".format(maxRT, minRT))
    if maxRT is not None and minRT is not None and avgRT is not None and (maxRT < avgRT or minRT > avgRT):
        ckResult.append("【笔数于金额特征】平均利率 [{2}] 未介于最高利率 [{0}] 和最低利率 [{1}] 之间,数据对应关系错误".format(maxRT, minRT, avgRT))

    #累计违约率有可能会降低，所以暂时不验证是否逐期递增，只验证是否每期的数字都在0.1以内
    for d in ds[7]:
        cdr = getNumberVal(d[1].val)
        if cdr is not None and cdr != 0:
            if (str(d[1].val).endswith('%') and cdr > 10) or (not str(d[1].val).endswith('%') and cdr > 0.1):
                ckResult.append("【累计违约率】违约率数值 [{0}] 大于10%，取值范围过大错误".format(cdr))

    return ckResult

def checkFileFormat(wb, cfgItems):
    sheet = wb['Sheet1']
    ckResult = []
    tp = [[], [], [], [], [], [], [], []]
    for i in range(len(cfgItems)):
        cfgItem = cfgItems[i]
        itemdesc = cfgItem.attrib['desc']
        itemtype = cfgItem.attrib['type'] if 'type' in cfgItem.attrib else ''

        itemCkResult = []
        r = []
        if itemtype == 'SpecificCells':
            itemCkResult, r = specificCellsCheck(cfgItem, sheet)
        else:
            itemCkResult, r = specificColsCheck(cfgItem, sheet)

        tp[i] = r
        if len(itemCkResult) > 0:
            ckResult.append("【{0}】".format(itemdesc))
            ckResult.extend(itemCkResult)

    if len(ckResult) == 0:
        ckResult = getValidationResult(tp)
    if len(ckResult) > 0:
        writeSheetLog(wb, ckResult)

    return len(ckResult)

scriptFolder = sys.argv[0].replace(sys.argv[0].split('\\')[-1], '')
logtxtFilePath = os.path.join(scriptFolder, 'Logs', 'Log_FormatCheck_{0}.txt'.format(datetime.datetime.now().strftime('%m-%d %H%M%S')))
configFilePath = os.path.join(scriptFolder, 'FileFormatCheck_Trustee.xml')
mappingTree = XETree.parse(configFilePath)
cfgRoot = mappingTree.getroot()
sourceFolderPath = cfgRoot.attrib['sourcefolder']
for dirPath, dirNames, fileNames in os.walk(sourceFolderPath):
    for fileName in fileNames:
        if not fileName.endswith('.xlsx') or not fileName.startswith('00受托报告数据提取;'):
            print("【跳过】文件名称不符合，已跳过文件{0}".format(fileName))
            continue

        sourceFilePath = os.path.join(dirPath, fileName)

        msg = "\n{0}".format(sourceFilePath)
        writeLog(msg)

        fileNameAry = fileName.split(';')
        if len(fileNameAry) != 4:
            msg = "【文件名错误】文件名称命名不规范"
            writeLog(msg)
            continue

        paymentPeriodID = 0

        paymentPeriodID = fileNameAry[3].rstrip('.xlsx')
        if not paymentPeriodID.isdigit() or paymentPeriodID == 0:
            msg = "【错误】文件名中的TrustCode或报告期数设置有误"
            writeLog(msg)
            continue

        excelwb = load_workbook(sourceFilePath)
        if '格式检查' in excelwb.sheetnames:
            excelwb.remove(excelwb['格式检查'])
            excelwb.save(sourceFilePath)

        hasError = checkFileFormat(excelwb, cfgRoot)
        if hasError > 0:
            excelwb.save(sourceFilePath)
            writeLog('【有格式错误】详情见文档[格式检查]sheet')

